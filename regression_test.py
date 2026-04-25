# -*- coding: utf-8 -*-
"""
最小样本回归测试脚本。

默认行为：
1. 不重复执行完整主链；
2. 自动定位 `_smoke_output` 下最近一次成功产物目录；
3. 对关键输出做结构化断言，尽快发现回归。

如需刷新产物，可加 `--run-smoke` 先执行一次 `smoke_test.py`。
"""

from __future__ import annotations

import argparse
import subprocess
import sys
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
DEFAULT_OUTPUT_DIR = ROOT / "_smoke_output"

MERGED_FILENAME = "4_合并表.xlsx"
SOURCE_FILENAME = "5_中间汇总表_效益审核数据源.xlsx"
AUTO_FILL_FILENAME = "自动填报完成_效益审核表.xlsx"
FINAL_FILENAME = "最终完美交付版_效益审核表.xlsx"


def _norm(value) -> str:
    if value is None:
        return ""
    return (
        str(value)
        .replace(" ", "")
        .replace("\u3000", "")
        .replace("（", "(")
        .replace("）", ")")
        .replace("：", ":")
        .replace("\n", "")
        .replace("\r", "")
    )


def _assert(condition: bool, message: str) -> None:
    if not condition:
        raise AssertionError(message)


def _print_ok(message: str) -> None:
    print(f"[PASS] {message}")


def _print_info(message: str) -> None:
    print(f"[INFO] {message}")


def _open_sheet(path: Path, *, read_only: bool = True):
    workbook = load_workbook(path, read_only=read_only, data_only=False)
    return workbook, workbook[workbook.sheetnames[0]]


def _row_values(ws, row_idx: int) -> list:
    return [cell.value for cell in next(ws.iter_rows(min_row=row_idx, max_row=row_idx))]


def _find_latest_result_dir(output_dir: Path) -> Path:
    candidates = []
    for child in output_dir.iterdir():
        if child.is_dir() and (child / MERGED_FILENAME).exists():
            candidates.append(child)

    if not candidates:
        raise FileNotFoundError(f"未在 {output_dir} 下找到包含 {MERGED_FILENAME} 的结果目录")

    return sorted(candidates, key=lambda path: path.stat().st_mtime, reverse=True)[0]


def _find_row_by_keyword(ws, keyword: str, cols=(2, 3, 4, 5)) -> int | None:
    expected = _norm(keyword)
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in cols:
            if expected in _norm(ws.cell(row_idx, col_idx).value):
                return row_idx
    return None


def _row_has_formula(ws, row_idx: int, start_col: int = 8) -> bool:
    for col_idx in range(start_col, ws.max_column + 1):
        value = ws.cell(row_idx, col_idx).value
        if isinstance(value, str) and value.startswith("="):
            return True
    return False


def run_smoke(output_dir: Path) -> None:
    command = [
        sys.executable,
        str(ROOT / "smoke_test.py"),
        "--output-dir",
        str(output_dir),
    ]
    _print_info("执行冒烟测试以刷新最新产物...")
    completed = subprocess.run(command, cwd=str(ROOT), check=False)
    if completed.returncode != 0:
        raise RuntimeError("smoke_test.py 执行失败，无法继续回归校验")


def validate_files(result_dir: Path) -> None:
    for filename in (MERGED_FILENAME, SOURCE_FILENAME, AUTO_FILL_FILENAME, FINAL_FILENAME):
        _assert((result_dir / filename).exists(), f"缺少关键输出文件：{result_dir / filename}")
    _print_ok(f"关键输出文件齐全：{result_dir.name}")


def validate_merged_table(result_dir: Path) -> None:
    workbook, ws = _open_sheet(result_dir / MERGED_FILENAME, read_only=True)
    try:
        header_row = _row_values(ws, 1)
        subtotal_row = _row_values(ws, 2)
        first_data_row = _row_values(ws, 3)

        required_headers = {"公司代码", "总账科目长文本", "借方本位币金额", "贷方本位币金额"}
        _assert(required_headers.issubset(set(header_row)), "4_合并表.xlsx 缺少关键表头")
        _assert(subtotal_row[0] == "筛选合计：", "4_合并表.xlsx 第2行未生成筛选合计标识")

        subtotal_formulas = [
            value for value in subtotal_row if isinstance(value, str) and value.upper().startswith("=SUBTOTAL(")
        ]
        _assert(len(subtotal_formulas) >= 5, "4_合并表.xlsx 第2行未生成足够的 SUBTOTAL 公式")
        _assert(ws.max_row >= 10, "4_合并表.xlsx 数据行数异常偏少")
        _assert(any(value not in (None, "") for value in first_data_row), "4_合并表.xlsx 首行数据为空")
        _print_ok(f"4_合并表结构正常，行数 {ws.max_row}")
    finally:
        workbook.close()


def validate_source_table(result_dir: Path) -> None:
    workbook, ws = _open_sheet(result_dir / SOURCE_FILENAME, read_only=True)
    try:
        header_row = _row_values(ws, 1)
        subtotal_row = _row_values(ws, 2)
        required_headers = {"利润中心", "工程名称", "成本_财务大类", "最终发生额"}
        _assert(required_headers.issubset(set(header_row)), "中间汇总表缺少关键表头")
        _assert(subtotal_row[0] == "筛选合计：", "中间汇总表第2行未生成筛选合计标识")
        _assert(
            isinstance(subtotal_row[9], str) and subtotal_row[9].upper().startswith("=SUBTOTAL("),
            "中间汇总表 J2 未生成 SUBTOTAL 公式",
        )

        found_data_row = False
        found_category = False
        for row in ws.iter_rows(min_row=3, values_only=True):
            if any(value not in (None, "") for value in row):
                found_data_row = True
            if _norm(row[7]) in {
                "(一)人工费",
                "(二)分包工程",
                "(三)材料费",
                "(四)机械租赁费",
                "(五)其他直接费",
                "(六)间接费",
                "(七)安全费",
            }:
                found_category = True
            if found_data_row and found_category:
                break

        _assert(found_data_row, "中间汇总表不存在有效数据行")
        _assert(found_category, "中间汇总表未发现关键成本分类")
        _print_ok(f"中间汇总表结构正常，行数 {ws.max_row}")
    finally:
        workbook.close()


def validate_delivery_workbooks(result_dir: Path) -> None:
    auto_workbook, auto_ws = _open_sheet(result_dir / AUTO_FILL_FILENAME, read_only=False)
    final_workbook, final_ws = _open_sheet(result_dir / FINAL_FILENAME, read_only=False)
    try:
        auto_title = _norm(auto_ws.cell(1, 1).value)
        final_title = _norm(final_ws.cell(1, 1).value)
        _assert("项目效益审核表" in auto_title, "自动填报版缺少主标题")
        _assert("项目效益审核表" in final_title, "最终交付版缺少主标题")

        for col_idx, label in ((2, "利润中心编码"), (3, "项目编码"), (4, "工程名称")):
            _assert(auto_ws.cell(6, col_idx).value not in (None, ""), f"自动填报版第6行缺少{label}")
            _assert(final_ws.cell(6, col_idx).value not in (None, ""), f"最终交付版第6行缺少{label}")
            _assert(
                _norm(auto_ws.cell(6, col_idx).value) == _norm(final_ws.cell(6, col_idx).value),
                f"最终交付版与自动填报版第6行{label}不一致",
            )

        key_rows = [
            "三、成本合计",
            "八、研发费用",
            "九、成本及费用合计",
            "十、增值税实际税负",
            "十一、税金及附加",
            "十二、利润",
            "十三、利润率",
        ]
        located_rows = {}
        for keyword in key_rows:
            row_idx = _find_row_by_keyword(final_ws, keyword)
            _assert(row_idx is not None, f"最终交付版缺少关键行：{keyword}")
            located_rows[keyword] = row_idx

        for keyword in ("八、研发费用", "九、成本及费用合计", "十二、利润", "十三、利润率"):
            _assert(_row_has_formula(final_ws, located_rows[keyword]), f"最终交付版关键行缺少公式：{keyword}")

        _assert(final_ws.max_row >= 100, "最终交付版行数异常偏少")
        _assert(auto_ws.max_row >= final_ws.max_row, "最终交付版行数不应大于自动填报版")
        _print_ok(
            "交付版结构正常，"
            f"自动填报版 {auto_ws.max_row} 行，最终交付版 {final_ws.max_row} 行"
        )
    finally:
        auto_workbook.close()
        final_workbook.close()


def build_parser():
    parser = argparse.ArgumentParser(description="校验最近一次样本产物是否发生结构回归")
    parser.add_argument("--output-dir", default=str(DEFAULT_OUTPUT_DIR), help="冒烟/回归输出根目录")
    parser.add_argument("--result-dir", help="直接指定待校验的结果目录")
    parser.add_argument("--run-smoke", action="store_true", help="先执行一次 smoke_test.py 再做回归校验")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    output_dir = Path(args.output_dir).resolve()

    try:
        if args.run_smoke:
            output_dir.mkdir(parents=True, exist_ok=True)
            run_smoke(output_dir)

        if args.result_dir:
            result_dir = Path(args.result_dir).resolve()
        else:
            result_dir = _find_latest_result_dir(output_dir)

        _assert(result_dir.exists(), f"结果目录不存在：{result_dir}")
        _print_info(f"校验结果目录：{result_dir}")

        validate_files(result_dir)
        validate_merged_table(result_dir)
        validate_source_table(result_dir)
        validate_delivery_workbooks(result_dir)

        print("\n回归测试通过")
        print(f"结果目录: {result_dir}")
        return 0

    except Exception as exc:
        print(f"\n回归测试失败: {exc}", file=sys.stderr)
        return 1


if __name__ == "__main__":
    raise SystemExit(main())
