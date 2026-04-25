# -*- coding: utf-8 -*-
"""
SAP 审计结果导出模块。
"""

from __future__ import annotations

import os

import openpyxl
import pandas as pd
from openpyxl.utils import get_column_letter

from etl_builder import IntermediateTableBuilder


DEFAULT_AUXILIARY_COL_ORDER = [
    "公司代码", "财年", "凭证编号", "行项目", "期间", "过帐日期", "利润中心",
    "利润中心文本描述", "自定义的凭证编号", "总账科目", "总账科目长文本",
    "对方科目", "对方科目描述", "文本", "客户", "客户描述", "供应商",
    "供应商名称", "合同", "合同文本描述", "借方本位币金额", "贷方本位币金额",
    "余额方向-本位币", "余额（本币）", "本位币类型", "中台单据号", "冲销标识", "反记帐"
]

DEFAULT_RAW_COLS_TO_DROP = [
    '匹配状态', '金额校验', '方向校验', '发生额差额', '发生额差额_含符号', '查看影像',
    '分类账', '记帐期间', '自定义的凭证编号', '凭证类型', '输入日期', '输入时间',
    '冲销关于', '借/贷标识', '货币类型（交易货币）', '带符号的交易货币金额', '带符号的本位币金额',
    'WBS元素', 'WBS元素描述', '成本中心', '成本中心描述', '业务范围', '业务范围描述',
    '地区分类档案文本描述', '资金账户', '资金账户文本描述', '事务代码', '核对线索',
    '用户名', '用户名称', '期间', '供应商名称'
]


class SAPAuditExporter:
    def __init__(self, logger, process_dir: str, result_dir: str, auxiliary_col_order=None, raw_cols_to_drop=None):
        self.log = logger
        self.process_dir = process_dir
        self.result_dir = result_dir
        self.auxiliary_col_order = auxiliary_col_order or list(DEFAULT_AUXILIARY_COL_ORDER)
        raw_cols = raw_cols_to_drop or list(DEFAULT_RAW_COLS_TO_DROP)
        self.cols_to_drop = [col for col in raw_cols if col not in self.auxiliary_col_order]

    @staticmethod
    def duplicate_suffix_columns(merged):
        return [col for col in merged.columns if str(col).endswith("_重复待删")]

    def export_process_outputs(self, merged):
        unmatched = merged[merged["匹配状态"] == "❌ 仅明细有(单边账)"]
        if not unmatched.empty:
            unmatched.to_excel(os.path.join(self.process_dir, "1_未对应凭证的异常明细表_单边账.xlsx"), index=False)

        exception_df = merged[
            (merged["发生额差额"] != "✅ 清账/反记账豁免")
            & (
                (~merged["匹配状态"].isin(["✅ 完全匹配"]))
                | (merged["金额校验"] == "❌ 金额异常")
                | (merged["方向校验"] != "")
            )
        ]
        if not exception_df.empty:
            exception_df.to_excel(os.path.join(self.process_dir, "2_全口径异常明细清单.xlsx"), index=False)

        process_main_df = merged.drop(
            columns=[col for col in self.duplicate_suffix_columns(merged) if col in merged.columns]
        ).copy()
        process_main_df.to_excel(os.path.join(self.process_dir, "3_合并大表.xlsx"), index=False)

    def build_intermediate_tables(self, merged):
        etl_builder = IntermediateTableBuilder(self.result_dir, self.log)
        etl_builder.generate(merged)

    def build_result_dataframe(self, merged):
        duplicate_cols = self.duplicate_suffix_columns(merged)
        drop_list = duplicate_cols + self.cols_to_drop
        cleaned_main_df = merged.drop(columns=[col for col in drop_list if col in merged.columns]).copy()
        ordered_cols = [col for col in self.auxiliary_col_order if col in cleaned_main_df.columns]
        final_cols = ordered_cols + [col for col in cleaned_main_df.columns if col not in ordered_cols]
        return cleaned_main_df[final_cols]

    def export_result_outputs(self, merged):
        cleaned_main_df = self.build_result_dataframe(merged)
        clean_main_table_path = os.path.join(self.result_dir, "4_合并表.xlsx")
        cleaned_main_df.to_excel(clean_main_table_path, index=False)
        self.add_subtotal(clean_main_table_path, cleaned_main_df)

    def add_subtotal(self, path, df):
        try:
            workbook = openpyxl.load_workbook(path)
            worksheet = workbook.active
            worksheet.insert_rows(2)

            for index, col in enumerate(df.columns, 1):
                is_amount_col = pd.api.types.is_numeric_dtype(df[col]) or any(
                    key in str(col) for key in ['金额', '余额', '发生额', '差额']
                )
                if is_amount_col:
                    col_letter = get_column_letter(index)
                    worksheet.cell(row=2, column=index, value=f"=SUBTOTAL(9,{col_letter}3:{col_letter}{worksheet.max_row})")
                elif index == 1:
                    worksheet.cell(row=2, column=index, value="筛选合计：")

            worksheet.freeze_panes = "A3"
            workbook.save(path)
            self.log.info("🧮 成功为【4_合并表】添加 SUBTOTAL 筛选动态合计行！")
        except Exception as exc:
            self.log.error("合并表合计行添加失败: %s", exc)
