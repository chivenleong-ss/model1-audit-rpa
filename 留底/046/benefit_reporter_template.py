# -*- coding: utf-8 -*-
"""
效益表模板定位与模板头部写入。
"""

from __future__ import annotations

import openpyxl


MATERIAL_ROW_ALIASES = {
    '材料库存': ['材料库存'],
    '材料调入': ['材料调入'],
    '材料调出': ['材料调出'],
    '材料调入+CFK': ['材料调入+CFK', 'CFK材料调入'],
    '材料调出+XSD': ['材料调出+XSD', 'XSD材料调出'],
    '材料调出+SQD': ['材料调出+SQD', 'SQD材料调出'],
    '废旧物资消耗': ['废旧物资消耗', '废旧物资处置'],
    '废旧物资消耗（SQD）': ['废旧物资消耗（SQD）', '废旧物资消耗(SQD)', '废旧物资处置+SQD'],
    '其他': ['其他'],
}

COL_A, COL_B, COL_C, COL_D = 1, 2, 3, 4
COL_H, COL_I, COL_J, COL_K = 8, 9, 10, 11
COL_L, COL_M, COL_N = 12, 13, 14
COL_O, COL_P, COL_Q, COL_R = 15, 16, 17, 18
COL_S, COL_T = 19, 20

SUBTOTAL_MAP = {
    '（一）人工费': '人工费小计',
    '（二）分包工程': '分包工程小计',
    '(三)材料费': '材料费小计',
    '(四）机械租赁费': '机械费小计',
    '（五）其他直接费': '其他直接费小计',
    '（六）间接费': '间接费小计',
    '（七）安全费': '安全费小计',
}


class BenefitTemplateLocator:
    def __init__(self, logger):
        self.log = logger

    @staticmethod
    def norm(val) -> str:
        if val is None:
            return ""
        return (
            str(val)
            .replace(" ", "")
            .replace("\u3000", "")
            .replace("（", "(")
            .replace("）", ")")
            .replace("：", ":")
            .replace("*", "")
            .replace(":", "")
        )

    def find_row(self, ws, keyword, cols=(2, 3, 4, 5)) -> int | None:
        normalized = self.norm(keyword)
        if not normalized:
            return None
        for row in range(1, ws.max_row + 1):
            for col in cols:
                if normalized in self.norm(ws.cell(row, col).value):
                    return row
        return None

    def find_material_row(self, ws, keyword) -> int | None:
        for alias in MATERIAL_ROW_ALIASES.get(keyword, [keyword]):
            row = self.find_row(ws, alias)
            if row:
                return row
        return None

    def find_anchor(self, ws, subtotal_kw, min_row: int = 1) -> int | None:
        subtotal_row = self.find_row(ws, subtotal_kw)
        if not subtotal_row:
            return None
        for row in range(subtotal_row - 1, max(min_row, subtotal_row - 300), -1):
            for col in range(2, 6):
                if '财务未列部分需增列' in self.norm(ws.cell(row, col).value):
                    return row
        return None

    def load_template_sheet(self, template_path: str):
        workbook = openpyxl.load_workbook(template_path)
        worksheet = workbook.active
        for sheet_name in workbook.sheetnames:
            if any(key in sheet_name for key in ['效益审核', '附表']):
                worksheet = workbook[sheet_name]
                break
        return workbook, worksheet

    @staticmethod
    def first_non_empty(df, col):
        if col not in df.columns:
            return ''
        values = df[col].replace('', None).dropna()
        return values.iloc[0] if not values.empty else ''

    def fill_header_fields(self, ws, df, df4) -> None:
        ws['B6'] = self.first_non_empty(df, '利润中心')
        ws['C6'] = self.first_non_empty(df, '项目编码')
        ws['D6'] = self.first_non_empty(df, '工程名称')

        if not df4.empty and '总账科目长文本' in df4.columns:
            revenue = df4[df4['总账科目长文本'].fillna('').str.startswith('主营业务收入')]['贷方本位币金额'].sum()
            ws['M6'] = round(-float(revenue), 2)
        else:
            revenue = df[df['成本_财务大类'] == '财务累计入账收入(不含增值税)']['最终发生额'].sum()
            ws['M6'] = round(-float(revenue), 2)

        vat_amount = df[df['成本_财务大类'] == '已价税分离的增值税金额(财务账面数据)']['最终发生额'].sum()
        ws['N6'] = round(-float(vat_amount), 2)
        self.log.info("   ✅ 表头写入（B6/C6/D6/M6/N6）")
