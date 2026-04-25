# -*- coding: utf-8 -*-
"""
excel_beautifier.py — Excel 精准排版与美化引擎 (纯画笔终极版 v4)
核心修复：
1. 【彻底根治公式断层】：强行剔除“删空行”和“标序号”的逻辑（已完全交由 benefit_reporter 处理）。
2. 本模块只进行安全渲染：边框、底色、千位分隔符、格式克隆，绝对不改变表格物理结构！
"""

import os
import openpyxl
from openpyxl.styles import Border, Side, PatternFill
import copy as py_copy

class ExcelBeautifier:
    def __init__(self, result_dir, logger):
        self.result_dir = result_dir
        self.log = logger
        self.input_path = os.path.join(self.result_dir, "自动填报完成_效益审核表.xlsx")
        self.output_path = os.path.join(self.result_dir, "交付版_效益审核表.xlsx")

    def normalize(self, text):
        return str(text).replace(" ", "").replace("（", "(").replace("）", ")").replace("：", ":")

    def execute_beautify(self):
        if not os.path.exists(self.input_path):
            self.log.error(f"❌ 美化引擎未找到输入文件: {self.input_path}")
            return False

        self.log.info("💅 启动独立排版美化引擎，执行纯视觉渲染(绝对保护公式)...")
        
        try:
            wb = openpyxl.load_workbook(self.input_path)
            ws = wb.active
            for sn in wb.sheetnames:
                if "效益审核" in sn or "附表" in sn:
                    ws = wb[sn]
                    break

            # 统一列数上限：扩展至 AL 列 (第38列)
            MAX_COL = 38

            # ========================================================
            # 1. 复制指定标题格式 (覆盖至 AL 列)
            # ========================================================
            self.log.info("   🎨 正在统一指定汇总行格式...")
            r_start = None
            for r in range(7, ws.max_row + 1):
                if "三、成本合计" in self.normalize(ws.cell(r, 4).value or ""):
                    r_start = r
                    break
            
            if r_start:
                target_keywords = ["四、计提", "五、过程", "六、资金", "七、局投资", "八、研发"]
                for r in range(r_start + 1, ws.max_row + 1):
                    val_d = self.normalize(ws.cell(r, 4).value or "")
                    if any(kw in val_d for kw in target_keywords):
                        for c in range(1, MAX_COL + 1):
                            src = ws.cell(r_start, c)
                            tgt = ws.cell(r, c)
                            if src.has_style:
                                if src.font: tgt.font = py_copy.copy(src.font)
                                if src.fill: tgt.fill = py_copy.copy(src.fill)
                                if src.alignment: tgt.alignment = py_copy.copy(src.alignment)

            # ========================================================
            # 2. 消除明细行底色 (防解析崩溃：采用纯白色实体填充)
            # ========================================================
            self.log.info("   🖌️ 正在安全擦除明细行底色...")
            safe_white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
            
            r_yanfa = None
            for r in range(7, ws.max_row + 1):
                if "八、研发" in self.normalize(ws.cell(r, 4).value or ""):
                    r_yanfa = r
                    break

            if r_yanfa:
                for r in range(r_yanfa + 1, ws.max_row + 1):
                    val_d = self.normalize(ws.cell(r, 4).value or "")
                    if "小计" in val_d or "合计" in val_d:
                        break 
                    if val_d:
                        for c in range(1, MAX_COL + 1):
                            ws.cell(r, c).fill = safe_white_fill

            # ========================================================
            # 3. 全局补齐缺失的单元格边框 (扩展至 AL 列)
            # ========================================================
            self.log.info("   🔲 正在全局补齐数据区单元格边框...")
            thin = Side(border_style="thin", color="000000")
            standard_border = Border(left=thin, right=thin, top=thin, bottom=thin)

            for r in range(7, ws.max_row + 1):
                val_d = ws.cell(r, 4).value
                val_m = ws.cell(r, 13).value
                if not val_d and not val_m:
                    continue
                
                for c in range(1, MAX_COL + 1):
                    ws.cell(r, c).border = standard_border

            # ========================================================
            # 4. 千位分隔符格式化 (E列至AK列，二、成本 至 十一、利润)
            # ========================================================
            self.log.info("   💲 正在批量设置金额列的千位分隔符...")
            r_cost = None
            r_profit = None
            for r in range(1, ws.max_row + 1):
                val_d = self.normalize(ws.cell(r, 4).value or "")
                if "二、成本" in val_d: r_cost = r
                if "十一、利润" in val_d: r_profit = r
            
            if r_cost and r_profit:
                for r in range(r_cost, r_profit + 1):
                    # E列(5) 到 AK列(37)
                    for c in range(5, 38): 
                        cell = ws.cell(r, c)
                        if isinstance(cell.value, (int, float)):
                            cell.number_format = '#,##0.00'

            # ========================================================
            # 5. L列(12)一致性填充 (第10行至 不含税利润率)
            # ========================================================
            self.log.info("   🎨 正在统一 L列 的底色填充，消除断层空白...")
            r_margin = None
            for r in range(1, ws.max_row + 1):
                if "不含税利润率" in self.normalize(ws.cell(r, 4).value or ""):
                    r_margin = r
                    break
            
            if r_margin and r_margin >= 10:
                ref_fill = ws.cell(9, 12).fill
                if ref_fill and ref_fill.fill_type and ref_fill.fill_type != "none":
                    target_fill = py_copy.copy(ref_fill)
                else:
                    target_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
                
                for r in range(10, r_margin + 1):
                    ws.cell(r, 12).fill = target_fill

            wb.save(self.output_path)
            self.log.info(f"✨ 完美收官！底层结构完整无损，【交付版_效益审核表.xlsx】已生成！")
            return True

        except Exception as e:
            self.log.error(f"❌ 美化时出现异常: {e}")
            import traceback
            self.log.error(traceback.format_exc())
            return False