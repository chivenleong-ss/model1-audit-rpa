# -*- coding: utf-8 -*-
import os
import logging
import pandas as pd
import numpy as np
import openpyxl
from openpyxl.utils import get_column_letter
from datetime import datetime

# ★ 引入拆分出去的 ETL 新模块 ★
from etl_builder import IntermediateTableBuilder

# ==========================================
# 日志配置
# ==========================================
def setup_logger(log_dir: str = ".") -> logging.Logger:
    logger = logging.getLogger("sap_audit")
    if logger.handlers: return logger
    logger.setLevel(logging.INFO)
    fmt = logging.Formatter("%(asctime)s  %(levelname)-7s  %(message)s", datefmt="%Y-%m-%d %H:%M:%S")
    ch = logging.StreamHandler()
    ch.setFormatter(fmt)
    logger.addHandler(ch)
    return logger

log = setup_logger()

class SAPAuditModel:
    def __init__(self, voucher_file, detail_file, output_base_dir='2_输出审计底稿', tolerance=0.01):
        self.voucher_path = voucher_file
        self.detail_path = detail_file
        self.tolerance = tolerance 
        self.join_keys = ['公司代码', '财年', '凭证编号', '行项目']
        
        self.alias_map = {
            '公司代码': ['公司', 'Company Code', 'BUKRS'],
            '财年': ['会计年度', '年度', 'Fiscal Year', 'GJAHR', '年度/期间'],
            '凭证编号': ['凭证号', '会计凭证', 'Document Number', 'BELNR'],
            '行项目': ['项目', '行号', 'Item No', 'BUZEI'],
            '本位币金额': ['金额', '主表金额', 'Amount in LC', 'DMBTR'],
            '借方本位币金额': ['借方金额', '借方', 'Debit'],
            '贷方本位币金额': ['贷方金额', '贷方', 'Credit'],
            '文本': ['摘要', 'Item Text', 'SGTXT', 'Text'],
            '反记帐': ['Reverse posting', 'XNEGP', '反记账标识', '反记账']
        }

        # 双输出文件夹架构
        time_str = datetime.now().strftime('%Y%m%d_%H%M%S')
        self.process_dir = os.path.join(output_base_dir, f"过程表单_{time_str}")
        os.makedirs(self.process_dir, exist_ok=True)
        self.result_dir = os.path.join(output_base_dir, f"审计核对结果_{time_str}")
        os.makedirs(self.result_dir, exist_ok=True)

        self.auxiliary_col_order = [
            "公司代码", "财年", "凭证编号", "行项目", "期间", "过帐日期", "利润中心", 
            "利润中心文本描述", "自定义的凭证编号", "总账科目", "总账科目长文本", 
            "对方科目", "对方科目描述", "文本", "客户", "客户描述", "供应商", 
            "供应商名称", "合同", "合同文本描述", "借方本位币金额", "贷方本位币金额", 
            "余额方向-本位币", "余额（本币）", "本位币类型", "中台单据号", "冲销标识", "反记帐"
        ]

        raw_cols_to_drop = [
            '匹配状态', '金额校验', '方向校验', '发生额差额', '发生额差额_含符号', '查看影像',
            '分类账', '记帐期间', '自定义的凭证编号', '凭证类型', '输入日期', '输入时间',
            '冲销关于', '借/贷标识', '货币类型（交易货币）','带符号的交易货币金额','带符号的本位币金额',
            'WBS元素', 'WBS元素描述', '成本中心', '成本中心描述', '业务范围', '业务范围描述', 
            '地区分类档案文本描述', '资金账户', '资金账户文本描述', '事务代码', '核对线索', 
            '用户名', '用户名称', '期间', '供应商名称'
        ]
        self.cols_to_drop = [c for c in raw_cols_to_drop if c not in self.auxiliary_col_order]

    def _smart_read(self, path):
        log.info(f"读取文件: {os.path.basename(path)}")
        if path.lower().endswith('.csv'):
            try: return pd.read_csv(path, low_memory=False)
            except: return pd.read_csv(path, encoding='gbk', low_memory=False)
        return pd.read_excel(path).dropna(how='all')

    def _preprocess(self, df, table_name):
        for std_name, aliases in self.alias_map.items():
            for alias in aliases:
                if alias in df.columns and std_name not in df.columns:
                    df = df.rename(columns={alias: std_name})
        
        for key in self.join_keys:
            if key in df.columns:
                df[key] = df[key].astype(str).str.strip().str.replace(r'\.0$', '', regex=True).replace(['nan', 'None', ''], np.nan)
                
        key_cols = [k for k in self.join_keys if k in df.columns]
        if key_cols:
            before_len = len(df)
            df = df.dropna(subset=key_cols, how='any')
            if before_len - len(df) > 0:
                log.info(f"   🗑️ 清洗拦截了 {before_len - len(df)} 行无效数据。")
        return df

    def execute_audit(self):
        log.info("=" * 55)
        log.info("🚀 启动自动化审计校验模块 (基础比对)")
        log.info("=" * 55)

        v_df = self._preprocess(self._smart_read(self.voucher_path), "凭证主表")
        d_df = self._preprocess(self._smart_read(self.detail_path),  "明细账表")

        merged = pd.merge(v_df, d_df, on=self.join_keys, how="outer", indicator="匹配状态", suffixes=("", "_重复待删"))
        merged["匹配状态"] = merged["匹配状态"].astype(str).replace({"both": "✅ 完全匹配", "left_only": "❌ 仅主表有(缺明细)", "right_only": "❌ 仅明细有(单边账)"})

        for col in ["本位币金额", "借方本位币金额", "贷方本位币金额"]:
            if col in merged.columns: merged[col] = pd.to_numeric(merged[col], errors="coerce").fillna(0)

        # ★★★ 智能补齐客商与合同信息 ★★★
        log.info("🔄 正在执行同凭证内信息智能广播补齐 (客商、合同)...")
        fill_cols = ['供应商', '供应商名称', '合同', '合同文本描述', '客户', '客户描述']
        existing_fill_cols = [c for c in fill_cols if c in merged.columns]
        group_keys = [k for k in ['公司代码', '财年', '凭证编号'] if k in merged.columns]

        if group_keys and existing_fill_cols:
            for c in existing_fill_cols:
                merged[c] = merged[c].replace(r'^\s*$', np.nan, regex=True)
            merged[existing_fill_cols] = merged.groupby(group_keys, dropna=False)[existing_fill_cols].transform(lambda x: x.ffill().bfill())

        v_amt = merged["本位币金额"].abs()
        d_amt_signed = merged["借方本位币金额"].fillna(0) - merged["贷方本位币金额"].fillna(0)
        d_amt = d_amt_signed.abs()

        merged["发生额差额"] = (v_amt - d_amt).round(2).astype(str)
        merged["发生额差额_含符号"] = (merged["本位币金额"] - d_amt_signed).round(2).astype(str)
        
        merged["金额校验"] = np.where((v_amt - d_amt).round(2).abs() <= self.tolerance, "✅ 金额正确", "❌ 金额异常")
        merged["方向校验"] = np.where((merged["金额校验"] == "✅ 金额正确") & ((merged["本位币金额"] - d_amt_signed).round(2).abs() > self.tolerance), "⚠️ 借贷方向疑似反向", "")

        ex_mask = pd.Series(False, index=merged.index)
        if "文本" in merged.columns: ex_mask = ex_mask | (merged["文本"] == "自动清账剩余项目")
        if "反记帐" in merged.columns: ex_mask = ex_mask | (merged["反记帐"].notna() & (merged["反记帐"] != ""))

        merged.loc[ex_mask, "发生额差额"] = "✅ 清账/反记账豁免"
        merged.loc[ex_mask, "发生额差额_含符号"] = "✅ 清账/反记账豁免"
        merged.loc[ex_mask, "方向校验"] = ""

        # ---------------- 1. 导出至【过程表单】 ----------------
        unmatched = merged[merged["匹配状态"] == "❌ 仅明细有(单边账)"]
        if not unmatched.empty: 
            unmatched.to_excel(os.path.join(self.process_dir, "1_未对应凭证的异常明细表_单边账.xlsx"), index=False)

        exception_df = merged[(merged["发生额差额"] != "✅ 清账/反记账豁免") & ((~merged["匹配状态"].isin(["✅ 完全匹配"])) | (merged["金额校验"] == "❌ 金额异常") | (merged["方向校验"] != ""))]
        if not exception_df.empty:
            exception_df.to_excel(os.path.join(self.process_dir, "2_全口径异常明细清单.xlsx"), index=False)

        base_drop = [c for c in merged.columns if str(c).endswith("_重复待删")]
        base_main_df = merged.drop(columns=[c for c in base_drop if c in merged.columns]).copy()
        base_main_df.to_excel(os.path.join(self.process_dir, "3_合并大表.xlsx"), index=False)

        # ---------------- 2. ★ 核心解耦：调用独立的 ETL 模块提取【5_中间汇总表】 ★ ----------------
        etl_builder = IntermediateTableBuilder(self.result_dir, log)
        etl_builder.generate(merged)

        # ---------------- 3. 导出至【审计核对结果】(4_合并表) ----------------
        drop_list = base_drop + self.cols_to_drop
        cleaned_main_df = merged.drop(columns=[c for c in drop_list if c in merged.columns]).copy()

        ordered_cols = [c for c in self.auxiliary_col_order if c in cleaned_main_df.columns]
        final_cols = ordered_cols + [c for c in cleaned_main_df.columns if c not in ordered_cols]
        cleaned_main_df = cleaned_main_df[final_cols]

        clean_main_table_path = os.path.join(self.result_dir, "4_合并表.xlsx")
        cleaned_main_df.to_excel(clean_main_table_path, index=False)
        self._add_subtotal(clean_main_table_path, cleaned_main_df)

        return True

    # ========================================================
    # 为合并表添加 SUBTOTAL 筛选动态合计行
    # ========================================================
    def _add_subtotal(self, path, df):
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            ws.insert_rows(2) # 在标题行下方插入空行
            
            for i, col in enumerate(df.columns, 1):
                # 强化识别逻辑：如果是数值类型，或者列名包含'金额'、'余额'、'发生额'、'差额'，都加公式
                is_amount_col = pd.api.types.is_numeric_dtype(df[col]) or any(k in str(col) for k in ['金额', '余额', '发生额', '差额'])
                
                if is_amount_col:
                    col_let = get_column_letter(i)
                    ws.cell(row=2, column=i, value=f"=SUBTOTAL(9,{col_let}3:{col_let}{ws.max_row})")
                elif i == 1: 
                    ws.cell(row=2, column=i, value="筛选合计：")
            
            ws.freeze_panes = "A3" 
            wb.save(path)
            log.info(f"🧮 成功为【4_合并表】添加 SUBTOTAL 筛选动态合计行！")
        except Exception as e: 
            log.error(f"合并表合计行添加失败: {e}")