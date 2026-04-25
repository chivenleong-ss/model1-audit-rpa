# -*- coding: utf-8 -*-
"""
excel_beautifier.py v3 — 排版美化引擎（全面修正版）

修复清单：
  1. L列填充色与J列保持一致（逐行复制）
  2. 删除研发费用节内的空公式占位行（无D、无M值的行）
  3. A列序号：补全研发费用节及其他各节的序号
  4. 修复「九、成本及费用合计」公式（#REF! → 正确行号）
  5. 修复「十二、利润」「十三、利润率」公式中失效的旧行引用
  6. 全局边框补齐（AL列）
  7. 汇总行格式统一（四~八 与三统一样式）
"""

import os, re, traceback
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, PatternFill
import copy as py_copy

GREY_SET = {"FFBFBFBF", "FFD9D9D9", "FFC0C0C0", "FFE0E0E0",
            "FFD3D3D3", "FFCCCCCC", "FFB8B8B8", "FFAEAAAA"}
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")


class ExcelBeautifier:

    def __init__(self, result_dir, logger):
        self.result_dir  = result_dir
        self.log         = logger
        self.input_path  = os.path.join(result_dir, "自动填报完成_效益审核表.xlsx")
        self.output_path = os.path.join(result_dir, "最终完美交付版_效益审核表.xlsx")

    # ── 工具 ─────────────────────────────────────────────────────
    @staticmethod
    def _norm(val) -> str:
        if val is None: return ""
        return (str(val)
                .replace(" ", "").replace("\u3000", "")
                .replace("（", "(").replace("）", ")")
                .replace("：", ":").replace("*", "")
                .replace(":", ""))

    def _find_row(self, ws, keyword, cols=(2, 3, 4, 5)) -> int | None:
        kw = self._norm(keyword)
        for r in range(1, ws.max_row + 1):
            for c in cols:
                if kw in self._norm(ws.cell(r, c).value):
                    return r
        return None

    @staticmethod
    def _get_fill_rgb(cell) -> str:
        try:
            f = cell.fill
            if f and f.fill_type not in (None, "none"):
                return f.fgColor.rgb
        except Exception:
            pass
        return ""

    # ════════════════════════════════════════════════════════════
    #  1. L列填充色与J列一致
    # ════════════════════════════════════════════════════════════
    def _sync_l_fill_to_j(self, ws):
        """逐行将 L 列的填充色设为与 J 列完全相同（含无填充情况）。"""
        NO_FILL = PatternFill(fill_type=None)
        fixed = 0
        for r in range(7, ws.max_row + 1):
            j_cell = ws.cell(r, 10)
            l_cell = ws.cell(r, 12)
            j_rgb  = self._get_fill_rgb(j_cell)
            l_rgb  = self._get_fill_rgb(l_cell)
            if j_rgb == l_rgb:
                continue
            try:
                l_cell.fill = py_copy.copy(j_cell.fill) if j_rgb else NO_FILL
                fixed += 1
            except Exception:
                pass
        self.log.info("   🎨 L列填充色与J列同步 (%d 个单元格)", fixed)

    # ════════════════════════════════════════════════════════════
    #  2. 删除研发费用节内的空占位行
    # ════════════════════════════════════════════════════════════
    def _delete_yanfa_ghost_rows(self, ws):
        """
        八、研发费用 ～ 九、成本及费用合计 之间，
        删除没有 D 值且 M 列不是真实数值的占位行（纯公式行）。
        """
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine  = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        if not r_yanfa or not r_nine:
            return
        to_del = []
        for r in range(r_yanfa + 1, r_nine):
            d = ws.cell(r, 4).value
            m = ws.cell(r, 13).value
            d_empty = (d is None) or (isinstance(d, str) and not d.strip())
            if d_empty and not isinstance(m, (int, float)):
                to_del.append(r)
        for r in reversed(to_del):
            ws.delete_rows(r)
        if to_del:
            self.log.info("   🗑 删除研发费用节空占位行 %d 行", len(to_del))

    def _fix_yanfa_subtotal_formula(self, ws):
        """
        修复“八、研发费用”行的 SUM 引用范围，避免插行/删行后仍指向模板旧行号。
        """
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine  = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        if not r_yanfa or not r_nine or r_nine <= r_yanfa:
            return
        start_row = r_yanfa + 1
        end_row = max(start_row, r_nine - 1)
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(r_yanfa, col)
            if cell.value and isinstance(cell.value, str) and 'SUM(' in cell.value.upper():
                cl = get_column_letter(col)
                cell.value = f"=SUM({cl}{start_row}:{cl}{end_row})"
        self.log.info("   🧮 研发费用小计公式修正（行 %d，范围 %d-%d）", r_yanfa, start_row, end_row)

    # ════════════════════════════════════════════════════════════
    #  3. A列序号补全（全节，含研发费用节）
    # ════════════════════════════════════════════════════════════
    def _fill_sequence_numbers(self, ws):
        """
        对每个成本节内有数据（D有内容 且 M为数值）的行写 A列序号。
        覆盖 一 至 七 所有节 + 八、研发费用 节。
        """
        SECTIONS = [
            ('（一）人工费',      '人工费小计'),
            ('（二）分包工程',    '分包工程小计'),
            ('(三)材料费',       '材料费小计'),
            ('(四）机械租赁费',   '机械费小计'),
            ('（五）其他直接费',  '其他直接费小计'),
            ('（六）间接费',      '间接费小计'),
            ('（七）安全费',      '安全费小计'),
        ]
        # 研发费用节单独处理（止于九）
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine  = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')

        def _number_section(hdr, end_excl):
            for r in range(hdr + 1, end_excl):
                ws.cell(r, 1).value = None
            seq = 1
            for r in range(hdr + 1, end_excl):
                d = ws.cell(r, 4).value
                m = ws.cell(r, 13).value
                if d and isinstance(m, (int, float)):
                    ws.cell(r, 1).value = seq
                    seq += 1

        for hdr_kw, sub_kw in SECTIONS:
            hdr = self._find_row(ws, hdr_kw)
            sub = self._find_row(ws, sub_kw)
            if not hdr or not sub:
                continue
            # 锚点（财务未列部分需增列）
            anchor = None
            for r in range(sub - 1, hdr, -1):
                for c in range(2, 6):
                    if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                        anchor = r; break
                if anchor: break
            end = (anchor - 1) if anchor else (sub - 1)
            _number_section(hdr, end + 1)

        # 研发费用节
        if r_yanfa and r_nine:
            _number_section(r_yanfa, r_nine)

        self.log.info("   🔢 A列序号全节补全完成")

    def _fix_yanfa_detail_fill(self, ws):
        """
        研发费用明细项（D列）应为明细行样式，不能继承节标题灰色底。
        这里优先用“研发费用节内第一个明细行”的 C/D 样式作为基准进行回填；
        如果找不到基准行，则退化为“无填充”。
        """
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine  = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        if not r_yanfa or not r_nine:
            return
        no_fill = PatternFill(fill_type=None)
        ref_fill_c = ref_fill_d = None
        for r in range(r_yanfa + 1, r_nine):
            d = ws.cell(r, 4).value
            m = ws.cell(r, 13).value
            if d and isinstance(m, (int, float)):
                ref_fill_c = py_copy.copy(ws.cell(r, 3).fill)
                ref_fill_d = py_copy.copy(ws.cell(r, 4).fill)
                break
        fixed = 0
        for r in range(r_yanfa + 1, r_nine):
            d = ws.cell(r, 4).value
            m = ws.cell(r, 13).value
            if d and isinstance(m, (int, float)):
                ws.cell(r, 3).fill = py_copy.copy(ref_fill_c) if ref_fill_c else py_copy.copy(no_fill)
                ws.cell(r, 4).fill = py_copy.copy(ref_fill_d) if ref_fill_d else py_copy.copy(no_fill)
                fixed += 1
        if fixed:
            self.log.info("   🎨 研发费用明细 C/D 列填充修正 %d 行", fixed)

    def _clear_yanfa_section_s_fill(self, ws):
        """
        审核表 S 列中，“八、研发费用”与“九、成本及费用合计”之间应保持空白填充，不继承灰底。
        """
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        if not r_yanfa or not r_nine:
            return
        no_fill = PatternFill(fill_type=None)
        fixed = 0
        for r in range(r_yanfa + 1, r_nine):
            ws.cell(r, 19).fill = py_copy.copy(no_fill)
            fixed += 1
        if fixed:
            self.log.info("   🎨 研发费用区间 S 列填充清空 %d 行", fixed)

    # ════════════════════════════════════════════════════════════
    #  4. 修复「九、成本及费用合计」公式
    # ════════════════════════════════════════════════════════════
    def _fix_nine_formula(self, ws):
        """
        动态定位 三-八 各汇总行，重建九的所有列公式。
        """
        ROW_KWS = [
            '三、成本合计',
            '四、计提保修金',
            '五、过程节点奖金',
            '六、资金占用费用',
            '七、局投资收益',
            '八、研发费用',
        ]
        row_nums = []
        for kw in ROW_KWS:
            r = self._find_row(ws, kw)
            if r:
                row_nums.append(r)

        r_nine = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        if not r_nine or not row_nums:
            self.log.warning("   ⚠️ 未能定位九、成本及费用合计，跳过公式修复")
            return

        for col in range(8, ws.max_column + 1):
            cell = ws.cell(r_nine, col)
            if cell.value is None:
                continue
            # 只修复含公式的格（不论是否有 #REF!）
            cl = get_column_letter(col)
            cell.value = '=' + '+'.join(f'{cl}{r}' for r in row_nums)

        self.log.info("   🧮 九、成本及费用合计 公式修正（参考行 %s）",
                      '+'.join(str(r) for r in row_nums))

    # ════════════════════════════════════════════════════════════
    #  5. 修复十二利润 / 十三利润率 公式
    # ════════════════════════════════════════════════════════════
    def _fix_profit_formulas(self, ws):
        """
        十二、利润 = J6 - 九 - 十增值税 - 十一税金
        十三、利润率 = 利润行 / 对应收入行
        """
        r6    = 6   # 收入行固定
        r_nine  = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        r_vat   = self._find_row(ws, '十、增值税实际税负')
        r_tax   = self._find_row(ws, '十一、税金及附加')
        r_profit = self._find_row(ws, '十二、利润')
        r_rate   = self._find_row(ws, '十三、利润率')

        # 十二、利润 — 对有公式的列重建
        if r_profit and r_nine:
            deduct = [r for r in [r_nine, r_vat, r_tax] if r]
            for col in range(8, ws.max_column + 1):
                cell = ws.cell(r_profit, col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cl = get_column_letter(col)
                    cell.value = f'={cl}{r6}' + ''.join(f'-{cl}{r}' for r in deduct)
            self.log.info("   🧮 十二、利润 公式修正（行 %d）", r_profit)

        # 十三、利润率 — 对有公式的列更新被除数行
        if r_rate and r_profit:
            for col in range(8, ws.max_column + 1):
                cell = ws.cell(r_rate, col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cl = get_column_letter(col)
                    # 重建为 =Xcurrent_profit / Xcol6
                    cell.value = f'={cl}{r_profit}/{cl}{r6}'
            self.log.info("   🧮 十三、利润率 公式修正（行 %d）", r_rate)

    # ════════════════════════════════════════════════════════════
    #  6. 删除尾部空行
    # ════════════════════════════════════════════════════════════
    def _delete_trailing_empty_rows(self, ws):
        deleted = 0
        for r in range(ws.max_row, 10, -1):
            if all(ws.cell(r, c).value is None for c in range(1, 40)):
                ws.delete_rows(r)
                deleted += 1
            else:
                break
        if deleted:
            self.log.info("   🗑 删除尾部空行 %d 行", deleted)

    # ════════════════════════════════════════════════════════════
    #  7. 汇总行格式统一（四~八 与三成本合计一致）
    # ════════════════════════════════════════════════════════════
    def _unify_summary_rows(self, ws, max_col=38):
        r_base = self._find_row(ws, '三、成本合计')
        if not r_base: return
        KEYWORDS = ["四、计提", "五、过程", "六、资金", "七、局投资", "八、研发"]
        for r in range(r_base + 1, ws.max_row + 1):
            val = self._norm(ws.cell(r, 4).value or "")
            if any(k in val for k in KEYWORDS):
                for c in range(1, max_col + 1):
                    src = ws.cell(r_base, c)
                    tgt = ws.cell(r, c)
                    try:
                        if src.font:      tgt.font      = py_copy.copy(src.font)
                        if src.fill:      tgt.fill      = py_copy.copy(src.fill)
                        if src.alignment: tgt.alignment = py_copy.copy(src.alignment)
                    except Exception:
                        pass

    # ════════════════════════════════════════════════════════════
    #  8. 全局边框补齐
    # ════════════════════════════════════════════════════════════
    def _apply_borders(self, ws, max_col=38):
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in range(7, ws.max_row + 1):
            if not ws.cell(r, 4).value and not ws.cell(r, 13).value:
                continue
            for c in range(1, max_col + 1):
                ws.cell(r, c).border = border

    # ════════════════════════════════════════════════════════════
    #  主流程（顺序固定，不可调换）
    # ════════════════════════════════════════════════════════════
    def execute_beautify(self):
        if not os.path.exists(self.input_path):
            self.log.error("❌ 找不到输入文件：%s", self.input_path)
            return False

        self.log.info("💅 美化引擎 v3 启动...")
        try:
            wb = openpyxl.load_workbook(self.input_path)
            ws = wb.active
            for sn in wb.sheetnames:
                if any(k in sn for k in ["效益审核", "附表"]):
                    ws = wb[sn]; break

            # 顺序：结构修复 → 公式修复 → 格式美化
            self._delete_trailing_empty_rows(ws)   # ① 尾部空行
            self._delete_yanfa_ghost_rows(ws)       # ② 研发节空占位行
            self._fix_yanfa_subtotal_formula(ws)    # ③ 研发费用小计
            self._fix_nine_formula(ws)              # ④ 九、成本及费用合计
            self._fix_profit_formulas(ws)           # ⑤ 十二利润 / 十三利润率
            self._fill_sequence_numbers(ws)         # ⑥ A列序号全节
            self._fix_yanfa_detail_fill(ws)         # ⑦ 研发费用明细填色纠偏
            self._clear_yanfa_section_s_fill(ws)    # ⑧ 研发费用区间 S列去灰底
            self._unify_summary_rows(ws)            # ⑨ 汇总行样式统一
            self._sync_l_fill_to_j(ws)              # ⑩ L列填充=J列（必须在格式操作后）
            self._apply_borders(ws)                 # ⑪ 全局边框

            wb.save(self.output_path)
            self.log.info("✨ 交付版生成完成：%s", self.output_path)
            return True

        except Exception as e:
            self.log.error("❌ 美化异常:\n%s", traceback.format_exc())
            return False
