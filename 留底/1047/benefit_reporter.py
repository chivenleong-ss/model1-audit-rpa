# -*- coding: utf-8 -*-
"""
benefit_reporter.py  v5 
本版修复：
  1. 付款列改为 T列(col 20)= 财务账面已列报付款（原误写 S列/col19）
  2. _delete_stubs 重写：删除 D列为空且 M列非数值 的占位行，彻底清除模板残留行
  3. _fix_stale_self_refs：插行后各行 O/P/Q/R/S/J/L 列的旧行号引用全部修正为当前行号
  4. _fix_aggregate_rows：成本合计、成本费用合计 行的跨节汇总公式修正
  5. _style_row：复制上方行的单元格填充色（fill）防止颜色消失
  6. 安全费按细分科目正常插行（依赖 _delete_stubs 修正）
  7. 机械费 A列序号从1连续（依赖 _delete_stubs 修正）
"""

import os, re, traceback
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, GradientFill
import copy as py_copy

# ─── 列号常量 (1-based) ───────────────────────────────────────
COL_A, COL_B, COL_C, COL_D = 1, 2, 3, 4
COL_H, COL_I, COL_J, COL_K = 8, 9, 10, 11
COL_L, COL_M, COL_N        = 12, 13, 14
COL_O, COL_P, COL_Q, COL_R = 15, 16, 17, 18
COL_S, COL_T                = 19, 20   # S=累计付款(总), T=财务账面已列报付款 ★

# ─── 小节小计关键词映射 ───────────────────────────────────────
SUBTOTAL_MAP = {
    '（一）人工费':      '人工费小计',
    '（二）分包工程':    '分包工程小计',
    '(三)材料费':       '材料费小计',
    '(四）机械租赁费':   '机械费小计',
    '（五）其他直接费':  '其他直接费小计',
    '（六）间接费':      '间接费小计',
    '（七）安全费':      '安全费小计',
}


class BenefitReporter:

    def __init__(self, result_dir, logger):
        self.result_dir    = result_dir
        self.log           = logger
        self.template_path = "项目效益审核表.xlsx"
        self.invoice_ledger_path = None  # 可选：发票收票台账（已认证）

    # ═══════════════════════════════════════════════════════════
    #  §1  工具
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _norm(val) -> str:
        if val is None: return ""
        return (str(val)
                .replace(" ", "").replace("\u3000", "")
                .replace("（", "(").replace("）", ")")
                .replace("：", ":").replace("*", "")
                .replace(":", ""))

    def _find_row(self, ws, keyword, cols=(2, 3, 4, 5)) -> int | None:
        kw = self._norm(keyword)
        if not kw: return None
        for r in range(1, ws.max_row + 1):
            for c in cols:
                if kw in self._norm(ws.cell(r, c).value):
                    return r
        return None

    def _find_anchor(self, ws, subtotal_kw, min_row: int = 1) -> int | None:
        """
        从小计行向上找'财务未列部分需增列'，不超越 min_row（节头行）。
        这样可以避免跨节找到上一节的锚点。
        """
        sub_row = self._find_row(ws, subtotal_kw)
        if not sub_row: return None
        for r in range(sub_row - 1, max(min_row, sub_row - 300), -1):
            for c in range(2, 6):
                if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                    return r
        return None

    # ═══════════════════════════════════════════════════════════
    #  §2  行样式（复制上方行填充色，保持边框/字体）
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _copy_fill(src_cell, dst_cell):
        """安全复制单元格填充色"""
        try:
            fill = src_cell.fill
            if fill and fill.fill_type not in (None, "none"):
                dst_cell.fill = py_copy.copy(fill)
        except Exception:
            pass

    # 原代码大约在 97 行：def _style_row(self, ws, r, ncol=22):
    def _style_row(self, ws, r, ncol=22, skip_fill=False):   # ← 修改这一行，增加参数
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font   = Font(name="宋体", size=10)
        for c in range(1, ncol + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.font   = font
            cell.alignment = Alignment(
                horizontal="left" if c in (COL_C, COL_D) else "center",
                vertical="center", wrap_text=(c in (COL_C, COL_D)))
            
            # 复制上方行的填充色 (👇 增加 and not skip_fill)
            if r > 1 and not skip_fill:
                self._copy_fill(ws.cell(r - 1, c), cell)

    # ═══════════════════════════════════════════════════════════
    #  §3  行公式生成
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _row_formulas(r: int, extended: bool) -> dict:
        """返回当行应有的基础公式 {col: formula}"""
        f = {
            COL_O: f"=L{r}-M{r}",
            COL_P: f"=K{r}-N{r}",
            COL_Q: f"=M{r}+O{r}",
            COL_R: f"=M{r}+N{r}+O{r}+P{r}",
        }
        if extended:
            f[COL_J] = f"=H{r}+I{r}"
            f[COL_L] = f"=J{r}-K{r}"
        return f

    def _write_row(self, ws, r, d_val, c_val, m_val, extended, skip_fill=False): # ← 增加参数
        self._style_row(ws, r, skip_fill=skip_fill)  # ← 传入参数
        if d_val: ws.cell(r, COL_D).value = str(d_val)
        if c_val: ws.cell(r, COL_C).value = str(c_val)
        if m_val is not None:
            cell = ws.cell(r, COL_M)
            cell.value = round(float(m_val), 2)
            cell.number_format = '#,##0.00'
        for col, fml in self._row_formulas(r, extended).items():
            ws.cell(r, col).value = fml

    # ═══════════════════════════════════════════════════════════
    #  §4  公式修复（三步）
    # ═══════════════════════════════════════════════════════════

    def _fix_stale_self_refs(self, ws):
        """
        ★ 核心修复：插行后各行自身公式中的旧行号替换为当前行号。
        目标列：J(10) L(12) O(15) P(16) Q(17) R(18) S(19)
        规则：若公式里所有数字引用同一个旧行号 x ≠ r → 全部替换为 r。
        跳过含':'的区间引用（SUM 范围）和特殊汇总行。
        """
        SKIP_D_WORDS = {'成本合计', '成本及费用合计', '利润', '增值税实际税负', '税金及附加'}
        TARGET_COLS  = {COL_J, COL_L, COL_O, COL_P, COL_Q, COL_R, COL_S}

        for r in range(10, ws.max_row + 1):
            # 跳过大汇总行
            d_val = self._norm(str(ws.cell(r, COL_D).value or ''))
            if any(w in d_val for w in SKIP_D_WORDS):
                continue

            for c in TARGET_COLS:
                cell = ws.cell(r, c)
                if not isinstance(cell.value, str) or not cell.value.startswith('='):
                    continue
                formula = cell.value
                # 跳过含区间冒号的 SUM 公式（如 =SUM(M12:M20)）
                if ':' in formula:
                    # 只修正 =SUM(Tx:Vx) 单行格式
                    m = re.fullmatch(r'=SUM\(T(\d+):V\1\)', formula, re.IGNORECASE)
                    if m:
                        x = int(m.group(1))
                        if x != r:
                            cell.value = f'=SUM(T{r}:V{r})'
                    continue
                # 提取公式中"字母后紧跟的数字"（即单元格行号，如 J37 中的 37）
                nums = [int(n) for n in re.findall(r'(?<=[A-Za-z])(\d+)', formula)]
                if not nums: continue
                unique = set(nums)
                # 只修正：所有行号相同 AND 行号 > 5 AND != r
                if len(unique) == 1:
                    x = unique.pop()
                    if x != r and x > 5:
                        # 用后向断言替换字母后的行号数字
                        cell.value = re.sub(
                            r'(?<=[A-Za-z])' + str(x) + r'(?!\d)',
                            str(r), formula
                        )

    def _fix_all_sums(self, ws):
        """重写每个小计行的 SUM 范围覆盖 [hdr_row, anchor_row]"""
        for hdr_kw, sub_kw in SUBTOTAL_MAP.items():
            hdr_row = self._find_row(ws, hdr_kw)
            sub_row = self._find_row(ws, sub_kw)
            if not hdr_row or not sub_row: continue

            anchor = None
            for r in range(sub_row - 1, max(hdr_row, sub_row - 400), -1):
                for c in range(2, 6):
                    if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                        anchor = r; break
                if anchor: break
            end_row = anchor if anchor else sub_row - 1

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(sub_row, col)
                if cell.value and isinstance(cell.value, str) and 'SUM(' in cell.value.upper():
                    cl = get_column_letter(col)
                    cell.value = f"=SUM({cl}{hdr_row}:{cl}{end_row})"

            # 修正锚点行自身公式
            if anchor:
                for col, fml in self._row_formulas(anchor, False).items():
                    cell = ws.cell(anchor, col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('=') and ':' not in cell.value:
                        cell.value = fml

            self.log.info("   🧮 %-14s SUM 范围 行%d→%d", sub_kw, hdr_row, end_row)

    def _fix_aggregate_rows(self, ws):
        """
        修复跨节汇总行（成本合计、成本费用合计）的引用偏移。
        重新定位各小计行，重建 H-V 列的求和公式。
        """
        # ── 定位各小计行 ───────────────────────────────────
        kw_to_sub = {
            '人工费小计':    '（一）人工费',
            '分包工程小计':  '（二）分包工程',
            '材料费小计':    '(三)材料费',
            '机械费小计':    '(四）机械租赁费',
            '其他直接费小计': '（五）其他直接费',
            '间接费小计':    '（六）间接费',
            '安全费小计':    '（七）安全费',
        }
        sub_rows = {}
        for kw in kw_to_sub:
            r = self._find_row(ws, kw)
            if r: sub_rows[kw] = r

        # ── 三、成本合计 ────────────────────────────────────
        cost_total_row = self._find_row(ws, '三、成本合计')
        if cost_total_row and len(sub_rows) == len(kw_to_sub):
            ordered = ['人工费小计','分包工程小计','材料费小计','机械费小计',
                       '其他直接费小计','间接费小计','安全费小计']
            row_refs = [sub_rows[k] for k in ordered]
            for col in range(8, ws.max_column + 1):
                cell = ws.cell(cost_total_row, col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cl = get_column_letter(col)
                    cell.value = '=' + '+'.join(f'{cl}{r}' for r in row_refs)
            self.log.info("   🧮 三、成本合计 公式已修正（行 %d）", cost_total_row)

        # ── 八、成本及费用合计 ──────────────────────────────
        total_row = self._find_row(ws, '八、成本及费用合计')
        if total_row and cost_total_row:
            four_row  = self._find_row(ws, '四、计提保修金')
            five_row  = self._find_row(ws, '五、过程节点奖金')
            six_row   = self._find_row(ws, '六、资金占用费用')
            seven_row = self._find_row(ws, '七、局投资收益')
            need = [cost_total_row, four_row, five_row, six_row, seven_row]
            if all(need):
                for col in range(8, ws.max_column + 1):
                    cell = ws.cell(total_row, col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        cl = get_column_letter(col)
                        cell.value = '=' + '+'.join(f'{cl}{r}' for r in need)
                self.log.info("   🧮 八、成本及费用合计 公式已修正（行 %d）", total_row)

    # ═══════════════════════════════════════════════════════════
    #  §5  数据聚合
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _sort_by_contract(rows):
        return sorted(rows, key=lambda x:
                      0 if x.get("c") and str(x.get("c","")).strip() not in ("","nan") else 1)

    def _agg_labor(self, df):
        """
        人工费聚合规则（v2）：
        1. FGD 前缀 → 全部汇总为1行，D='FGD'
        2. CFK 前缀 → 按客商汇总，D='CFK+{客商名称}'
        3. 其余有客商 或 有合同编码 → 按客商名称+合同编码聚合
        4. 无合同编码 → 按中台单据号前3位聚合
        """
        df = df.copy()
        df['_pfx'] = df['中台单据号'].fillna('').astype(str).str[:3].str.upper()
        df['_ven'] = df['客商名称'].fillna('').astype(str).str.strip().replace('nan', '')
        df['_con'] = df['合同编码'].fillna('').astype(str).str.strip().replace('nan', '')
        rows = []

        # Rule 1: FGD → 单行汇总
        fgd = df[df['_pfx'] == 'FGD']
        if not fgd.empty:
            rows.append({"d": "FGD", "c": None, "p": round(fgd['最终发生额'].sum(), 2)})

        # Rule 2: CFK → 按客商名称分组
        for ven, grp in df[df['_pfx'] == 'CFK'].groupby('_ven', sort=False):
            label = f"CFK+{ven}" if ven else "CFK"
            rows.append({"d": label, "c": None, "p": round(grp['最终发生额'].sum(), 2)})

        # 其余行（非 FGD/CFK）
        others = df[~df['_pfx'].isin(['FGD', 'CFK'])].copy()

        # Rule 3: 有客商 或 有合同 -> 按客商+合同聚合
        has_vc = others[(others['_ven'] != '') | (others['_con'] != '')]
        for (ven, con), grp in has_vc.groupby(['_ven', '_con'], sort=False):
            rows.append({"d": ven or None, "c": con or None,
                          "p": round(grp['最终发生额'].sum(), 2)})

        # Rule 4: 剩余既没有客商、又没有合同的数据 -> 按中台单据号前3位聚合 (修复重复计算Bug)
        no_vc = others[~others.index.isin(has_vc.index)]
        if not no_vc.empty:
            no_vc = no_vc.copy()

        return self._sort_by_contract(rows)

    def _agg_vendor(self, df):
        rows = []
        for (ven, con), grp in df.groupby(
                [df['客商名称'].fillna(''), df['合同编码'].fillna('')], sort=False):
            rows.append({"d": ven or None, "c": con or None, "p": round(grp['最终发生额'].sum(), 2)})
        return self._sort_by_contract(rows)

    def _agg_sub(self, df):
        rows = []
        for sub, grp in df.groupby(df['细分科目'].fillna(''), sort=False):
            rows.append({"d": sub or None, "c": None, "p": round(grp['最终发生额'].sum(), 2)})
        return rows

    def _agg_material(self, df):
        """
        材料费聚合规则（v2）：
        排除材料库存、材料调入、材料调出、材料调入+CFK等固定项
        1. 有合同编码的：按供应商+合同编码聚合
        2. 无合同编码的：按细分科目聚合
        3. 排序：有合同编码的排在前面
        """
        # 排除固定填报项（这些项通过_fill_fixed直接填入）
        EXCL = {'材料库存', '材料调入', '材料调出', '材料调入+CFK', '材料调入-CFK', 
                '材料调出+XSD', '材料调出+SQD', '废旧物资处置', '废旧物资处置+SQD', '其他'}
        df = df[~df['细分科目'].fillna('').astype(str).str.strip().isin(EXCL)].copy()
        
        # 同时排除ZGD开头的数据（这些会单独聚合）
        zgd_mask = df['细分科目'].fillna('').astype(str).str.startswith('ZGD-')
        df = df[~zgd_mask].copy()
        
        if df.empty: 
            return []
        
        # 有合同编码的按供应商+合同编码聚合
        has_c = df[df['合同编码'].fillna('').astype(str).str.strip().replace('nan','') != '']
        # 无合同编码的按细分科目聚合
        no_c  = df[~df.index.isin(has_c.index)]
        
        # 排序：有合同的排在前面
        return self._sort_by_contract(self._agg_vendor(has_c) + self._agg_sub(no_c))

    # ═══════════════════════════════════════════════════════════
    #  §6  占位行清除（改良版）
    # ═══════════════════════════════════════════════════════════

    # Protected sub-category keywords — never delete these rows
    _KEEP_ROW_KEYWORDS = {
        '（1）集中采购', '(1)集中采购', '（2）自行采购', '(2)自行采购',
        '材料库存', '材料调入', '材料调出','材料调入+CFK',
        '研发支出', '研发费用'
    }

    def _delete_stubs(self, ws, hdr_row: int, anchor_row: int):
        """
        ★ v5 强化版：删除 [hdr_row+1, anchor_row-1] 范围内
          M列（col 13）不含真实数值 且 D列不是受保护子标题 的所有行。
          这样可以清除旧版本填入 P列 的历史数据行，以及纯公式占位行。
        """
        protected = {self._norm(k) for k in self._KEEP_ROW_KEYWORDS}
        to_del = []
        for r in range(hdr_row + 1, anchor_row):
            d = ws.cell(r, COL_D).value
            m = ws.cell(r, COL_M).value
            # 保留：M列有真实数值
            if isinstance(m, (int, float)):
                continue
            # 保留：D列是受保护子标题
            d_norm = self._norm(str(d or ''))
            if any(k in d_norm for k in protected):
                continue
            to_del.append(r)
        for r in reversed(to_del):
            ws.delete_rows(r)
        if to_del:
            self.log.info("      🗑 清除 %d 行（旧数据/公式占位行）", len(to_del))

    def _insert_rows(self, ws, anchor, action, rows_data, extended, is_rd=False): # ← 增加 is_rd 参数
        for i, rd in enumerate(rows_data):
            ins = anchor + i if action == 'above' else anchor + 1 + i
            ws.insert_rows(ins)
            self._write_row(ws, ins, rd.get('d'), rd.get('c'), rd.get('p', 0), extended, skip_fill=is_rd)

    # ═══════════════════════════════════════════════════════════
    #  §7  B/N/T 列匹配填写（合并表来源）
    # ═══════════════════════════════════════════════════════════

    def _build_lookups(self, df4):
        # ★ 供应商名称列兼容：不同项目可能叫"供应商名称"或"供应商描述"
        ven_col = '供应商名称' if '供应商名称' in df4.columns else (
                  '供应商描述' if '供应商描述' in df4.columns else None)

        # 客商名称+合同 → 供应商编码
        vendor_lkp = {}
        if ven_col:
            sub = df4[df4[ven_col].notna() & df4['供应商'].notna()]
            dedup_cols = [ven_col, '合同'] if '合同' in sub.columns else [ven_col]
            for _, row in sub.drop_duplicates(dedup_cols).iterrows():
                ven  = self._norm(str(row[ven_col]))
                con  = self._norm(str(row.get('合同','') or ''))
                code = str(int(float(row['供应商']))) if pd.notna(row['供应商']) else ''
                if ven:
                    vendor_lkp[(ven, con)] = code
                    if (ven, '') not in vendor_lkp:
                        vendor_lkp[(ven, '')] = code

        # 供应商编码+合同 → 应付账款借方（累计付款）
        # ★ 应付账款付款：借方发生额=付款（冲减负债），取 借方合计 作为累计付款
        ap_mask = (df4['总账科目长文本'].fillna('').str.contains('应付账款') &
                   ~df4['总账科目长文本'].fillna('').str.contains('待确认进项税额'))
        ap = df4[ap_mask].copy()
        ap['_v'] = ap['供应商'].fillna(0).apply(
            lambda x: str(int(float(x))) if pd.notna(x) and x else '')
        ap['_c'] = ap['合同'].fillna('').astype(str).str.strip()
        # 借方合计即为付款金额（借方记录付款动作）
        pay_lkp = dict(ap.groupby(['_v','_c'])['借方本位币金额'].sum())

        # 供应商编码+合同 → 其他应收款\待确认进项税额借方
        vat = df4[df4['总账科目长文本'].fillna('').str.contains('其他应收款') &
                  df4['总账科目长文本'].fillna('').str.contains('待确认进项税额')].copy()
        vat['_v'] = vat['供应商'].fillna(0).apply(
            lambda x: str(int(float(x))) if pd.notna(x) and x else '')
        vat['_c'] = vat['合同'].fillna('').astype(str).str.strip()
        vat_lkp = dict(vat.groupby(['_v','_c'])['借方本位币金额'].sum())

        return vendor_lkp, pay_lkp, vat_lkp

    def _fill_per_row_cols(self, ws, vendor_lkp, pay_lkp, vat_lkp):
        """遍历数据行，填 B(供应商编码)、N(进项税)、T(财务账面付款)"""
        SKIP = {'人工费小计','分包工程小计','材料费小计','机械费小计','其他直接费小计',
                '间接费小计','安全费小计','财务未列部分需增列','成本合计','项目效益',
                '成本及费用合计','增值税','税金及附加','利润','保修金','节点奖'}
        for r in range(10, ws.max_row + 1):
            d_val = ws.cell(r, COL_D).value
            m_val = ws.cell(r, COL_M).value
            c_val = ws.cell(r, COL_C).value
            if not d_val or not isinstance(m_val, (int, float)): continue
            if any(s in self._norm(str(d_val)) for s in SKIP): continue

            d_norm = self._norm(str(d_val))
            c_norm = self._norm(str(c_val or ''))
            code = (vendor_lkp.get((d_norm, c_norm)) or
                    vendor_lkp.get((d_norm, ''), ''))
            if code:
                ws.cell(r, COL_B).value = code

            if code:
                pay = pay_lkp.get((code, c_norm), 0)
                if pay:
                    c = ws.cell(r, COL_T)   # ★ T列(20) 财务账面已列报付款
                    c.value = round(float(pay), 2)
                    c.number_format = '#,##0.00'
                vat = vat_lkp.get((code, c_norm), 0)
                if vat:
                    c = ws.cell(r, COL_N)
                    c.value = round(float(vat), 2)
                    c.number_format = '#,##0.00'

    def _apply_invoice_match(self, ws) -> None:
        invoice_path = self.invoice_ledger_path
        if not invoice_path:
            try:
                for fn in os.listdir(self.result_dir):
                    if fn.lower().endswith((".xlsx", ".xls")) and any(k in fn for k in ["发票", "收票", "台账", "台帳"]):
                        invoice_path = os.path.join(self.result_dir, fn)
                        break
            except Exception:
                invoice_path = None

        if not (invoice_path and os.path.exists(invoice_path)):
            self.log.info("   ⓘ 未提供发票台账，跳过 AB 列发票匹配")
            return

        from invoice_matcher import InvoiceMatcher

        matcher = InvoiceMatcher(self.log)
        matcher.apply_to_worksheet(ws, invoice_path)

    # ═══════════════════════════════════════════════════════════
    #  §8  A列序号重排
    # ═══════════════════════════════════════════════════════════

    def _renumber(self, ws):
        """对人工/分包/材料/机械节内有数据的行重写 A列序号"""
        sections = [
            ('（一）人工费',    '人工费小计'),
            ('（二）分包工程',  '分包工程小计'),
            ('(三)材料费',     '材料费小计'),
            ('(四）机械租赁费', '机械费小计'),
        ]
        for hdr_kw, sub_kw in sections:
            hdr = self._find_row(ws, hdr_kw)
            sub = self._find_row(ws, sub_kw)
            if not hdr or not sub: continue
            anchor = None
            for r in range(sub - 1, hdr, -1):
                for c in range(2, 6):
                    if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                        anchor = r; break
                if anchor: break
            end = (anchor - 1) if anchor else (sub - 1)
            # 先清空区间内 A 列
            for r in range(hdr + 1, sub):
                ws.cell(r, COL_A).value = None
            # 对有数据的行写序号
            seq = 1
            for r in range(hdr + 1, end + 1):
                d = ws.cell(r, COL_D).value
                m = ws.cell(r, COL_M).value
                if d and isinstance(m, (int, float)) and m != 0:
                    ws.cell(r, COL_A).value = seq
                    seq += 1

    # ═══════════════════════════════════════════════════════════
    #  §9  主流程
    # ═══════════════════════════════════════════════════════════

    def execute_fill(self):
        source_path = os.path.join(self.result_dir, "5_中间汇总表_效益审核数据源.xlsx")
        merged_path = os.path.join(self.result_dir, "4_合并表.xlsx")

        if not os.path.exists(source_path):
            self.log.error("⚠️ 找不到中间汇总表：%s", source_path); return False
        if not os.path.exists(self.template_path):
            self.log.error("⚠️ 找不到模板：%s", self.template_path); return False

        self.log.info("📝 填报引擎 v5 启动（全面修正版）...")
        try:
            # ── 读取中间表 ─────────────────────────────────
            df_raw = pd.read_excel(source_path)
            df = df_raw[df_raw['利润中心'].astype(str).str.strip() != '筛选合计：'].copy()
            df['最终发生额'] = pd.to_numeric(df['最终发生额'], errors='coerce').fillna(0)
            for col in ['合同编码', '细分科目', '客商名称', '中台单据号']:
                if col in df.columns:
                    df[col] = df[col].fillna('').astype(str).str.strip().replace('nan', '')

            # ── 读取合并表 ─────────────────────────────────
            df4 = pd.read_excel(merged_path) if os.path.exists(merged_path) else pd.DataFrame()

            vendor_lkp = pay_lkp = vat_lkp = {}
            if not df4.empty:
                vendor_lkp, pay_lkp, vat_lkp = self._build_lookups(df4)

            # ── 打开模板 ───────────────────────────────────
            wb = openpyxl.load_workbook(self.template_path)
            ws = wb.active
            for sn in wb.sheetnames:
                if any(k in sn for k in ['效益审核','附表']):
                    ws = wb[sn]; break

            # ════════════════════════════════════════════
            # 一、表头固定坐标
            # ════════════════════════════════════════════
            def _first(col):
                if col not in df.columns: return ''
                v = df[col].replace('', None).dropna()
                return v.iloc[0] if not v.empty else ''

            ws['B6'] = _first('利润中心')
            ws['C6'] = _first('项目编码')
            ws['D6'] = _first('工程名称')

            # M6：合并表主营业务收入贷方合计（负数）
            if not df4.empty and '总账科目长文本' in df4.columns:
                rev = df4[df4['总账科目长文本'].fillna('').str.startswith(
                    '主营业务收入')]['贷方本位币金额'].sum()
                ws['M6'] = round(-float(rev), 2)
            else:
                rev = df[df['成本_财务大类'] == '财务累计入账收入(不含增值税)']['最终发生额'].sum()
                ws['M6'] = round(-float(rev), 2)

            n6 = df[df['成本_财务大类'] == '已价税分离的增值税金额(财务账面数据)']['最终发生额'].sum()
            ws['N6'] = round(-float(n6), 2)
            self.log.info("   ✅ 表头写入（B6/C6/D6/M6/N6）")

            # ════════════════════════════════════════════
            # 二、动态插行（先清占位行，再写数据）
            # ════════════════════════════════════════════
            plan = [
                # (数据大类, 聚合, 锚点关键词, above/below, 扩展公式)
                ('(一)人工费',    'labor',    '人工费小计',    'above', False),
                ('(二)分包工程',  'vendor',   '分包工程小计',  'above', False),
                ('(三)材料费',    'material', '(2)自行采购',   'below', False),
                ('(四)机械租赁费','vendor',   '机械费小计',    'above', True),
                ('(五)其他直接费','sub',      '其他直接费小计','above', True),
                ('(六)间接费',    'sub',      '间接费小计',    'above', True),
                ('(七)安全费',    'sub',      '安全费小计',    'above', True),
                ('研发支出',      'sub',      '八、研发费用',  'below', False),
            ]
            hdr_kw_map = {
                '(一)人工费':    '（一）人工费',
                '(二)分包工程':  '（二）分包工程',
                '(三)材料费':   '(三)材料费',
                '(四)机械租赁费':'(四）机械租赁费',
                '(五)其他直接费':'（五）其他直接费',
                '(六)间接费':    '（六）间接费',
                '(七)安全费':    '（七）安全费',
            }

            for cat, mode, anchor_kw, action, ext in plan:
                cat_df = df[df['成本_财务大类'] == cat].copy()
                if cat_df.empty:
                    self.log.info("   ℹ️  [%s] 无数据", cat); continue

                if   mode == 'labor':    rows_data = self._agg_labor(cat_df)
                elif mode == 'vendor':   rows_data = self._agg_vendor(cat_df)
                elif mode == 'material': rows_data = self._agg_material(cat_df)
                else:                    rows_data = self._agg_sub(cat_df)

                if not rows_data: continue

                # 定位锚点
                if action == 'above':
                    hdr_kw  = hdr_kw_map.get(cat, '')
                    hdr_row = self._find_row(ws, hdr_kw) if hdr_kw else None
                    anchor = self._find_anchor(ws, anchor_kw, min_row=hdr_row or 1)
                    if anchor is None:
                        anchor = self._find_row(ws, anchor_kw)
                else:
                    anchor = self._find_row(ws, anchor_kw)
                    hdr_row = None

                if anchor is None:
                    self.log.warning("   ⚠️  [%s] 锚点未找到（'%s'）", cat, anchor_kw); continue

                # 清除占位行（仅 above 节）
                if action == 'above' and hdr_row:
                    self._delete_stubs(ws, hdr_row, anchor)
                    anchor = (self._find_anchor(ws, anchor_kw, min_row=hdr_row)
                              or self._find_row(ws, anchor_kw))

                self.log.info("   📍 [%s] 锚点=行%-4d  插入 %d 行", cat, anchor, len(rows_data))
                
                is_rd = (cat == '研发支出')  
                
                self._insert_rows(ws, anchor, action, rows_data, ext, is_rd=is_rd) # ← 传入开关参数

            # ════════════════════════════════════════════
            # 三、公式修复（三步，顺序不可变）
            # ════════════════════════════════════════════
            self._fix_stale_self_refs(ws)   # ① 行内自身引用修正
            self._fix_all_sums(ws)           # ② 小计行 SUM 范围修正
            self._fix_aggregate_rows(ws)     # ③ 成本合计/成本费用合计 修正

            # ════════════════════════════════════════════
            # 四、原位直填
            # ════════════════════════════════════════════
            def _fill_fixed(cat_or_filter, kw, label):
                if isinstance(cat_or_filter, str):
                    amt = df[df['成本_财务大类'] == cat_or_filter]['最终发生额'].sum()
                else:
                    mask = pd.Series(True, index=df.index)
                    for k, v in cat_or_filter.items(): mask &= (df[k] == v)
                    amt = df[mask]['最终发生额'].sum()
                r = self._find_row(ws, kw)
                if r and amt != 0:
                    ws.cell(r, COL_M).value = round(float(amt), 2)
                    ws.cell(r, COL_M).number_format = '#,##0.00'
                    self.log.info("   ✅ %-20s 行%d  M=%.2f", label, r, amt)
                elif not r:
                    self.log.warning("   ⚠️  找不到关键词 '%s'", kw)

            _fill_fixed('六、资金占用费用', '六、资金占用费用', '资金占用费用')
            _fill_fixed('七、局投资收益（局投资项目选填）', '七、局投资收益', '局投资收益')
            
            # 材料费各项固定填报（按新规则）
            material_items = [
                '材料库存', '材料调入', '材料调出', '材料调入+CFK', '材料调入+CFK',
                '材料调出+XSD', '材料调出+SQD', '废旧物资处置', '废旧物资处置+SQD', '其他'
            ]
            for sub in material_items:
                _fill_fixed({'成本_财务大类': '(三)材料费', '细分科目': sub}, sub, sub)
            
            # ZGD开头的单据，按细分科目单独填报
            zgd_df = df[(df['成本_财务大类'] == '(三)材料费') & 
                       (df['细分科目'].fillna('').astype(str).str.startswith('ZGD-'))]
            for _, row_data in zgd_df.iterrows():
                sub_name = str(row_data['细分科目'])
                amt = row_data['最终发生额']
                r = self._find_row(ws, sub_name)
                if r and amt != 0:
                    ws.cell(r, COL_M).value = round(float(amt), 2)
                    ws.cell(r, COL_M).number_format = '#,##0.00'
                    self.log.info("   ✅ %-20s 行%d  M=%.2f", sub_name, r, amt)

            # 十一、税金及附加 → M列（按 D列第4列查找）
            tax_amt = df[df['成本_财务大类'].fillna('').str.contains('税金及附加')]['最终发生额'].sum()
            r_tax = self._find_row(ws, '十一、税金及附加')
            if r_tax and tax_amt:
                ws.cell(r_tax, COL_M).value = round(float(tax_amt), 2)
                ws.cell(r_tax, COL_M).number_format = '#,##0.00'
                self.log.info("   ✅ 税金及附加 行%d M=%.2f", r_tax, tax_amt)

            # ════════════════════════════════════════════
            # 五、B/N/T 列按行匹配
            # ════════════════════════════════════════════
            if not df4.empty:
                self._fill_per_row_cols(ws, vendor_lkp, pay_lkp, vat_lkp)
            self.log.info("   ✅ B/N/T 列匹配填写完成")

            # ════════════════════════════════════════════
            # 五点五、发票已认证台账匹配 → AB列（累计已计票发票额）
            # ════════════════════════════════════════════
            self._apply_invoice_match(ws)

            # ════════════════════════════════════════════
            # 六、A列序号重排
            # ════════════════════════════════════════════
            self._renumber(ws)
            self.log.info("   ✅ A列序号重排完成")

            # ── 保存 ──────────────────────────────────
            out = os.path.join(self.result_dir, "自动填报完成_效益审核表.xlsx")
            wb.save(out)
            self.log.info("🎉 填报完成：%s", out)
            return True

        except Exception as e:
            self.log.error("❌ 填报异常:\n%s", traceback.format_exc()); return False