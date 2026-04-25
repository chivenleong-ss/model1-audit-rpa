# -*- coding: utf-8 -*-
"""
benefit_reporter.py  ── v3 (完全修正版)
核心修复：
  1. 所有金额先按规则汇总合计，再按行插入（原版逐笔插入已修正）
  2. 锚点关键词修正（机械费小计 / 财务未列部分需增列的 * 号处理）
  3. 关键词统一搜索列 B-E (openpyxl col 2-5)，适配模板合并单元格
  4. 材料库存/调入/调出原位直填到模板专属行，不参与动态插行
"""

import os
import traceback
import pandas as pd
import openpyxl
from openpyxl.styles import Border, Side, Font, Alignment


class BenefitReporter:
    """智能填报引擎：先聚合，再精准插行，零损耗填入《项目效益审核表》"""

    # ── 模板关键词（与实际 Excel 模板中 C 列文字一致，去符号后匹配）──
    FILL_COL = 13   # M 列（金额写入列）
    VENDOR_COL = 4  # D 列（客商/细分名称）
    CONTRACT_COL = 3  # C 列（合同编码）

    def __init__(self, result_dir, logger):
        self.result_dir = result_dir
        self.log = logger
        self.template_path = "项目效益审核表.xlsx"

    # ────────────────────────────────────────────────────────────
    #   工具函数
    # ────────────────────────────────────────────────────────────

    @staticmethod
    def _norm(val) -> str:
        """标准化：去空格、星号，全角→半角括号/冒号，便于模糊匹配"""
        if val is None:
            return ""
        return (
            str(val)
            .replace(" ", "").replace("\u3000", "")
            .replace("（", "(").replace("）", ")")
            .replace("：", ":").replace("*", "")
            .replace(":", "")  # 统一去掉冒号，避免"小计："vs"小计"失配
        )

    def _find_row(self, ws, keyword, search_cols=(2, 3, 4, 5)) -> int | None:
        """在指定列中搜索包含 keyword（标准化后）的第一行，返回 1-based 行号"""
        kw = self._norm(keyword)
        if not kw:
            return None
        for r in range(1, ws.max_row + 1):
            for c in search_cols:
                if kw in self._norm(ws.cell(r, c).value):
                    return r
        return None

    def _find_row_above(self, ws, base_kw, target_kw,
                        search_cols=(2, 3, 4, 5)) -> int | None:
        """
        先找 base_kw 行，再从该行向上搜索 target_kw，返回 target_kw 所在行号。
        用于定位"财务未列部分需增列"（紧挨在各小计行的正上方）。
        """
        base_row = self._find_row(ws, base_kw, search_cols)
        if base_row is None:
            return None
        tgt = self._norm(target_kw)
        for r in range(base_row - 1, max(0, base_row - 50), -1):  # 最多向上找50行
            for c in search_cols:
                if tgt in self._norm(ws.cell(r, c).value):
                    return r
        return None

    @staticmethod
    def _make_style():
        thin = Side(border_style="thin", color="000000")
        return (
            Border(left=thin, right=thin, top=thin, bottom=thin),
            Font(name="宋体", size=10),
        )

    def _apply_row_style(self, ws, row_num: int, col_count: int = 20):
        border, font = self._make_style()
        for col in range(1, col_count + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.border = border
            cell.font = font
            cell.alignment = Alignment(
                horizontal="left" if col in (self.VENDOR_COL, self.CONTRACT_COL) else "center",
                vertical="center",
                wrap_text=(col in (self.VENDOR_COL, self.CONTRACT_COL)),
            )

    def _write_data_row(self, ws, row_num: int, c_val, d_val, p_val):
        """写入 D（客商/细分）、C（合同）、M（金额）三列"""
        self._apply_row_style(ws, row_num)
        if c_val:
            ws.cell(row=row_num, column=self.VENDOR_COL, value=str(c_val))
        if d_val:
            ws.cell(row=row_num, column=self.CONTRACT_COL, value=str(d_val))
        amt_cell = ws.cell(row=row_num, column=self.FILL_COL, value=round(float(p_val), 2))
        amt_cell.number_format = '#,##0.00'

    # ────────────────────────────────────────────────────────────
    #   聚合函数（所有金额先汇总，再写入）
    # ────────────────────────────────────────────────────────────

    @staticmethod
    def _clean_str(val) -> str:
        return str(val).strip() if pd.notna(val) else ""

    def _agg_by_vendor(self, df: pd.DataFrame) -> list[dict]:
        """按 客商名称 + 合同编码 聚合，适用于：人工费、分包工程、机械租赁费"""
        grp = (
            df.groupby(
                [df["客商名称"].fillna("").astype(str).str.strip(),
                 df["合同编码"].fillna("").astype(str).str.strip()],
                sort=False,
            )["最终发生额"]
            .sum()
            .reset_index()
        )
        grp.columns = ["客商名称", "合同编码", "金额"]
        rows = []
        for _, r in grp.iterrows():
            rows.append({
                "d": r["客商名称"] or None,
                "c": r["合同编码"] or None,
                "m": r["金额"],
            })
        return rows

    def _agg_by_sub(self, df: pd.DataFrame) -> list[dict]:
        """按 细分科目 聚合，适用于：其他直接费、间接费、安全费、研发支出"""
        grp = (
            df.groupby(
                df["细分科目"].fillna("").astype(str).str.strip(),
                sort=False,
            )["最终发生额"]
            .sum()
            .reset_index()
        )
        grp.columns = ["细分科目", "金额"]
        rows = []
        for _, r in grp.iterrows():
            rows.append({
                "d": r["细分科目"] or None,
                "c": None,
                "m": r["金额"],
            })
        return rows

    def _agg_material(self, df: pd.DataFrame) -> list[dict]:
        """
        材料费聚合（已排除库存/调入/调出）：
        • 有合同编码 → 按 客商+合同 聚合，D=客商，C=合同
        • 无合同编码 → 按 细分科目 聚合，D=细分科目
        """
        EXCLUDE = {"材料库存", "材料调入", "材料调出"}
        df = df[~df["细分科目"].fillna("").astype(str).str.strip().isin(EXCLUDE)].copy()
        if df.empty:
            return []

        has_contract = df[
            df["合同编码"].fillna("").astype(str).str.strip().replace("nan", "") != ""
        ]
        no_contract = df[~df.index.isin(has_contract.index)]

        rows = []
        if not has_contract.empty:
            rows.extend(self._agg_by_vendor(has_contract))
        if not no_contract.empty:
            rows.extend(self._agg_by_sub(no_contract))
        return rows

    # ────────────────────────────────────────────────────────────
    #   插行核心：在锚点前/后批量插入聚合数据行
    # ────────────────────────────────────────────────────────────

    def _insert_rows(self, ws, anchor_row: int, action: str, rows_data: list[dict]):
        """
        action='above': 在 anchor_row 上方逐行插入（数据行按顺序出现在锚点之前）
        action='below': 在 anchor_row 下方逐行插入
        """
        for i, rd in enumerate(rows_data):
            if action == "above":
                ins = anchor_row + i   # 每次插入后锚点下移，i 补偿偏移
            else:
                ins = anchor_row + 1 + i
            ws.insert_rows(ins)
            self._write_data_row(ws, ins, rd.get("d"), rd.get("c"), rd.get("m", 0))

    # ────────────────────────────────────────────────────────────
    #   主流程
    # ────────────────────────────────────────────────────────────

    def execute_fill(self):
        source_path = os.path.join(self.result_dir, "5_中间汇总表_效益审核数据源.xlsx")
        if not os.path.exists(source_path):
            self.log.error("⚠️ 找不到数据源：%s", source_path)
            return False
        if not os.path.exists(self.template_path):
            self.log.error("⚠️ 找不到模板：%s", self.template_path)
            return False

        self.log.info("📝 启动智能填报引擎 v3（聚合版）...")
        try:
            # ── 读取并清洗数据源 ──────────────────────────────────
            df_raw = pd.read_excel(source_path)
            df = df_raw[
                df_raw["利润中心"].astype(str).str.strip() != "筛选合计："
            ].copy()
            df["最终发生额"] = pd.to_numeric(df["最终发生额"], errors="coerce").fillna(0)
            for col in ["合同编码", "细分科目", "客商名称"]:
                if col in df.columns:
                    df[col] = df[col].fillna("").astype(str).str.strip().replace("nan", "")

            # ── 打开模板 ──────────────────────────────────────────
            wb = openpyxl.load_workbook(self.template_path)
            ws = wb.active
            for sn in wb.sheetnames:
                if any(k in sn for k in ["效益审核", "附表"]):
                    ws = wb[sn]
                    break

            # ════════════════════════════════════════════════════
            # 一、表头固定坐标直填
            # ════════════════════════════════════════════════════
            def _first_val(col):
                sub = df[df[col].astype(str).str.strip() != ""] if col in df.columns else df
                return sub[col].iloc[0] if not sub.empty else ""

            ws["B6"] = _first_val("利润中心")
            ws["C6"] = _first_val("项目编码")
            ws["D6"] = _first_val("工程名称")

            rev = df[df["成本_财务大类"] == "财务累计入账收入(不含增值税)"]["最终发生额"].sum()
            tax = df[df["成本_财务大类"] == "已价税分离的增值税金额(财务账面数据)"]["最终发生额"].sum()
            ws["M6"] = round(float(rev), 2) if rev else 0
            ws["N6"] = round(float(tax), 2) if tax else 0
            self.log.info("   ✅ 表头信息写入完毕")

            # ════════════════════════════════════════════════════
            # 二、原位直填（不插行，找关键词行→写 M 列）
            # ════════════════════════════════════════════════════
            def _fixed_fill(cat_filter: dict | str, row_kw: str, label: str):
                """汇总指定类别金额，找行写 M 列"""
                if isinstance(cat_filter, str):
                    amt = df[df["成本_财务大类"] == cat_filter]["最终发生额"].sum()
                else:
                    mask = pd.Series([True] * len(df), index=df.index)
                    for col, val in cat_filter.items():
                        mask &= (df[col] == val)
                    amt = df[mask]["最终发生额"].sum()

                if amt == 0:
                    self.log.info("   ℹ️ [%s] 金额为0，跳过", label)
                    return
                r = self._find_row(ws, row_kw)
                if r:
                    ws.cell(row=r, column=self.FILL_COL, value=round(float(amt), 2)).number_format = "#,##0.00"
                    self.log.info("   ✅ %-12s → 行 %-5d M列 = %.2f", label, r, amt)
                else:
                    self.log.warning("   ⚠️ [%s] 在模板中找不到关键词'%s'", label, row_kw)

            # 资金占用费用（合计）
            _fixed_fill("六、资金占用费用", "六、资金占用费用", "资金占用费用")

            # 局投资收益（若有）
            _fixed_fill("七、局投资收益（局投资项目选填）", "七、局投资收益", "局投资收益")

            # 材料库存 / 材料调入 / 材料调出 —— 各自找模板专属行
            for sub_kw in ["材料库存", "材料调入", "材料调出"]:
                _fixed_fill(
                    {"成本_财务大类": "(三)材料费", "细分科目": sub_kw},
                    sub_kw,
                    sub_kw,
                )

            # ════════════════════════════════════════════════════
            # 三、动态插行（先聚合，按规则定位锚点，批量插入）
            # 注意：所有插行需按模板从上到下顺序执行，
            #       否则行号已偏移会导致后续锚点找错
            # ════════════════════════════════════════════════════

            insert_plan = [
                # ── (一) 人工费 ─────────────────────────────────────────
                # 锚点：人工费小计 上方的"财务未列部分需增列"，在其上方插行
                # 写入：客商→D，合同→C，金额→M
                {
                    "label": "(一)人工费",
                    "cat":   "(一)人工费",
                    "base_kw":  "人工费小计",
                    "tgt_kw":   "财务未列部分需增列",
                    "action":   "above",
                    "agg":      "vendor",
                },
                # ── (二) 分包工程 ────────────────────────────────────────
                {
                    "label": "(二)分包工程",
                    "cat":   "(二)分包工程",
                    "base_kw":  "分包工程小计",
                    "tgt_kw":   "财务未列部分需增列",
                    "action":   "above",
                    "agg":      "vendor",
                },
                # ── (三) 材料费 ──────────────────────────────────────────
                # 锚点："（2）自行采购" 行，在其下方插行
                # 写入：客商→D，合同→C，金额→M；无合同时细分科目→D
                {
                    "label": "(三)材料费",
                    "cat":   "(三)材料费",
                    "base_kw":  "2)自行采购",   # 去括号后的标准化结果
                    "tgt_kw":   None,
                    "action":   "below",
                    "agg":      "material",
                },
                # ── (四) 机械租赁费 ──────────────────────────────────────
                # 注意：模板内关键词是"机械费小计"，非"机械租赁费小计"
                {
                    "label": "(四)机械租赁费",
                    "cat":   "(四)机械租赁费",
                    "base_kw":  "机械费小计",
                    "tgt_kw":   "财务未列部分需增列",
                    "action":   "above",
                    "agg":      "vendor",
                },
                # ── (五) 其他直接费 ──────────────────────────────────────
                # 按细分科目聚合写入 D 列，金额→M
                {
                    "label": "(五)其他直接费",
                    "cat":   "(五)其他直接费",
                    "base_kw":  "其他直接费小计",
                    "tgt_kw":   "财务未列部分需增列",
                    "action":   "above",
                    "agg":      "sub",
                },
                # ── (六) 间接费 ──────────────────────────────────────────
                # 按细分科目聚合写入 D 列，金额→M
                {
                    "label": "(六)间接费",
                    "cat":   "(六)间接费",
                    "base_kw":  "间接费小计",
                    "tgt_kw":   "财务未列部分需增列",
                    "action":   "above",
                    "agg":      "sub",
                },
                # ── (七) 安全费 ──────────────────────────────────────────
                # 按细分科目聚合写入 D 列，金额→M
                {
                    "label": "(七)安全费",
                    "cat":   "(七)安全费",
                    "base_kw":  "安全费小计",
                    "tgt_kw":   None,
                    "action":   "above",
                    "agg":      "sub",
                },
                # ── 八、研发费用 ─────────────────────────────────────────
                # 按细分科目聚合，在关键词行下方插行
                {
                    "label": "八、研发费用",
                    "cat":   "研发支出",
                    "base_kw":  "八、研发费用",
                    "tgt_kw":   None,
                    "action":   "below",
                    "agg":      "sub",
                },
            ]

            for rule in insert_plan:
                cat_df = df[df["成本_财务大类"] == rule["cat"]].copy()
                if cat_df.empty:
                    self.log.info("   ℹ️ [%s] 中间表无数据，跳过", rule["label"])
                    continue

                # 定位锚点行
                if rule["tgt_kw"]:
                    anchor = self._find_row_above(ws, rule["base_kw"], rule["tgt_kw"])
                    if anchor is None:
                        self.log.warning(
                            "   ⚠️ [%s] 未找到'%s'→退到'%s'直接定位",
                            rule["label"], rule["tgt_kw"], rule["base_kw"],
                        )
                        anchor = self._find_row(ws, rule["base_kw"])
                else:
                    anchor = self._find_row(ws, rule["base_kw"])

                if anchor is None:
                    self.log.warning(
                        "   ⚠️ [%s] 在模板中找不到关键词'%s'，跳过",
                        rule["label"], rule["base_kw"],
                    )
                    continue

                # 聚合数据
                agg = rule["agg"]
                if agg == "vendor":
                    rows_data = self._agg_by_vendor(cat_df)
                elif agg == "sub":
                    rows_data = self._agg_by_sub(cat_df)
                elif agg == "material":
                    rows_data = self._agg_material(cat_df)
                else:
                    rows_data = self._agg_by_vendor(cat_df)

                self.log.info(
                    "   📍 [%s] 锚点=行%-5d 动作=%-5s 聚合后=%d行",
                    rule["label"], anchor, rule["action"], len(rows_data),
                )

                self._insert_rows(ws, anchor, rule["action"], rows_data)

            # ── 保存输出 ──────────────────────────────────────────
            out_path = os.path.join(self.result_dir, "自动填报完成_效益审核表.xlsx")
            wb.save(out_path)
            self.log.info("🎉 填报完成：%s", out_path)
            return True

        except Exception as e:
            self.log.error("❌ 填报异常: %s\n%s", e, traceback.format_exc())
            return False