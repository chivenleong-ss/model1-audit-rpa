# benefit_reporter.py（调整后）
# -*- coding: utf-8 -*-
"""
benefit_reporter.py  v6
基于 v5，重构消除冗余和易错点：
  1. 引入共享配置 project_config，统一管理固定项、ZGD、研发冲销映射。
  2. 材料固定项从配置读取，消除重复和遗漏。
  3. 材料费聚合不再排除 ZGD，由插行自动生成行；删除 ZGD 旁路填充代码。
  4. 研发冲销行显示为“减：研发领用材料”，保留负数以保证成本正确。
  5. 供应商匹配失败时输出调试日志，支持简称映射尝试。
"""

import os, re, traceback
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side, Font, Alignment, PatternFill, GradientFill
import copy as py_copy
from project_config import MATERIAL_FIXED_SUBCATS, VENDOR_SHORTNAME_MAP

MATERIAL_ROW_ALIASES = {
    '材料库存': ['材料库存'],
    '材料调入': ['材料调入'],
    '材料调出': ['材料调出'],
    '材料调入+CFK': ['材料调入+CFK', 'CFK材料调入'],
    '材料调出+XSD': ['材料调出+XSD', 'XSD材料调出'],
    '材料调出+SQD': ['材料调出+SQD', 'SQD材料调出'],
    '废旧物资消耗': ['废旧物资消耗', '废旧物资处置'],
    '废旧物资消耗（SQD）': ['废旧物资消耗（SQD）', '废旧物资消耗(SQD)', '废旧物资处置+SQD'],
    '其他': ['其他'],
}

# ─── 列号常量 (1-based) ───────────────────────────────────────
COL_A, COL_B, COL_C, COL_D = 1, 2, 3, 4
COL_H, COL_I, COL_J, COL_K = 8, 9, 10, 11
COL_L, COL_M, COL_N        = 12, 13, 14
COL_O, COL_P, COL_Q, COL_R = 15, 16, 17, 18
COL_S, COL_T                = 19, 20   # S=累计付款(总), T=财务账面已列报付款

# ─── 小节小计关键词映射 ───────────────────────────────────────
SUBTOTAL_MAP = {
    '（一）人工费':      '人工费小计',
    '（二）分包工程':    '分包工程小计',
    '(三)材料费':       '材料费小计',
    '(四）机械租赁费':   '机械费小计',
    '（五）其他直接费':  '其他直接费小计',
    '（六）间接费':      '间接费小计',
    '（七）安全费':      '安全费小计',
}


class BenefitReporter:

    def __init__(self, result_dir, logger):
        self.result_dir    = result_dir
        self.log           = logger
        self.template_path = "项目效益审核表.xlsx"
        self.invoice_ledger_path = None

    # ═══════════════════════════════════════════════════════════
    #  §1  工具
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _norm(val) -> str:
        if val is None: return ""
        return (str(val)
                .replace(" ", "").replace("\u3000", "")
                .replace("（", "(").replace("）", ")")
                .replace("：", ":").replace("*", "")
                .replace(":", ""))

    def _find_row(self, ws, keyword, cols=(2, 3, 4, 5)) -> int | None:
        kw = self._norm(keyword)
        if not kw: return None
        for r in range(1, ws.max_row + 1):
            for c in cols:
                if kw in self._norm(ws.cell(r, c).value):
                    return r
        return None

    def _find_material_row(self, ws, keyword) -> int | None:
        for alias in MATERIAL_ROW_ALIASES.get(keyword, [keyword]):
            row = self._find_row(ws, alias)
            if row:
                return row
        return None

    def _find_anchor(self, ws, subtotal_kw, min_row: int = 1) -> int | None:
        sub_row = self._find_row(ws, subtotal_kw)
        if not sub_row: return None
        for r in range(sub_row - 1, max(min_row, sub_row - 300), -1):
            for c in range(2, 6):
                if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                    return r
        return None

    # ═══════════════════════════════════════════════════════════
    #  §2  行样式
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _copy_fill(src_cell, dst_cell):
        try:
            fill = src_cell.fill
            if fill and fill.fill_type not in (None, "none"):
                dst_cell.fill = py_copy.copy(fill)
        except Exception:
            pass

    def _style_row(self, ws, r, ncol=22, skip_fill=False):
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font   = Font(name="宋体", size=10)
        for c in range(1, ncol + 1):
            cell = ws.cell(r, c)
            cell.border = border
            cell.font   = font
            cell.alignment = Alignment(
                horizontal="left" if c in (COL_C, COL_D) else "center",
                vertical="center", wrap_text=(c in (COL_C, COL_D)))
            if r > 1 and not skip_fill:
                self._copy_fill(ws.cell(r - 1, c), cell)

    # ═══════════════════════════════════════════════════════════
    #  §3  行公式生成
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _row_formulas(r: int, extended: bool) -> dict:
        f = {
            COL_O: f"=L{r}-M{r}",
            COL_P: f"=K{r}-N{r}",
            COL_Q: f"=M{r}+O{r}",
            COL_R: f"=M{r}+N{r}+O{r}+P{r}",
        }
        if extended:
            f[COL_J] = f"=H{r}+I{r}"
            f[COL_L] = f"=J{r}-K{r}"
        return f

    def _write_row(self, ws, r, d_val, c_val, m_val, extended, skip_fill=False):
        self._style_row(ws, r, skip_fill=skip_fill)
        if d_val: ws.cell(r, COL_D).value = str(d_val)
        if c_val: ws.cell(r, COL_C).value = str(c_val)
        if m_val is not None:
            cell = ws.cell(r, COL_M)
            cell.value = round(float(m_val), 2)
            cell.number_format = '#,##0.00'
        for col, fml in self._row_formulas(r, extended).items():
            ws.cell(r, col).value = fml

    # ═══════════════════════════════════════════════════════════
    #  §4  公式修复
    # ═══════════════════════════════════════════════════════════

    def _fix_stale_self_refs(self, ws):
        SKIP_D_WORDS = {'成本合计', '成本及费用合计', '利润', '增值税实际税负', '税金及附加'}
        TARGET_COLS  = {COL_J, COL_L, COL_O, COL_P, COL_Q, COL_R, COL_S}

        for r in range(10, ws.max_row + 1):
            d_val = self._norm(str(ws.cell(r, COL_D).value or ''))
            if any(w in d_val for w in SKIP_D_WORDS):
                continue
            for c in TARGET_COLS:
                cell = ws.cell(r, c)
                if not isinstance(cell.value, str) or not cell.value.startswith('='):
                    continue
                formula = cell.value
                if ':' in formula:
                    m = re.fullmatch(r'=SUM\(T(\d+):V\1\)', formula, re.IGNORECASE)
                    if m:
                        x = int(m.group(1))
                        if x != r:
                            cell.value = f'=SUM(T{r}:V{r})'
                    continue
                nums = [int(n) for n in re.findall(r'(?<=[A-Za-z])(\d+)', formula)]
                if not nums: continue
                unique = set(nums)
                if len(unique) == 1:
                    x = unique.pop()
                    if x != r and x > 5:
                        cell.value = re.sub(
                            r'(?<=[A-Za-z])' + str(x) + r'(?!\d)',
                            str(r), formula
                        )

    def _fix_all_sums(self, ws):
        for hdr_kw, sub_kw in SUBTOTAL_MAP.items():
            hdr_row = self._find_row(ws, hdr_kw)
            sub_row = self._find_row(ws, sub_kw)
            if not hdr_row or not sub_row: continue

            anchor = None
            for r in range(sub_row - 1, max(hdr_row, sub_row - 400), -1):
                for c in range(2, 6):
                    if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                        anchor = r; break
                if anchor: break
            end_row = anchor if anchor else sub_row - 1

            for col in range(1, ws.max_column + 1):
                cell = ws.cell(sub_row, col)
                if cell.value and isinstance(cell.value, str) and 'SUM(' in cell.value.upper():
                    cl = get_column_letter(col)
                    cell.value = f"=SUM({cl}{hdr_row}:{cl}{end_row})"

            if anchor:
                for col, fml in self._row_formulas(anchor, False).items():
                    cell = ws.cell(anchor, col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('=') and ':' not in cell.value:
                        cell.value = fml

            self.log.info("   🧮 %-14s SUM 范围 行%d→%d", sub_kw, hdr_row, end_row)

    def _fix_aggregate_rows(self, ws):
        kw_to_sub = {
            '人工费小计':    '（一）人工费',
            '分包工程小计':  '（二）分包工程',
            '材料费小计':    '(三)材料费',
            '机械费小计':    '(四）机械租赁费',
            '其他直接费小计': '（五）其他直接费',
            '间接费小计':    '（六）间接费',
            '安全费小计':    '（七）安全费',
        }
        sub_rows = {}
        for kw in kw_to_sub:
            r = self._find_row(ws, kw)
            if r: sub_rows[kw] = r

        cost_total_row = self._find_row(ws, '三、成本合计')
        if cost_total_row and len(sub_rows) == len(kw_to_sub):
            ordered = ['人工费小计','分包工程小计','材料费小计','机械费小计',
                       '其他直接费小计','间接费小计','安全费小计']
            row_refs = [sub_rows[k] for k in ordered]
            for col in range(8, ws.max_column + 1):
                cell = ws.cell(cost_total_row, col)
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    cl = get_column_letter(col)
                    cell.value = '=' + '+'.join(f'{cl}{r}' for r in row_refs)
            self.log.info("   🧮 三、成本合计 公式已修正（行 %d）", cost_total_row)

        total_row = self._find_row(ws, '八、成本及费用合计')
        if total_row and cost_total_row:
            four_row  = self._find_row(ws, '四、计提保修金')
            five_row  = self._find_row(ws, '五、过程节点奖金')
            six_row   = self._find_row(ws, '六、资金占用费用')
            seven_row = self._find_row(ws, '七、局投资收益')
            need = [cost_total_row, four_row, five_row, six_row, seven_row]
            if all(need):
                for col in range(8, ws.max_column + 1):
                    cell = ws.cell(total_row, col)
                    if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                        cl = get_column_letter(col)
                        cell.value = '=' + '+'.join(f'{cl}{r}' for r in need)
                self.log.info("   🧮 八、成本及费用合计 公式已修正（行 %d）", total_row)

    # ═══════════════════════════════════════════════════════════
    #  §5  数据聚合
    # ═══════════════════════════════════════════════════════════

    @staticmethod
    def _sort_rows(rows):
        """
        D列排序新规则：
        1. 带有客商实体的行 (is_v=True) 绝对排在前面；
        2. 在此基础上，有合同编码的行优先；
        """
        return sorted(rows, key=lambda x: (
            0 if x.get("is_v", False) else 1,
            0 if x.get("c") and str(x.get("c", "")).strip() not in ("", "nan", "None") else 1
        ))

    def _agg_labor(self, df):
        df = df.copy()
        df['_pfx'] = df['中台单据号'].fillna('').astype(str).str[:3].str.upper()
        df['_ven'] = df['客商名称'].fillna('').astype(str).str.strip().replace('nan', '')
        df['_con'] = df['合同编码'].fillna('').astype(str).str.strip().replace('nan', '')
        rows = []

        fgd = df[df['_pfx'] == 'FGD']
        if not fgd.empty:
            rows.append({"d": "FGD", "c": None, "p": round(fgd['最终发生额'].sum(), 2), "is_v": True})

        for ven, grp in df[df['_pfx'] == 'CFK'].groupby('_ven', sort=False):
            label = f"CFK+{ven}" if ven else "CFK"
            rows.append({"d": label, "c": None, "p": round(grp['最终发生额'].sum(), 2), "is_v": True})

        others = df[~df['_pfx'].isin(['FGD', 'CFK'])].copy()
        has_vc = others[(others['_ven'] != '') | (others['_con'] != '')]
        for (ven, con), grp in has_vc.groupby(['_ven', '_con'], sort=False):
            rows.append({"d": ven or None, "c": con or None, "p": round(grp['最终发生额'].sum(), 2), "is_v": True})

        no_vc = others[~others.index.isin(has_vc.index)]
        if not no_vc.empty:
            pass # 沿用原有规则处理无客商无合同的数据（如有）

        return self._sort_rows(rows)

    def _agg_vendor(self, df):
        rows = []
        for (ven, con), grp in df.groupby(
                [df['客商名称'].fillna(''), df['合同编码'].fillna('')], sort=False):
            rows.append({"d": ven or None, "c": con or None, "p": round(grp['最终发生额'].sum(), 2), "is_v": True})
        return self._sort_rows(rows)

    def _agg_sub(self, df):
        rows = []
        for sub, grp in df.groupby(df['细分科目'].fillna(''), sort=False):
            # 细分科目统一视作非客商
            rows.append({"d": sub or None, "c": None, "p": round(grp['最终发生额'].sum(), 2), "is_v": False})
        return rows

    def _agg_material(self, df):
        """
        材料费聚合规则更新：合并完成后统一应用全局排序（优先排有客商的行）
        """
        excl_set = set(MATERIAL_FIXED_SUBCATS)
        df = df[~df['细分科目'].fillna('').astype(str).str.strip().isin(excl_set)].copy()

        if df.empty:
            return []

        has_c = df[df['合同编码'].fillna('').astype(str).str.strip().replace('nan', '') != '']
        no_c  = df[~df.index.isin(has_c.index)]

        # 拼接客商聚合池与细分科目聚合池，并统一用 _sort_rows 进行全局“客商优先”排序
        rows = self._sort_rows(self._agg_vendor(has_c) + self._agg_sub(no_c))

        # 美化研发冲销行的显示名称
        for rd in rows:
            d_val = str(rd.get('d', ''))
            if d_val.startswith('研发支出-'):
                rd['d'] = '减：研发费用-材料费'
        return rows

    # ═══════════════════════════════════════════════════════════
    #  §6  占位行清除
    # ═══════════════════════════════════════════════════════════

    _KEEP_ROW_KEYWORDS = {
        '（1）集中采购', '(1)集中采购', '（2）自行采购', '(2)自行采购',
        '材料库存', '材料调入', '材料调出', '材料调入+CFK',
        '研发支出', '研发费用'
    }

    def _delete_stubs(self, ws, hdr_row: int, anchor_row: int):
        protected = {self._norm(k) for k in self._KEEP_ROW_KEYWORDS}
        to_del = []
        for r in range(hdr_row + 1, anchor_row):
            d = ws.cell(r, COL_D).value
            m = ws.cell(r, COL_M).value
            if isinstance(m, (int, float)):
                continue
            d_norm = self._norm(str(d or ''))
            if any(k in d_norm for k in protected):
                continue
            to_del.append(r)
        for r in reversed(to_del):
            ws.delete_rows(r)
        if to_del:
            self.log.info("      🗑 清除 %d 行（旧数据/公式占位行）", len(to_del))

    def _insert_rows(self, ws, anchor, action, rows_data, extended, is_rd=False):
        for i, rd in enumerate(rows_data):
            ins = anchor + i if action == 'above' else anchor + 1 + i
            ws.insert_rows(ins)
            self._write_row(ws, ins, rd.get('d'), rd.get('c'), rd.get('p', 0), extended, skip_fill=is_rd)

    def _ensure_material_fixed_rows(self, ws, extended) -> None:
        hdr_row = self._find_row(ws, '(三)材料费')
        subtotal_row = self._find_row(ws, '材料费小计')
        if not subtotal_row:
            return

        missing = [sub for sub in MATERIAL_FIXED_SUBCATS if self._find_material_row(ws, sub) is None]
        if not missing:
            return

        anchor_row = self._find_anchor(ws, '材料费小计', min_row=hdr_row or 1)
        insert_at = anchor_row or subtotal_row
        for sub in missing:
            ws.insert_rows(insert_at)
            self._write_row(ws, insert_at, sub, None, None, extended)
            insert_at += 1

        self.log.info("   已自动补插材料费固定行: %s", "、".join(missing))

    # ═══════════════════════════════════════════════════════════
    #  §7  B/N/T 列匹配填写
    # ═══════════════════════════════════════════════════════════

    def _build_lookups(self, df4):
        ven_col = '供应商名称' if '供应商名称' in df4.columns else (
                  '供应商描述' if '供应商描述' in df4.columns else None)

        vendor_lkp = {}
        if ven_col:
            sub = df4[df4[ven_col].notna() & df4['供应商'].notna()]
            dedup_cols = [ven_col, '合同'] if '合同' in sub.columns else [ven_col]
            for _, row in sub.drop_duplicates(dedup_cols).iterrows():
                ven  = self._norm(str(row[ven_col]))
                con  = self._norm(str(row.get('合同','') or ''))
                code = str(int(float(row['供应商']))) if pd.notna(row['供应商']) else ''
                if ven:
                    vendor_lkp[(ven, con)] = code
                    if (ven, '') not in vendor_lkp:
                        vendor_lkp[(ven, '')] = code

        ap_mask = (df4['总账科目长文本'].fillna('').str.contains('应付账款') &
                   ~df4['总账科目长文本'].fillna('').str.contains('待确认进项税额'))
        ap = df4[ap_mask].copy()
        ap['_v'] = ap['供应商'].fillna(0).apply(
            lambda x: str(int(float(x))) if pd.notna(x) and x else '')
        ap['_c'] = ap['合同'].fillna('').astype(str).str.strip()
        pay_lkp = dict(ap.groupby(['_v','_c'])['借方本位币金额'].sum())

        vat = df4[df4['总账科目长文本'].fillna('').str.contains('其他应收款') &
                  df4['总账科目长文本'].fillna('').str.contains('待确认进项税额')].copy()
        vat['_v'] = vat['供应商'].fillna(0).apply(
            lambda x: str(int(float(x))) if pd.notna(x) and x else '')
        vat['_c'] = vat['合同'].fillna('').astype(str).str.strip()
        vat_lkp = dict(vat.groupby(['_v','_c'])['借方本位币金额'].sum())

        return vendor_lkp, pay_lkp, vat_lkp

    def _fill_per_row_cols(self, ws, vendor_lkp, pay_lkp, vat_lkp):
        SKIP = {'人工费小计','分包工程小计','材料费小计','机械费小计','其他直接费小计',
                '间接费小计','安全费小计','财务未列部分需增列','成本合计','项目效益',
                '成本及费用合计','增值税','税金及附加','利润','保修金','节点奖'}
        for r in range(10, ws.max_row + 1):
            d_val = ws.cell(r, COL_D).value
            m_val = ws.cell(r, COL_M).value
            c_val = ws.cell(r, COL_C).value
            if not d_val or not isinstance(m_val, (int, float)): continue
            if any(s in self._norm(str(d_val)) for s in SKIP): continue

            d_norm = self._norm(str(d_val))
            c_norm = self._norm(str(c_val or ''))
            code = (vendor_lkp.get((d_norm, c_norm)) or
                    vendor_lkp.get((d_norm, ''), ''))

            # 尝试简称映射
            if not code and d_norm in VENDOR_SHORTNAME_MAP:
                mapped = VENDOR_SHORTNAME_MAP[d_norm]
                code = vendor_lkp.get((self._norm(mapped), c_norm)) or vendor_lkp.get((self._norm(mapped), ''))
            if not code and d_norm:
                log_debug = getattr(self.log, 'debug', None)
                if callable(log_debug):
                    log_debug("未匹配到供应商编码：行%d D='%s' C='%s'", r, d_val, c_val)

            if code:
                ws.cell(r, COL_B).value = code
                pay = pay_lkp.get((code, c_norm), 0)
                if pay:
                    c = ws.cell(r, COL_T)
                    c.value = round(float(pay), 2)
                    c.number_format = '#,##0.00'
                vat = vat_lkp.get((code, c_norm), 0)
                if vat:
                    c = ws.cell(r, COL_N)
                    c.value = round(float(vat), 2)
                    c.number_format = '#,##0.00'

    def _apply_invoice_match(self, ws) -> None:
        invoice_path = self.invoice_ledger_path
        if not invoice_path:
            try:
                for fn in os.listdir(self.result_dir):
                    if fn.lower().endswith((".xlsx", ".xls")) and any(k in fn for k in ["发票", "收票", "台账", "台帳"]):
                        invoice_path = os.path.join(self.result_dir, fn)
                        break
            except Exception:
                invoice_path = None

        if not (invoice_path and os.path.exists(invoice_path)):
            self.log.info("   ⓘ 未提供发票台账，跳过 AB 列发票匹配")
            return

        from invoice_matcher import InvoiceMatcher
        matcher = InvoiceMatcher(self.log)
        matcher.apply_to_worksheet(ws, invoice_path)

    # ═══════════════════════════════════════════════════════════
    #  §8  A列序号重排
    # ═══════════════════════════════════════════════════════════

    def _renumber(self, ws):
        sections = [
            ('（一）人工费',    '人工费小计'),
            ('（二）分包工程',  '分包工程小计'),
            ('(三)材料费',     '材料费小计'),
            ('(四）机械租赁费', '机械费小计'),
        ]
        for hdr_kw, sub_kw in sections:
            hdr = self._find_row(ws, hdr_kw)
            sub = self._find_row(ws, sub_kw)
            if not hdr or not sub: continue
            anchor = None
            for r in range(sub - 1, hdr, -1):
                for c in range(2, 6):
                    if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                        anchor = r; break
                if anchor: break
            end = (anchor - 1) if anchor else (sub - 1)
            for r in range(hdr + 1, sub):
                ws.cell(r, COL_A).value = None
            seq = 1
            for r in range(hdr + 1, end + 1):
                d = ws.cell(r, COL_D).value
                m = ws.cell(r, COL_M).value
                if d and isinstance(m, (int, float)) and m != 0:
                    ws.cell(r, COL_A).value = seq
                    seq += 1

    # ═══════════════════════════════════════════════════════════
    #  §9  主流程
    # ═══════════════════════════════════════════════════════════

    def execute_fill(self):
        source_path = os.path.join(self.result_dir, "5_中间汇总表_效益审核数据源.xlsx")
        merged_path = os.path.join(self.result_dir, "4_合并表.xlsx")

        if not os.path.exists(source_path):
            self.log.error("⚠️ 找不到中间汇总表：%s", source_path); return False
        if not os.path.exists(self.template_path):
            self.log.error("⚠️ 找不到模板：%s", self.template_path); return False

        self.log.info("📝 填报引擎 v6 启动（冗余消除版）...")
        try:
            df_raw = pd.read_excel(source_path)
            df = df_raw[df_raw['利润中心'].astype(str).str.strip() != '筛选合计：'].copy()
            df['最终发生额'] = pd.to_numeric(df['最终发生额'], errors='coerce').fillna(0)
            for col in ['合同编码', '细分科目', '客商名称', '中台单据号']:
                if col in df.columns:
                    df[col] = df[col].fillna('').astype(str).str.strip().replace('nan', '')

            df4 = pd.read_excel(merged_path) if os.path.exists(merged_path) else pd.DataFrame()
            vendor_lkp = pay_lkp = vat_lkp = {}
            if not df4.empty:
                vendor_lkp, pay_lkp, vat_lkp = self._build_lookups(df4)

            wb = openpyxl.load_workbook(self.template_path)
            ws = wb.active
            for sn in wb.sheetnames:
                if any(k in sn for k in ['效益审核','附表']):
                    ws = wb[sn]; break

            def _first(col):
                if col not in df.columns: return ''
                v = df[col].replace('', None).dropna()
                return v.iloc[0] if not v.empty else ''

            ws['B6'] = _first('利润中心')
            ws['C6'] = _first('项目编码')
            ws['D6'] = _first('工程名称')

            if not df4.empty and '总账科目长文本' in df4.columns:
                rev = df4[df4['总账科目长文本'].fillna('').str.startswith('主营业务收入')]['贷方本位币金额'].sum()
                ws['M6'] = round(-float(rev), 2)
            else:
                rev = df[df['成本_财务大类'] == '财务累计入账收入(不含增值税)']['最终发生额'].sum()
                ws['M6'] = round(-float(rev), 2)

            n6 = df[df['成本_财务大类'] == '已价税分离的增值税金额(财务账面数据)']['最终发生额'].sum()
            ws['N6'] = round(-float(n6), 2)
            self.log.info("   ✅ 表头写入（B6/C6/D6/M6/N6）")

            plan = [
                ('(一)人工费',    'labor',    '人工费小计',    'above', False),
                ('(二)分包工程',  'vendor',   '分包工程小计',  'above', False),
                ('(三)材料费',    'material', '材料费小计',    'above', False),
                ('(四)机械租赁费','vendor',   '机械费小计',    'above', True),
                ('(五)其他直接费','sub',      '其他直接费小计','above', True),
                ('(六)间接费',    'sub',      '间接费小计',    'above', True),
                ('(七)安全费',    'sub',      '安全费小计',    'above', True),
                ('研发支出',      'sub',      '八、研发费用',  'below', False),
            ]
            hdr_kw_map = {
                '(一)人工费':    '（一）人工费',
                '(二)分包工程':  '（二）分包工程',
                '(三)材料费':   '(三)材料费',
                '(四)机械租赁费':'(四）机械租赁费',
                '(五)其他直接费':'（五）其他直接费',
                '(六)间接费':    '（六）间接费',
                '(七)安全费':    '（七）安全费',
            }

            for cat, mode, anchor_kw, action, ext in plan:
                cat_df = df[df['成本_财务大类'] == cat].copy()
                if cat_df.empty:
                    self.log.info("   ℹ️  [%s] 无数据", cat); continue

                if   mode == 'labor':    rows_data = self._agg_labor(cat_df)
                elif mode == 'vendor':   rows_data = self._agg_vendor(cat_df)
                elif mode == 'material': rows_data = self._agg_material(cat_df)
                else:                    rows_data = self._agg_sub(cat_df)

                if not rows_data: continue

                if action == 'above':
                    hdr_kw  = hdr_kw_map.get(cat, '')
                    hdr_row = self._find_row(ws, hdr_kw) if hdr_kw else None
                    anchor = self._find_anchor(ws, anchor_kw, min_row=hdr_row or 1)
                    if anchor is None:
                        anchor = self._find_row(ws, anchor_kw)
                else:
                    anchor = self._find_row(ws, anchor_kw)
                    hdr_row = None

                if anchor is None:
                    self.log.warning("   ⚠️  [%s] 锚点未找到（'%s'）", cat, anchor_kw); continue

                if action == 'above' and hdr_row:
                    self._delete_stubs(ws, hdr_row, anchor)
                    anchor = (self._find_anchor(ws, anchor_kw, min_row=hdr_row)
                              or self._find_row(ws, anchor_kw))

                self.log.info("   📍 [%s] 锚点=行%-4d  插入 %d 行", cat, anchor, len(rows_data))
                is_rd = (cat == '研发支出')
                self._insert_rows(ws, anchor, action, rows_data, ext, is_rd=is_rd)

            self._ensure_material_fixed_rows(ws, ext)
            self._fix_stale_self_refs(ws)
            self._fix_all_sums(ws)
            self._fix_aggregate_rows(ws)

            # 固定项填写（包括材料费固定子项，从配置读取）
            def _fill_fixed(cat_or_filter, kw, label):
                if isinstance(cat_or_filter, str):
                    amt = df[df['成本_财务大类'] == cat_or_filter]['最终发生额'].sum()
                else:
                    mask = pd.Series(True, index=df.index)
                    for k, v in cat_or_filter.items(): mask &= (df[k] == v)
                    amt = df[mask]['最终发生额'].sum()
                r = self._find_material_row(ws, kw) if kw in MATERIAL_ROW_ALIASES else self._find_row(ws, kw)
                if r and amt != 0:
                    ws.cell(r, COL_M).value = round(float(amt), 2)
                    ws.cell(r, COL_M).number_format = '#,##0.00'
                    self.log.info("   ✅ %-20s 行%d  M=%.2f", label, r, amt)
                elif not r:
                    self.log.warning("   ⚠️  找不到关键词 '%s'", kw)

            _fill_fixed('六、资金占用费用', '六、资金占用费用', '资金占用费用')
            _fill_fixed('七、局投资收益（局投资项目选填）', '七、局投资收益', '局投资收益')

            for sub in MATERIAL_FIXED_SUBCATS:
                _fill_fixed({'成本_财务大类': '(三)材料费', '细分科目': sub}, sub, sub)

            # ZGD 已由插行处理，不再单独填充

            tax_amt = df[df['成本_财务大类'].fillna('').str.contains('税金及附加')]['最终发生额'].sum()
            r_tax = self._find_row(ws, '十一、税金及附加')
            if r_tax and tax_amt:
                ws.cell(r_tax, COL_M).value = round(float(tax_amt), 2)
                ws.cell(r_tax, COL_M).number_format = '#,##0.00'
                self.log.info("   ✅ 税金及附加 行%d M=%.2f", r_tax, tax_amt)

            if not df4.empty:
                self._fill_per_row_cols(ws, vendor_lkp, pay_lkp, vat_lkp)
            self.log.info("   ✅ B/N/T 列匹配填写完成")

            self._apply_invoice_match(ws)
            self._renumber(ws)
            self.log.info("   ✅ A列序号重排完成")

            out = os.path.join(self.result_dir, "自动填报完成_效益审核表.xlsx")
            wb.save(out)
            self.log.info("🎉 填报完成：%s", out)
            return True

        except Exception as e:
            self.log.error("❌ 填报异常:\n%s", traceback.format_exc()); return False
