# -*- coding: utf-8 -*-
"""
invoice_matcher.py

将“发票收票台账”匹配进《项目效益审核表》：
  - AC(合同客商列号) -> 审核表 B(客商编码)
  - AE(合同编码列号) -> 审核表 C(合同编码)
  - Q(价税总额) -> 无论是否认证，均汇总写入 审核表 AB(累计已计票发票额, 第28列)
  - P(税额) -> 无论是否认证，均汇总写入 审核表 AD(第30列)
  - P(税额) -> 当 AK 列为“已认证”时，汇总写入 审核表 AE(第31列)

说明：
  - 本文件只负责发票匹配；由 benefit_reporter.py 在填报完成后调用并一并输出交付版审核表。
"""

from __future__ import annotations

import os

import openpyxl
import pandas as pd


class InvoiceMatcher:
    COL_AUDIT_B = 2
    COL_AUDIT_C = 3
    COL_AUDIT_D = 4
    COL_AUDIT_M = 13
    
    COL_AUDIT_AB = 28  # Q列全量价税总额
    COL_AUDIT_AD = 30  # P列全量税额
    COL_AUDIT_AE = 31  # P列已认证税额

    def __init__(self, logger=None):
        self.log = logger
        # 缓存结构: (code, con) -> {'q_all': x, 'p_all': y, 'p_cert': z}
        self._lookup_cache: dict[tuple[str, float, int], dict[tuple[str, str], dict[str, float]]] = {}

    @staticmethod
    def _norm_key(v) -> str:
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        s = str(v).strip()
        if s.endswith(".0"):
            try:
                return str(int(float(s)))
            except Exception:
                pass
        return "" if s.lower() == "nan" else s

    @staticmethod
    def _num(v) -> float:
        n = pd.to_numeric(v, errors="coerce")
        return 0.0 if pd.isna(n) else float(n)

    def build_lookup(self, invoice_path: str) -> dict[tuple[str, str], dict[str, float]]:
        if not invoice_path or (not os.path.exists(invoice_path)):
            return {}
        stat = os.stat(invoice_path)
        cache_key = (os.path.abspath(invoice_path), stat.st_mtime, stat.st_size)
        if cache_key in self._lookup_cache:
            return self._lookup_cache[cache_key]

        # 按用户给定的固定列位读取，不依赖表头文字：
        # P=16, Q=17, AC=29, AE=31, AK=37（均为 1-based，代码里是 0-based）
        rows = []
        if invoice_path.lower().endswith(".xls"):
            inv = pd.read_excel(invoice_path, header=0, usecols=[15, 16, 28, 30, 36])
            for _, row in inv.iterrows():
                rows.append({
                    "_tax": row.iloc[0] if len(row) > 0 else None,
                    "_amt": row.iloc[1] if len(row) > 1 else None,
                    "_code": row.iloc[2] if len(row) > 2 else None,
                    "_con": row.iloc[3] if len(row) > 3 else None,
                    "_status": row.iloc[4] if len(row) > 4 else None,
                })
        else:
            wb = openpyxl.load_workbook(invoice_path, data_only=True, read_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=2, values_only=True):
                tax = row[15] if len(row) > 15 else None
                amt = row[16] if len(row) > 16 else None
                code = row[28] if len(row) > 28 else None
                con = row[30] if len(row) > 30 else None
                status = row[36] if len(row) > 36 else None
                rows.append({"_tax": tax, "_amt": amt, "_code": code, "_con": con, "_status": status})
            wb.close()

        if not rows:
            return {}

        inv = pd.DataFrame(rows)
        inv["_code"] = inv["_code"].apply(self._norm_key)
        inv["_con"] = inv["_con"].apply(self._norm_key)
        inv["_status"] = inv["_status"].apply(self._norm_key)
        inv["_amt"] = inv["_amt"].apply(self._num)
        inv["_tax"] = inv["_tax"].apply(self._num)

        # 过滤掉 客商 和 合同 均为空的无效行
        inv = inv[(inv["_code"] != "") & (inv["_con"] != "")]
        if inv.empty:
            if self.log:
                self.log.warning("   ⚠️ 发票台账中 AC/AE 无可匹配的客商或合同数据")
            return {}

        # 1. 不管是否认证，全量计算 Q列(amt) 和 P列(tax)
        agg_all = inv.groupby(["_code", "_con"], dropna=False)[["_amt", "_tax"]].sum()

        # 2. 仅筛选已认证的，计算 P列(tax)
        inv_cert = inv[inv["_status"] == "已认证"]
        if inv_cert.empty:
            agg_cert = pd.Series(dtype=float, name="_tax_cert")
        else:
            agg_cert = inv_cert.groupby(["_code", "_con"], dropna=False)["_tax"].sum().rename("_tax_cert")

        # 3. 将全量数据与已认证数据合并
        agg = agg_all.join(agg_cert).fillna(0).round(2)

        lookup = {}
        for idx, row in agg.iterrows():
            lookup[idx] = {
                "q_all": float(row["_amt"]),
                "p_all": float(row["_tax"]),
                "p_cert": float(row["_tax_cert"])
            }
            
        self._lookup_cache[cache_key] = lookup
        if self.log:
            self.log.info("   🧾 发票台账读取完成: %s (%d 组匹配键)", os.path.basename(invoice_path), len(lookup))
        return lookup

    def apply_to_worksheet(self, ws, invoice_path: str) -> int:
        """
        将台账匹配结果写入 ws 的 AB、AD、AE 列；返回写入行数。
        仅对“明细行”写入：D列有值 且 M列为数值。
        """
        lkp = self.build_lookup(invoice_path)
        if not lkp:
            return 0

        filled = 0
        for r in range(7, ws.max_row + 1):
            d_val = ws.cell(r, self.COL_AUDIT_D).value
            m_val = ws.cell(r, self.COL_AUDIT_M).value
            
            # 判断是否是有效的数据行
            if not (d_val and isinstance(m_val, (int, float))):
                continue
                
            code = self._norm_key(ws.cell(r, self.COL_AUDIT_B).value)
            con = self._norm_key(ws.cell(r, self.COL_AUDIT_C).value)
            if (not code) or (not con):
                continue
                
            vals = lkp.get((code, con))
            if vals:
                # 写入 Q列 全量价税总额 到 AB 列
                if vals["q_all"] != 0:
                    c_ab = ws.cell(r, self.COL_AUDIT_AB)
                    c_ab.value = vals["q_all"]
                    c_ab.number_format = "#,##0.00"
                    
                # 写入 P列 全量税额 到 AD 列
                if vals["p_all"] != 0:
                    c_ad = ws.cell(r, self.COL_AUDIT_AD)
                    c_ad.value = vals["p_all"]
                    c_ad.number_format = "#,##0.00"
                    
                # 写入 P列 已认证税额 到 AE 列
                if vals["p_cert"] != 0:
                    c_ae = ws.cell(r, self.COL_AUDIT_AE)
                    c_ae.value = vals["p_cert"]
                    c_ae.number_format = "#,##0.00"
                    
                filled += 1

        if self.log:
            if filled:
                self.log.info("   ✅ 发票明细匹配并写入 AB/AD/AE 列：%d 行", filled)
            else:
                self.log.warning("   ⚠️ 发票台账已读取，但未匹配到审核表 B/C 对应行")
        return filled