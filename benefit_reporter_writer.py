# -*- coding: utf-8 -*-
"""
效益表写表动作与后处理动作。
"""

from __future__ import annotations

import copy as py_copy
import os
import re

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side
from openpyxl.utils import get_column_letter

from benefit_reporter_template import (
    COL_A,
    COL_B,
    COL_C,
    COL_D,
    COL_H,
    COL_J,
    COL_L,
    COL_M,
    COL_N,
    COL_O,
    COL_P,
    COL_Q,
    COL_R,
    COL_S,
    COL_T,
    SUBTOTAL_MAP,
)
from project_config import MATERIAL_FIXED_SUBCATS, VENDOR_SHORTNAME_MAP


class BenefitSheetWriter:
    _KEEP_ROW_KEYWORDS = {
        '（1）集中采购', '(1)集中采购', '（2）自行采购', '(2)自行采购',
        '材料库存', '材料调入', '材料调出', '材料调入+CFK',
        '研发支出', '研发费用'
    }
    _SECTION_HEADER_MAP = {
        '(一)人工费': 'labor',
        '(二)分包工程': 'vendor',
        '(三)材料费': 'material',
        '(四)机械租赁费': 'machine',
        '(五)其他直接费': 'other',
        '(六)间接费': 'other',
        '(七)安全费': 'other',
    }

    def __init__(self, logger, locator):
        self.log = logger
        self.locator = locator


    @staticmethod
    def _set_value(ws, row: int, col: int, value):
        if row <= 10:
            return False
        ws.cell(row, col).value = value
        return True

    @staticmethod
    def _normalize_code(value) -> str:
        if value is None:
            return ''
        text = str(value).strip()
        if text in ('', 'nan', 'None'):
            return ''
        try:
            num = float(text.replace(',', ''))
            if num.is_integer():
                return str(int(num))
        except Exception:
            pass
        return text

    @staticmethod
    def _classify_gl_to_section(gl_text: str) -> str:
        """从总账科目长文本判断属于审核表的哪个节（labor/vendor/material/machine/other）"""
        if not gl_text or not isinstance(gl_text, str):
            return 'other'
        if '直接人工费' in gl_text:
            return 'labor'
        if '分包工程支出' in gl_text:
            return 'vendor'
        if '直接材料费' in gl_text or '原材料' in gl_text:
            return 'material'
        if '机械使用费' in gl_text:
            return 'machine'
        return 'other'

    @staticmethod
    def _classify_safety_by_opponent(opp_gl: str) -> str:
        """根据对方科目描述判断安全生产费归属节（labor/vendor/machine/material/''）"""
        if not opp_gl or not isinstance(opp_gl, str):
            return ''
        s = opp_gl.replace(' ', '')
        if '直接人工' in s:
            return 'labor'
        if '分包' in s:
            return 'vendor'
        if '机械' in s:
            return 'machine'
        if '材料' in s or '原材料' in s:
            return 'material'
        return ''

    @staticmethod
    def copy_fill(src_cell, dst_cell):
        try:
            fill = src_cell.fill
            if fill and fill.fill_type not in (None, "none"):
                dst_cell.fill = py_copy.copy(fill)
        except Exception:
            pass

    @staticmethod
    def row_formulas(row_num: int, extended: bool) -> dict:
        formulas = {
            COL_O: f"=L{row_num}-M{row_num}",
            COL_P: f"=K{row_num}-N{row_num}",
            COL_Q: f"=M{row_num}+O{row_num}",
            COL_R: f"=M{row_num}+N{row_num}+O{row_num}+P{row_num}",
        }
        if extended:
            formulas[COL_J] = f"=H{row_num}+I{row_num}"
            formulas[COL_L] = f"=J{row_num}-K{row_num}"
        return formulas

    def style_row(self, ws, row_num, ncol=22, skip_fill=False):
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        font = Font(name="宋体", size=10)
        for col in range(1, ncol + 1):
            cell = ws.cell(row_num, col)
            cell.border = border
            cell.font = font
            cell.alignment = Alignment(
                horizontal="left" if col in (COL_C, COL_D) else "center",
                vertical="center",
                wrap_text=(col in (COL_C, COL_D)),
            )
            if row_num > 1 and not skip_fill:
                self.copy_fill(ws.cell(row_num - 1, col), cell)

    def write_row(self, ws, row_num, d_val, c_val, m_val, extended, skip_fill=False):
        self.style_row(ws, row_num, skip_fill=skip_fill)
        if d_val:
            self._set_value(ws, row_num, COL_D, str(d_val))
        if c_val:
            self._set_value(ws, row_num, COL_C, str(c_val))
        if m_val is not None:
            cell = ws.cell(row_num, COL_M)
            self._set_value(ws, row_num, COL_M, round(float(m_val), 2))
            cell.number_format = '#,##0.00'
        for col, formula in self.row_formulas(row_num, extended).items():
            self._set_value(ws, row_num, col, formula)

    def delete_stubs(self, ws, hdr_row: int, anchor_row: int):
        protected = {self.locator.norm(k) for k in self._KEEP_ROW_KEYWORDS}
        to_delete = []
        for row in range(hdr_row + 1, anchor_row):
            c_val = ws.cell(row, COL_C).value
            d_val = ws.cell(row, COL_D).value
            m_val = ws.cell(row, COL_M).value
            if isinstance(m_val, (int, float)):
                continue
            c_norm = self.locator.norm(str(c_val or ''))
            d_norm = self.locator.norm(str(d_val or ''))
            if any(key in c_norm or key in d_norm for key in protected):
                continue
            to_delete.append(row)
        for row in reversed(to_delete):
            ws.delete_rows(row)
        if to_delete:
            self.locator.clear_cache()
            self.log.info("      🗑 清除 %d 行（旧数据/公式占位行）", len(to_delete))

    def insert_rows(self, ws, anchor, action, rows_data, extended, is_rd=False):
        for index, row_data in enumerate(rows_data):
            insert_at = anchor + index if action == 'above' else anchor + 1 + index
            ws.insert_rows(insert_at)
            self.write_row(
                ws,
                insert_at,
                row_data.get('d'),
                row_data.get('c'),
                row_data.get('p', 0),
                extended,
                skip_fill=is_rd,
            )
        if rows_data:
            self.locator.clear_cache()

    def ensure_material_fixed_rows(self, ws, extended) -> None:
        hdr_row = self.locator.find_row(ws, '(三)材料费')
        subtotal_row = self.locator.find_row(ws, '材料费小计')
        if not subtotal_row:
            return

        missing = [sub for sub in MATERIAL_FIXED_SUBCATS if self.locator.find_material_row(ws, sub) is None]
        if not missing:
            return

        anchor_row = self.locator.find_anchor(ws, '材料费小计', min_row=hdr_row or 1)
        insert_at = anchor_row or subtotal_row
        for sub in missing:
            ws.insert_rows(insert_at)
            self.write_row(ws, insert_at, sub, None, None, extended)
            insert_at += 1

        self.locator.clear_cache()
        self.log.info("   已自动补插材料费固定行: %s", "、".join(missing))

    def ensure_material_group_headers(self, ws) -> None:
        hdr_row = self.locator.find_row(ws, '(三)材料费')
        subtotal_row = self.locator.find_row(ws, '材料费小计')
        if not hdr_row or not subtotal_row or subtotal_row <= hdr_row:
            return

        labels = ['（1）集中采购', '（2）自行采购']
        inserted = []
        insert_at = hdr_row + 1
        for label in labels:
            if self.locator.find_row(ws, label, cols=(3, 4)):
                continue
            ws.insert_rows(insert_at)
            self.style_row(ws, insert_at)
            self._set_value(ws, insert_at, COL_D, label)
            inserted.append(label)
            insert_at += 1

        if inserted:
            self.locator.clear_cache()
            self.log.info("   已自动补插材料费分组标题: %s", "、".join(inserted))

    def fix_stale_self_refs(self, ws):
        skip_words = {'成本合计', '成本及费用合计', '利润', '增值税实际税负', '税金及附加'}
        target_cols = {COL_J, COL_L, COL_O, COL_P, COL_Q, COL_R, COL_S}

        for row in range(11, ws.max_row + 1):
            d_val = self.locator.norm(str(ws.cell(row, COL_D).value or ''))
            if any(word in d_val for word in skip_words):
                continue
            for col in target_cols:
                cell = ws.cell(row, col)
                if not isinstance(cell.value, str) or not cell.value.startswith('='):
                    continue
                formula = cell.value
                if ':' in formula:
                    matched = re.fullmatch(r'=SUM\(T(\d+):V\1\)', formula, re.IGNORECASE)
                    if matched:
                        old_row = int(matched.group(1))
                        if old_row != row:
                            self._set_value(ws, row, cell.column, f'=SUM(T{row}:V{row})')
                    continue
                nums = [int(num) for num in re.findall(r'(?<=[A-Za-z])(\d+)', formula)]
                if not nums:
                    continue
                unique = set(nums)
                if len(unique) == 1:
                    old_row = unique.pop()
                    if old_row != row and old_row > 5:
                        self._set_value(ws, row, cell.column, re.sub(r'(?<=[A-Za-z])' + str(old_row) + r'(?!\d)', str(row), formula))


    def fix_all_sums(self, ws):
        """
        对所有成本节小计行，强制重写全部列的 SUM 公式，
        并强制重写锚点行和自身小计行的行内运算公式（O/P/Q/R/J/L 列）。
        """
        for hdr_kw, sub_kw in SUBTOTAL_MAP.items():
            hdr_row = self.locator.find_row(ws, hdr_kw)
            sub_row = self.locator.find_row(ws, sub_kw)
            if not hdr_row or not sub_row:
                continue

            anchor = None
            for row in range(sub_row - 1, max(hdr_row, sub_row - 400), -1):
                for col in range(2, 6):
                    if '财务未列部分需增列' in self.locator.norm(ws.cell(row, col).value):
                        anchor = row
                        break
                if anchor:
                    break
            end_row = (anchor - 1) if anchor else sub_row - 1
            if end_row >= sub_row:
                end_row = sub_row - 1

            start_row = hdr_row + 1
            for col in range(COL_H, COL_S):
                cell = ws.cell(sub_row, col)
                col_letter = get_column_letter(col)
                self._set_value(ws, sub_row, col, f'=SUM({col_letter}{start_row}:{col_letter}{end_row})')
                if col == COL_M and cell.number_format == 'General':
                    cell.number_format = '#,##0.00'

            self._set_value(ws, sub_row, COL_J, f'=H{sub_row}+I{sub_row}')
            self._set_value(ws, sub_row, COL_L, f'=J{sub_row}-K{sub_row}')
            self._set_value(ws, sub_row, COL_O, f'=L{sub_row}-M{sub_row}')
            self._set_value(ws, sub_row, COL_P, f'=K{sub_row}-N{sub_row}')
            self._set_value(ws, sub_row, COL_Q, f'=M{sub_row}+O{sub_row}')
            self._set_value(ws, sub_row, COL_R, f'=M{sub_row}+N{sub_row}+O{sub_row}+P{sub_row}')

            if anchor:
                self._set_value(ws, anchor, COL_J, f'=H{anchor}+I{anchor}')
                self._set_value(ws, anchor, COL_L, f'=J{anchor}-K{anchor}')
                self._set_value(ws, anchor, COL_O, f'=L{anchor}-M{anchor}')
                self._set_value(ws, anchor, COL_P, f'=K{anchor}-N{anchor}')
                self._set_value(ws, anchor, COL_Q, f'=M{anchor}+O{anchor}')
                self._set_value(ws, anchor, COL_R, f'=M{anchor}+N{anchor}+O{anchor}+P{anchor}')

            self.log.info("   🧮 %-14s H:R 重写 行%d→%d；S列以后保持模板原样",
                         sub_kw, start_row, end_row)

    def fix_aggregate_rows(self, ws):
        """
        强制重写汇总行的所有列公式。
        覆盖行：三、成本合计，四~七各固定项，八、研发费用，八/九、成本及费用合计
        """
        cost_total_row = self.locator.find_row(ws, '三、成本合计')
        if cost_total_row:
            subtotal_keywords = ['人工费小计', '分包工程小计', '材料费小计', '机械费小计', '其他直接费小计', '间接费小计', '安全费小计']
            refs = []
            for kw in subtotal_keywords:
                r = self.locator.find_row(ws, kw)
                if r and r != cost_total_row:
                    refs.append(r)
            if len(refs) == 7:
                for col in range(COL_H, COL_S):
                    cl = get_column_letter(col)
                    self._set_value(ws, cost_total_row, col, '=' + '+'.join(f'{cl}{r}' for r in refs))
                self.log.info("   🧮 三、成本合计 强制重写 行%d", cost_total_row)

        r_yanfa = self.locator.find_row(ws, '八、研发费用')
        if r_yanfa:
            r_nine = self.locator.find_row(ws, '九、成本及费用合计') or self.locator.find_row(ws, '八、成本及费用合计')
            if r_nine and r_nine > r_yanfa + 1:
                start_row = r_yanfa + 1
                end_row = r_nine - 1
                if end_row < start_row:
                    end_row = start_row
                for col in range(COL_H, COL_S):
                    cl = get_column_letter(col)
                    self._set_value(ws, r_yanfa, col, f'=SUM({cl}{start_row}:{cl}{end_row})')
                self._set_value(ws, r_yanfa, COL_J, f'=H{r_yanfa}+I{r_yanfa}')
                self._set_value(ws, r_yanfa, COL_L, f'=J{r_yanfa}-K{r_yanfa}')
                self._set_value(ws, r_yanfa, COL_O, f'=L{r_yanfa}-M{r_yanfa}')
                self._set_value(ws, r_yanfa, COL_P, f'=K{r_yanfa}-N{r_yanfa}')
                self._set_value(ws, r_yanfa, COL_Q, f'=M{r_yanfa}+O{r_yanfa}')
                self._set_value(ws, r_yanfa, COL_R, f'=M{r_yanfa}+N{r_yanfa}+O{r_yanfa}+P{r_yanfa}')
                self.log.info("   🧮 八、研发费用 强制重写 行%d", r_yanfa)

        for kw in ['四、计提保修金', '五、过程节点奖金', '六、资金占用费用', '七、局投资收益']:
            r = self.locator.find_row(ws, kw)
            if r:
                self._set_value(ws, r, COL_J, f'=H{r}+I{r}')
                self._set_value(ws, r, COL_L, f'=J{r}-K{r}')
                self._set_value(ws, r, COL_O, f'=L{r}-M{r}')
                self._set_value(ws, r, COL_P, f'=K{r}-N{r}')
                self._set_value(ws, r, COL_Q, f'=M{r}+O{r}')
                self._set_value(ws, r, COL_R, f'=M{r}+N{r}+O{r}+P{r}')

        total_row = self.locator.find_row(ws, '八、成本及费用合计') or self.locator.find_row(ws, '九、成本及费用合计')
        if total_row and cost_total_row:
            related_kws = ['三、成本合计', '四、计提保修金', '五、过程节点奖金', '六、资金占用费用', '七、局投资收益']
            related_rows = []
            for kw in related_kws:
                r = self.locator.find_row(ws, kw)
                if r and r != total_row:
                    related_rows.append(r)
            if related_rows:
                for col in range(COL_H, COL_S):
                    cl = get_column_letter(col)
                    self._set_value(ws, total_row, col, '=' + '+'.join(f'{cl}{r}' for r in related_rows))
                self.log.info("   🧮 成本及费用合计 强制重写 行%d", total_row)

    @staticmethod
    def _build_nested_lkp(df, v_col, c_col, amt_col, cat_col):
        """将 groupby 结果转为嵌套 dict: {(v, c): {cat: amount}}"""
        lkp = {}
        grouped = df.groupby([v_col, c_col, cat_col], dropna=False)[amt_col].sum()
        for (_v, _c, _cat), amt in grouped.items():
            v = str(int(float(_v))) if pd.notna(_v) and _v else ''
            c = str(_c or '').strip()
            cat = str(_cat or 'other')
            lkp.setdefault((v, c), {})[cat] = round(float(amt), 2)
        return lkp

    def build_lookups(self, df4):
        required_cols = {'总账科目长文本', '供应商', '借方本位币金额'}
        missing = required_cols - set(df4.columns)
        if missing:
            self.log.warning("   ⚠️ 合并表缺少字段，跳过 B/N/T 匹配：%s", "、".join(sorted(missing)))
            return {}, {}, {}, {}, {}

        # --- vendor_lkp: 客商名称 → 供应商编码（不受窜项影响，保持不变） ---
        ven_col = '供应商名称' if '供应商名称' in df4.columns else ('供应商描述' if '供应商描述' in df4.columns else None)
        vendor_lkp = {}
        if ven_col:
            subset = df4[df4[ven_col].notna() & df4['供应商'].notna()]
            dedup_cols = [ven_col, '合同'] if '合同' in subset.columns else [ven_col]
            value_cols = [ven_col, '供应商'] + (['合同'] if '合同' in subset.columns else [])
            for row in subset.drop_duplicates(dedup_cols)[value_cols].itertuples(index=False, name=None):
                ven_raw = row[0]
                code_raw = row[1]
                con_raw = row[2] if len(row) > 2 else ''
                ven = self.locator.norm(str(ven_raw))
                con = self.locator.norm(str(con_raw or ''))
                code = str(int(float(code_raw))) if pd.notna(code_raw) else ''
                if ven:
                    vendor_lkp[(ven, con)] = code
                    if (ven, '') not in vendor_lkp:
                        vendor_lkp[(ven, '')] = code

        # --- 构建 中台单据号 → 成本大类 雷达字典（供 safe/rd 兜底使用） ---
        has_doc = '中台单据号' in df4.columns
        doc_cat_lkp = {}
        if has_doc:
            cost_mask = df4['总账科目长文本'].fillna('').str.contains(
                r'合同履约成本\\工程施工成本', na=False
            )
            cost_df = df4[cost_mask].copy()
            cost_df['_doc'] = (
                cost_df['中台单据号'].fillna('').astype(str).str.strip().str.upper()
            )
            cost_df['_cat'] = cost_df['总账科目长文本'].apply(self._classify_gl_to_section)
            valid = cost_df[cost_df['_doc'] != '']
            for doc, grp in valid.groupby('_doc', sort=False):
                if not grp.empty:
                    doc_cat_lkp[doc] = grp['_cat'].value_counts().index[0]

        def _norm_v(df):
            return df['供应商'].fillna(0).apply(
                lambda value: str(int(float(value))) if pd.notna(value) and value else ''
            )

        def _norm_c(df):
            return df['合同'].fillna('').astype(str).str.strip()

        # --- pay_lkp: 应付账款 → T列（保持原有(供应商,合同)扁平匹配，不做分类拆分） ---
        ap_mask = (
            df4['总账科目长文本'].fillna('').str.contains('应付账款')
            & ~df4['总账科目长文本'].fillna('').str.contains('待确认进项税额')
        )
        ap = df4[ap_mask].copy()
        ap['_v'] = ap['供应商'].fillna(0).apply(lambda value: str(int(float(value))) if pd.notna(value) and value else '')
        ap['_c'] = ap['合同'].fillna('').astype(str).str.strip()
        pay_lkp = dict(ap.groupby(['_v', '_c'])['借方本位币金额'].sum())

        # --- vat_lkp: 待确认进项税额 → N列（保持原有(供应商,合同)扁平匹配，不做分类拆分） ---
        vat = df4[
            df4['总账科目长文本'].fillna('').str.contains('其他应收款')
            & df4['总账科目长文本'].fillna('').str.contains('待确认进项税额')
        ].copy()
        vat['_v'] = vat['供应商'].fillna(0).apply(lambda value: str(int(float(value))) if pd.notna(value) and value else '')
        vat['_c'] = vat['合同'].fillna('').astype(str).str.strip()
        vat_lkp = dict(vat.groupby(['_v', '_c'])['借方本位币金额'].sum())

        # --- safe_lkp: 安全生产费 → M列加回，优先用对方科目描述分类，兜底用中台单据号 ---
        safe_mask = (
            df4['总账科目长文本'].fillna('').str.contains('专项储备')
            & df4['总账科目长文本'].fillna('').str.contains('安全生产费')
            & df4['总账科目长文本'].fillna('').str.contains('发生数')
        )
        safe_df = df4[safe_mask].copy()
        safe_df['_v'] = _norm_v(safe_df)
        safe_df['_c'] = _norm_c(safe_df)

        opp_col = '对方科目描述' if '对方科目描述' in df4.columns else (
            '对方科目名称' if '对方科目名称' in df4.columns else None
        )
        if opp_col:
            safe_df['_cat'] = safe_df[opp_col].fillna('').apply(self._classify_safety_by_opponent)
        else:
            safe_df['_cat'] = ''

        # 兜底：对方科目分类为空的，用 中台单据号 回查
        fallback_mask = safe_df['_cat'] == ''
        if fallback_mask.any() and has_doc:
            safe_df.loc[fallback_mask, '_doc'] = (
                safe_df.loc[fallback_mask, '中台单据号']
                .fillna('').astype(str).str.strip().str.upper()
            )
            safe_df.loc[fallback_mask, '_cat'] = (
                safe_df.loc[fallback_mask, '_doc'].map(doc_cat_lkp).fillna('other')
            )
        else:
            safe_df.loc[fallback_mask, '_cat'] = 'other'

        safe_lkp = self._build_nested_lkp(safe_df, '_v', '_c', '借方本位币金额', '_cat')

        # --- rd_lkp: 研发支出租赁及运行维护费 → M列加回（机械节），同时覆盖全部研发子类 ---
        rd_mask = (
            df4['总账科目长文本'].fillna('').str.contains(r'研发支出', regex=True)
        )
        rd_df = df4[rd_mask].copy()
        rd_df['_v'] = _norm_v(rd_df)
        rd_df['_c'] = _norm_c(rd_df)
        # 用总账科目末级分类
        rd_df['_cat'] = rd_df['总账科目长文本'].fillna('').apply(
            lambda x: 'machine' if '租赁及运行维护' in str(x) else (
                'material' if '材料费' in str(x) else (
                    'labor' if '人工' in str(x) else 'other'
                )
            )
        )
        rd_lkp = self._build_nested_lkp(rd_df, '_v', '_c', '借方本位币金额', '_cat')

        return vendor_lkp, pay_lkp, vat_lkp, safe_lkp, rd_lkp

    def fill_per_row_cols(self, ws, vendor_lkp, pay_lkp, vat_lkp, safe_lkp, rd_lkp):
        skip_words = {
            '人工费小计', '分包工程小计', '材料费小计', '机械费小计', '其他直接费小计',
            '间接费小计', '安全费小计', '财务未列部分需增列', '成本合计', '项目效益',
            '成本及费用合计', '增值税', '税金及附加', '利润', '保修金', '节点奖'
        }
        current_section = None
        for row in range(11, ws.max_row + 1):
            b_val = ws.cell(row, COL_B).value
            d_val = ws.cell(row, COL_D).value
            m_val = ws.cell(row, COL_M).value
            c_val = ws.cell(row, COL_C).value
            d_norm = self.locator.norm(str(d_val or ''))
            current_section = self._SECTION_HEADER_MAP.get(d_norm, current_section)
            if not d_val or not isinstance(m_val, (int, float)):
                continue
            if any(word in self.locator.norm(str(d_val)) for word in skip_words):
                continue

            c_norm = self.locator.norm(str(c_val or ''))
            code = self._normalize_code(b_val)
            if not code:
                code = vendor_lkp.get((d_norm, c_norm)) or vendor_lkp.get((d_norm, ''), '')

            if not code and d_norm in VENDOR_SHORTNAME_MAP:
                mapped = VENDOR_SHORTNAME_MAP[d_norm]
                code = vendor_lkp.get((self.locator.norm(mapped), c_norm)) or vendor_lkp.get((self.locator.norm(mapped), ''))
            if not code and d_norm:
                log_debug = getattr(self.log, 'debug', None)
                if callable(log_debug):
                    log_debug("未匹配到供应商编码：行%d D='%s' C='%s'", row, d_val, c_val)

            if code:
                self._set_value(ws, row, COL_B, code)

                pay = pay_lkp.get((code, c_norm), 0)
                if pay:
                    cell = ws.cell(row, COL_T)
                    self._set_value(ws, row, COL_T, round(float(pay), 2))
                    cell.number_format = '#,##0.00'
                vat = vat_lkp.get((code, c_norm), 0)
                if vat:
                    cell = ws.cell(row, COL_N)
                    self._set_value(ws, row, COL_N, round(float(vat), 2))
                    cell.number_format = '#,##0.00'

                m_add = 0.0
                if current_section in {'labor', 'vendor', 'machine'}:
                    m_add += float(safe_lkp.get((code, c_norm), {}).get(current_section, 0) or 0)
                if current_section == 'machine':
                    m_add += float(rd_lkp.get((code, c_norm), {}).get('machine', 0) or 0)
                if m_add != 0:
                    current_m = ws.cell(row, COL_M).value or 0
                    self._set_value(ws, row, COL_M, round(float(current_m) + m_add, 2))
                    ws.cell(row, COL_M).number_format = '#,##0.00'

    def apply_invoice_match(self, ws, result_dir: str, invoice_ledger_path: str | None):
        invoice_path = invoice_ledger_path
        if not invoice_path:
            try:
                for filename in os.listdir(result_dir):
                    if filename.lower().endswith((".xlsx", ".xls")) and any(key in filename for key in ["发票", "收票", "台账", "台帳"]):
                        invoice_path = os.path.join(result_dir, filename)
                        break
            except Exception:
                invoice_path = None

        if not (invoice_path and os.path.exists(invoice_path)):
            self.log.info("   ⓘ 未提供发票台账，跳过 AB 列发票匹配")
            return

        from invoice_matcher import InvoiceMatcher

        matcher = InvoiceMatcher(self.log)
        matcher.apply_to_worksheet(ws, invoice_path)

    def renumber(self, ws):
        sections = [
            ('（一）人工费', '人工费小计'),
            ('（二）分包工程', '分包工程小计'),
            ('(三)材料费', '材料费小计'),
            ('(四）机械租赁费', '机械费小计'),
        ]
        for hdr_kw, sub_kw in sections:
            hdr = self.locator.find_row(ws, hdr_kw)
            sub = self.locator.find_row(ws, sub_kw)
            if not hdr or not sub:
                continue
            anchor = None
            for row in range(sub - 1, hdr, -1):
                for col in range(2, 6):
                    if '财务未列部分需增列' in self.locator.norm(ws.cell(row, col).value):
                        anchor = row
                        break
                if anchor:
                    break
            end = (anchor - 1) if anchor else (sub - 1)
            for row in range(hdr + 1, sub):
                self._set_value(ws, row, COL_A, None)
            seq = 1
            for row in range(hdr + 1, end + 1):
                d_val = ws.cell(row, COL_D).value
                m_val = ws.cell(row, COL_M).value
                if d_val and isinstance(m_val, (int, float)) and m_val != 0:
                    self._set_value(ws, row, COL_A, seq)
                    seq += 1
