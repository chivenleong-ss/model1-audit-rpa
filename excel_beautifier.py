# -*- coding: utf-8 -*-
"""
excel_beautifier.py v3.3 - 完全重写公式修复逻辑

修复策略（关键：先删行再强制重写，避免旧行号残留）：
  1. 先删除无效行（尾空行、研发空占位、成本零金额行）
  2. 根据删除后的实际行号，强制重写所有关键 SUM 范围
  3. 强制重写九合计、十二利润、十三利润率公式（不依赖原公式存在与否）
  4. 重写锚点行 O/P/Q/R 列公式，确保不引用自身
  5. 最后清理残留的纯自引用

保留原有风格修复功能：L列填充同步、研发小计、序号补全、边框、汇总行格式等。
"""

import os, re, traceback
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.styles import Border, Side, PatternFill
import copy as py_copy

GREY_SET = {"FFBFBFBF", "FFD9D9D9", "FFC0C0C0", "FFE0E0E0","FFD3D3D3", "FFCCCCCC", "FFB8B8B8", "FFAEAAAA"}
WHITE_FILL = PatternFill(start_color="FFFFFFFF", end_color="FFFFFFFF", fill_type="solid")
COL_S = 19

class ExcelBeautifier:

    def __init__(self, result_dir, logger):
        self.result_dir  = result_dir
        self.log         = logger
        self.input_path  = os.path.join(result_dir, "自动填报完成_效益审核表.xlsx")
        self.output_path = os.path.join(result_dir, "效益审核表（已填）.xlsx")

    @staticmethod
    def _set_value(ws, row: int, col: int, value):
        if row <= 10:
            return False
        ws.cell(row, col).value = value
        return True

    @staticmethod
    def _norm(val) -> str:
        if val is None: return ""
        return (str(val).replace(" ", "").replace("\u3000", "").replace("（", "(").replace("）", ")").replace("：", ":").replace("*", "").replace(":", ""))

    def _find_row(self, ws, keyword, cols=(2, 3, 4, 5)):
        kw = self._norm(keyword)
        for r in range(1, ws.max_row + 1):
            for c in cols:
                val = ws.cell(r, c).value
                if val is not None and kw in self._norm(val):
                    return r
        return None

    @staticmethod
    def _get_fill_rgb(cell) -> str:
        try:
            f = cell.fill
            if f and f.fill_type not in (None, "none"):
                return f.fgColor.rgb
        except Exception:
            pass
        return ""

    # ====================================================================
    #  删除无效行
    # ====================================================================
    def _delete_trailing_empty_rows(self, ws):
        deleted = 0
        for r in range(ws.max_row, 10, -1):
            if all(ws.cell(r, c).value is None for c in range(1, 40)):
                ws.delete_rows(r)
                deleted += 1
            else:
                break
        if deleted:
            self.log.info("   删除尾部空行 %d 行", deleted)

    def _delete_yanfa_ghost_rows(self, ws):
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine  = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        if not r_yanfa or not r_nine:
            return
        to_del = []
        for r in range(r_yanfa + 1, r_nine):
            d = ws.cell(r, 4).value
            m = ws.cell(r, 13).value
            d_empty = (d is None) or (isinstance(d, str) and not d.strip())
            if d_empty and not isinstance(m, (int, float)):
                to_del.append(r)
        for r in reversed(to_del):
            ws.delete_rows(r)
        if to_del:
            self.log.info("   删除研发费用节空占位行 %d 行", len(to_del))

    def _delete_cost_zero_m_rows(self, ws):
        SECTIONS = [
            ('（一）人工费',      '人工费小计'),
            ('（二）分包工程',    '分包工程小计'),
            ('(三)材料费',       '材料费小计'),
            ('(四）机械租赁费',   '机械费小计'),
            ('（五）其他直接费',  '其他直接费小计'),
            ('（六）间接费',      '间接费小计'),
            ('（七）安全费',      '安全费小计'),
        ]
        total_deleted = 0
        for hdr_kw, sub_kw in SECTIONS:
            hdr = self._find_row(ws, hdr_kw)
            sub = self._find_row(ws, sub_kw)
            if not hdr or not sub or sub <= hdr:
                continue
            anchor = None
            for r in range(sub - 1, hdr, -1):
                for c in range(2, 6):
                    if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                        anchor = r
                        break
                if anchor:
                    break
            end_row = (anchor - 1) if anchor else (sub - 1)
            to_del = []
            for r in range(hdr + 1, end_row + 1):
                d_val = ws.cell(r, 4).value
                m_val = ws.cell(r, 13).value
                has_d = d_val is not None and (not isinstance(d_val, str) or d_val.strip())
                m_empty = (m_val is None) or (isinstance(m_val, (int, float)) and m_val == 0)
                if has_d and m_empty:
                    to_del.append(r)
            for r in reversed(to_del):
                ws.delete_rows(r)
            total_deleted += len(to_del)
            if to_del:
                self.log.info("   清除 %s 节无效行 %d 行", hdr_kw, len(to_del))
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine  = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        if r_yanfa and r_nine and r_nine > r_yanfa + 1:
            to_del = []
            for r in range(r_yanfa + 1, r_nine):
                d_val = ws.cell(r, 4).value
                m_val = ws.cell(r, 13).value
                has_d = d_val is not None and (not isinstance(d_val, str) or d_val.strip())
                m_empty = (m_val is None) or (isinstance(m_val, (int, float)) and m_val == 0)
                if has_d and m_empty:
                    to_del.append(r)
            for r in reversed(to_del):
                ws.delete_rows(r)
            if to_del:
                self.log.info("   清除 八、研发费用 节无效行 %d 行", len(to_del))
                total_deleted += len(to_del)
        if total_deleted:
            self.log.info("   总计清除成本节内空金额行 %d 行", total_deleted)

    # ====================================================================
    #  公式修复核心：删除行后强制重写，基于当前实际行号
    # ====================================================================
    def _fix_cost_section_sums(self, ws):
        """
        对每个成本节（一～七），强制重写小计行的所有列公式，
        SUM 范围从节标题下一行到锚点行或小计行的前一行的上一行，
        确保 end_row < 小计行自身行号。
        """
        SECTIONS = [
            ('（一）人工费',      '人工费小计'),
            ('（二）分包工程',    '分包工程小计'),
            ('(三)材料费',       '材料费小计'),
            ('(四）机械租赁费',   '机械费小计'),
            ('（五）其他直接费',  '其他直接费小计'),
            ('（六）间接费',      '间接费小计'),
            ('（七）安全费',      '安全费小计'),
        ]
        COL_O, COL_P, COL_Q, COL_R = 15, 16, 17, 18
        fixed = 0
        for hdr_kw, sub_kw in SECTIONS:
            hdr = self._find_row(ws, hdr_kw)
            sub = self._find_row(ws, sub_kw)
            if not hdr or not sub or sub <= hdr + 1:
                continue
            # 找到锚点行（财务未列部分需增列）
            anchor = None
            for r in range(sub - 1, hdr, -1):
                for c in range(2, 6):
                    if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                        anchor = r
                        break
                if anchor:
                    break
            # 确定 SUM 范围：从 hdr+1 到 anchor-1，如果没找到锚点则到 sub-1
            start_row = hdr + 1
            end_row = (anchor - 1) if anchor else (sub - 1)
            # 核心：确保 end_row < sub（小计行不包含自身）
            if end_row >= sub:
                end_row = sub - 1
            if start_row > end_row:
                continue

            for col_idx in range(1, COL_S):
                cell = ws.cell(sub, col_idx)
                cl = get_column_letter(col_idx)
                if col_idx in (8, 9, 11, 13, 14):  # H/I/K/M/N 列汇总明细区域
                    self._set_value(ws, sub, col_idx, f"=SUM({cl}{start_row}:{cl}{end_row})")
                    if col_idx == 13:
                        cell.number_format = '#,##0.00' if cell.number_format == 'General' else cell.number_format
                    fixed += 1
                elif col_idx == 10:  # J 列保持模板行内公式
                    self._set_value(ws, sub, col_idx, f"=H{sub}+I{sub}")
                    fixed += 1
                elif col_idx == 12:  # L 列保持模板行内公式
                    self._set_value(ws, sub, col_idx, f"=J{sub}-K{sub}")
                    fixed += 1
                elif col_idx == COL_O:
                    self._set_value(ws, sub, col_idx, f"=L{sub}-M{sub}")
                    fixed += 1
                elif col_idx == COL_P:
                    self._set_value(ws, sub, col_idx, f"=K{sub}-N{sub}")
                    fixed += 1
                elif col_idx == COL_Q:
                    self._set_value(ws, sub, col_idx, f"=M{sub}+O{sub}")
                    fixed += 1
                elif col_idx == COL_R:
                    self._set_value(ws, sub, col_idx, f"=M{sub}+N{sub}+O{sub}+P{sub}")
                    fixed += 1
            self.log.info("   %-12s SUM 强制重写 行%d = SUM(%d:%d)，J/L 保持模板行内公式", sub_kw, sub, start_row, end_row)
        if not fixed:
            self.log.info("   未找到成本节小计行，跳过 SUM 重写")

    def _fix_yanfa_sum(self, ws):
        """强制重写八、研发费用行的 SUM 公式，基于当前实际行号"""
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine  = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        if not r_yanfa or not r_nine or r_nine <= r_yanfa + 1:
            return
        start_row = r_yanfa + 1
        end_row = r_nine - 1
        if start_row > end_row:
            end_row = start_row
        COL_J, COL_L = 10, 12
        COL_O, COL_P, COL_Q, COL_R = 15, 16, 17, 18
        for col_idx in range(1, COL_S):
            cell = ws.cell(r_yanfa, col_idx)
            cl = get_column_letter(col_idx)
            if col_idx in (8, 9, 11, 13, 14):
                self._set_value(ws, r_yanfa, col_idx, f"=SUM({cl}{start_row}:{cl}{end_row})")
            elif col_idx == COL_J:
                self._set_value(ws, r_yanfa, col_idx, f"=H{r_yanfa}+I{r_yanfa}")
            elif col_idx == COL_L:
                self._set_value(ws, r_yanfa, col_idx, f"=J{r_yanfa}-K{r_yanfa}")
            elif col_idx == COL_O:
                self._set_value(ws, r_yanfa, col_idx, f"=L{r_yanfa}-M{r_yanfa}")
            elif col_idx == COL_P:
                self._set_value(ws, r_yanfa, col_idx, f"=K{r_yanfa}-N{r_yanfa}")
            elif col_idx == COL_Q:
                self._set_value(ws, r_yanfa, col_idx, f"=M{r_yanfa}+O{r_yanfa}")
            elif col_idx == COL_R:
                self._set_value(ws, r_yanfa, col_idx, f"=M{r_yanfa}+N{r_yanfa}+O{r_yanfa}+P{r_yanfa}")
        self.log.info("   研发费用 强制重写 行%d = SUM(%d:%d)", r_yanfa, start_row, end_row)

    def _fix_extended_section_sums(self, ws):
        """
        单独修复 S 列及之后的成本节小计公式。
        这些列属于债权债务、税务等独立区域，只允许按本节明细范围求和，禁止跨节。
        """
        if ws.max_column < COL_S:
            return
        sections = [
            ('（一）人工费',      '人工费小计'),
            ('（二）分包工程',    '分包工程小计'),
            ('(三)材料费',       '材料费小计'),
            ('(四）机械租赁费',   '机械费小计'),
            ('（五）其他直接费',  '其他直接费小计'),
            ('（六）间接费',      '间接费小计'),
            ('（七）安全费',      '安全费小计'),
        ]
        section_headers = [self._find_row(ws, hdr_kw) for hdr_kw, _ in sections]
        section_headers = sorted(r for r in section_headers if r)

        fixed = 0
        for hdr_kw, sub_kw in sections:
            hdr = self._find_row(ws, hdr_kw)
            sub = self._find_row(ws, sub_kw)
            if not hdr or not sub or sub <= hdr + 1:
                continue

            next_headers = [r for r in section_headers if r > hdr]
            next_hdr = min(next_headers) if next_headers else None
            anchor = None
            for r in range(sub - 1, hdr, -1):
                for c in range(2, 6):
                    if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                        anchor = r
                        break
                if anchor:
                    break

            start_row = hdr + 1
            end_row = (anchor - 1) if anchor else (sub - 1)
            if end_row >= sub:
                end_row = sub - 1
            if next_hdr and end_row >= next_hdr:
                self.log.error(
                    "   S列后公式跨节风险：%s 行%d:%d 超过下一节标题行%d，放弃写入",
                    sub_kw, start_row, end_row, next_hdr,
                )
                continue
            if start_row > end_row:
                continue

            for col_idx in range(COL_S, ws.max_column + 1):
                cl = get_column_letter(col_idx)
                self._set_value(ws, sub, col_idx, f"=SUM({cl}{start_row}:{cl}{end_row})")
                fixed += 1
            self.log.info("   %-12s S列后节内 SUM 行%d = SUM(%d:%d)", sub_kw, sub, start_row, end_row)
        if not fixed:
            self.log.info("   S列后节内 SUM 未发现需修复列或小计行")

    def _fix_anchor_rows(self, ws):
        """
        清空锚点行（“财务未列部分需增列”）的行内公式列，
        避免这些占位行出现自引用或被误识别为有效计算行。
        """
        found = 0
        for r in range(11, ws.max_row + 1):
            found_anchor = False
            for c in range(2, 6):
                if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                    found_anchor = True
                    break
            if not found_anchor:
                continue
            for col in (8, 9, 10, 11, 12, 14, 15, 16, 17, 18):
                self._set_value(ws, r, col, None)
            found += 1
        if found:
            self.log.info("   锚点行公式清空 %d 行", found)

    def _fix_fixed_amount_rows(self, ws):
        """重写固定金额行的行内公式，避免删行后残留旧行号。"""
        formula_rows = [
            '四、计提保修金',
            '五、过程节点奖金',
            '六、资金占用费用',
            '七、局投资收益',
        ]
        tax_rows = [
            '十、增值税实际税负',
            '十一、税金及附加',
        ]
        COL_J, COL_L = 10, 12
        COL_O, COL_P, COL_Q, COL_R = 15, 16, 17, 18
        fixed = 0
        for kw in formula_rows:
            r = self._find_row(ws, kw)
            if not r:
                continue
            for col in (8, 9, 10, 11, 12, 14):
                self._set_value(ws, r, col, None)
            self._set_value(ws, r, COL_O, f"=0-M{r}")
            self._set_value(ws, r, COL_P, 0)
            self._set_value(ws, r, COL_Q, f"=M{r}+O{r}")
            self._set_value(ws, r, COL_R, f"=M{r}+O{r}+P{r}")
            fixed += 1
        for kw in tax_rows:
            r = self._find_row(ws, kw)
            if not r:
                continue
            self._set_value(ws, r, COL_J, f"=H{r}+I{r}")
            self._set_value(ws, r, COL_L, f"=J{r}-K{r}")
            self._set_value(ws, r, COL_O, f"=L{r}-M{r}")
            self._set_value(ws, r, COL_P, f"=K{r}-N{r}")
            self._set_value(ws, r, COL_Q, f"=M{r}+O{r}")
            self._set_value(ws, r, COL_R, f"=M{r}+N{r}+O{r}+P{r}")
            fixed += 1
        if fixed:
            self.log.info("   固定金额/税费行公式重写 %d 行", fixed)

    def _fix_nine_formula(self, ws):
        """
        强制重写九、成本及费用合计公式，基于当前实际行号，
        确保引用的行号 != 自身行号。
        """
        ROW_KWS = ['三、成本合计','四、计提保修金','五、过程节点奖金','六、资金占用费用','七、局投资收益','八、研发费用']
        row_nums = []
        for kw in ROW_KWS:
            r = self._find_row(ws, kw)
            if r:
                row_nums.append(r)
        r_nine = self._find_row(ws, '九、成本及费用合计')
        if not r_nine or not row_nums:
            self.log.warning("   未能定位九、成本及费用合计，跳过")
            return
        safe_rows = [r for r in row_nums if r != r_nine]
        if not safe_rows:
            self.log.warning("   九、成本及费用合计 所有引用行均等于自身行，跳过")
            return
        for col_idx in range(8, COL_S):
            cl = get_column_letter(col_idx)
            self._set_value(ws, r_nine, col_idx, '=' + '+'.join(f'{cl}{r}' for r in safe_rows))
        self.log.info("   九、成本及费用合计 强制重写（行%d，引用%s）", r_nine, safe_rows)

    def _fix_cost_total_formula(self, ws):
        """
        强制重写三、成本合计，避免把自身行纳入求和而形成循环引用。
        """
        subtotal_keywords = [
            '人工费小计',
            '分包工程小计',
            '材料费小计',
            '机械费小计',
            '其他直接费小计',
            '间接费小计',
            '安全费小计',
        ]
        r_cost = self._find_row(ws, '三、成本合计')
        if not r_cost:
            return
        refs = []
        for kw in subtotal_keywords:
            r = self._find_row(ws, kw)
            if r and r != r_cost:
                refs.append(r)
        if not refs:
            self.log.warning("   未能定位成本节小计，跳过三、成本合计重写")
            return
        for col_idx in range(8, COL_S):
            cl = get_column_letter(col_idx)
            self._set_value(ws, r_cost, col_idx, '=' + '+'.join(f'{cl}{r}' for r in refs))
        self.log.info("   三、成本合计 强制重写（行%d，引用%s）", r_cost, refs)

    def _fix_profit_formulas(self, ws):
        """强制重写十二利润、十三利润率公式"""
        r6 = 6
        r_nine    = self._find_row(ws, '九、成本及费用合计')
        r_vat     = self._find_row(ws, '十、增值税实际税负')
        r_tax     = self._find_row(ws, '十一、税金及附加')
        r_profit  = self._find_row(ws, '十二、利润')
        r_rate    = self._find_row(ws, '十三、利润率')
        if r_profit and r_nine:
            deduct = [r for r in [r_nine, r_vat, r_tax] if r and r != r_profit]
            for col_idx in range(8, COL_S):
                cl = get_column_letter(col_idx)
                self._set_value(ws, r_profit, col_idx, f'={cl}{r6}' + ''.join(f'-{cl}{r}' for r in deduct))
            self.log.info("   十二、利润 强制重写（行%d）", r_profit)
        if r_rate and r_profit:
            for col_idx in range(8, COL_S):
                cl = get_column_letter(col_idx)
                self._set_value(ws, r_rate, col_idx, f'={cl}{r_profit}/{cl}{r6}')
            self.log.info("   十三、利润率 强制重写（行%d）", r_rate)

    def _rewrite_detail_row_formulas(self, ws):
        """
        重写所有非小计行、非锚点行、非标题行的 O/P/Q/R 列公式，
        确保使用当前行的行号。
        这解决了删除/插入行后公式行号残留旧行号的问题。
        """
        # 跳过行的关键词（这些行的公式由 _fix_cost_section_sums / _fix_yanfa_sum 等处理）
        skip_keywords = {
            '人工费小计', '分包工程小计', '材料费小计', '机械费小计',
            '其他直接费小计', '间接费小计', '安全费小计',
            '财务未列部分需增列', '一、', '二、', '三、', '四、', '五、',
            '六、', '七、', '八、', '九、', '十、', '十一、', '十二、', '十三、',
            '收入', '增值税', '成本合计', '成本及费用合计', '利润', '利润率',
            '税金及附加', '局投资收益', '资金占用', '计提保修金', '过程节点',
        }
        COL_J, COL_L = 10, 12
        COL_O, COL_P, COL_Q, COL_R = 15, 16, 17, 18
        fixed = 0
        for r in range(11, ws.max_row + 1):
            d_val = ws.cell(r, 4).value
            m_val = ws.cell(r, 13).value
            # 只在“明细行”重写：D列有文本值 且 M列为数值
            if not d_val or not isinstance(m_val, (int, float)):
                continue
            # 跳过标题/小计/锚点行
            d_norm = self._norm(str(d_val))
            if any(kw in d_norm for kw in skip_keywords):
                continue

            # 检查 O/P/Q/R 列，如果缺少公式或公式行号 != 当前行，则重写
            changed = False
            # O 列
            cell_o = ws.cell(r, COL_O)
            if not cell_o.value or not isinstance(cell_o.value, str) or f"L{r}" not in str(cell_o.value):
                cell_o.value = f"=L{r}-M{r}"
                changed = True
            # P 列
            cell_p = ws.cell(r, COL_P)
            if not cell_p.value or not isinstance(cell_p.value, str) or f"K{r}" not in str(cell_p.value):
                cell_p.value = f"=K{r}-N{r}"
                changed = True
            # Q 列
            cell_q = ws.cell(r, COL_Q)
            if not cell_q.value or not isinstance(cell_q.value, str) or f"M{r}" not in str(cell_q.value):
                cell_q.value = f"=M{r}+O{r}"
                changed = True
            # R 列
            cell_r = ws.cell(r, COL_R)
            if not cell_r.value or not isinstance(cell_r.value, str) or f"M{r}" not in str(cell_r.value):
                cell_r.value = f"=M{r}+N{r}+O{r}+P{r}"
                changed = True
            if changed:
                fixed += 1
        if fixed:
            self.log.info("   明细行 O/P/Q/R 列公式重写 %d 行（确保行号正确）", fixed)

    def _normalize_row_arithmetic_formulas(self, ws):
        """
        对仍然残留旧行号的 J/L/O/P/Q/R 行内公式做统一归正。
        只处理当前已经存在这些公式痕迹的行，避免误伤纯空白区域。
        """
        subtotal_keywords = [
            '人工费小计', '分包工程小计', '材料费小计', '机械费小计',
            '其他直接费小计', '间接费小计', '安全费小计',
        ]
        fixed_keywords = [
            '四、计提保修金', '五、过程节点奖金', '六、资金占用费用',
            '七、局投资收益', '八、研发费用',
        ]
        protected_rows = {
            self._find_row(ws, '三、成本合计'),
            self._find_row(ws, '九、成本及费用合计'),
            self._find_row(ws, '十、增值税实际税负'),
            self._find_row(ws, '十一、税金及附加'),
            self._find_row(ws, '十二、利润'),
            self._find_row(ws, '十三、利润率'),
        }
        protected_rows.update(self._find_row(ws, kw) for kw in subtotal_keywords)
        protected_rows.update(self._find_row(ws, kw) for kw in fixed_keywords)
        protected_rows = {r for r in protected_rows if r}
        target_cols = (10, 12, 15, 16, 17, 18)
        fixed = 0
        for r in range(11, ws.max_row + 1):
            if r in protected_rows:
                continue
            has_formula_trace = any(
                isinstance(ws.cell(r, c).value, str) and ws.cell(r, c).value.startswith('=')
                for c in target_cols
            )
            if not has_formula_trace:
                continue
            d_norm = self._norm(str(ws.cell(r, 4).value or ''))
            if '财务未列部分需增列' in d_norm:
                continue
            self._set_value(ws, r, 10, f'=H{r}+I{r}')
            self._set_value(ws, r, 12, f'=J{r}-K{r}')
            self._set_value(ws, r, 15, f'=L{r}-M{r}')
            self._set_value(ws, r, 16, f'=K{r}-N{r}')
            self._set_value(ws, r, 17, f'=M{r}+O{r}')
            self._set_value(ws, r, 18, f'=M{r}+N{r}+O{r}+P{r}')
            fixed += 1
        if fixed:
            self.log.info("   行内公式归正 %d 行（清理残留旧行号）", fixed)

    def _clean_circular_refs(self, ws):
        """
        兜底清理纯自引用公式（=A1 等情况），排除 SUM 范围（已由 _fix_cost_section_sums 等处理）。
        """
        cleared = 0
        for r in range(11, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                cell = ws.cell(r, c)
                if cell.value is None or not isinstance(cell.value, str) or not cell.value.startswith("="):
                    continue
                cl = get_column_letter(c)
                formula_upper = cell.value.upper()
                self_ref = f"{cl}{r}"
                # 纯自引用
                if formula_upper == f"={self_ref}":
                    self._set_value(ws, r, c, 0)
                    cleared += 1
                    continue
                # 非 SUM 公式中包含自身单元格地址
                if "SUM(" not in formula_upper and self_ref in formula_upper:
                    self._set_value(ws, r, c, 0)
                    cleared += 1
        if cleared:
            self.log.info("   兜底清理自引用 %d 个", cleared)

    # ====================================================================
    #  风格修复（保持原逻辑不变）
    # ====================================================================
    def _sync_l_fill_to_j(self, ws):
        NO_FILL = PatternFill(fill_type=None)
        fixed = 0
        for r in range(7, ws.max_row + 1):
            j_cell = ws.cell(r, 10)
            l_cell = ws.cell(r, 12)
            j_rgb  = self._get_fill_rgb(j_cell)
            l_rgb  = self._get_fill_rgb(l_cell)
            if j_rgb == l_rgb:
                continue
            try:
                l_cell.fill = py_copy.copy(j_cell.fill) if j_rgb else NO_FILL
                fixed += 1
            except Exception:
                pass
        self.log.info("   L列填充色与J列同步 (%d 个单元格)", fixed)

    def _fill_sequence_numbers(self, ws):
        SECTIONS = [
            ('（一）人工费','人工费小计'),('（二）分包工程','分包工程小计'),
            ('(三)材料费','材料费小计'),('(四）机械租赁费','机械费小计'),
            ('（五）其他直接费','其他直接费小计'),('（六）间接费','间接费小计'),
            ('（七）安全费','安全费小计')
        ]
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine  = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        def _number_section(hdr, end_excl):
            for r in range(max(hdr + 1, 11), end_excl):
                self._set_value(ws, r, 1, None)
            seq = 1
            for r in range(max(hdr + 1, 11), end_excl):
                d = ws.cell(r, 4).value
                m = ws.cell(r, 13).value
                if d and isinstance(m, (int, float)):
                    self._set_value(ws, r, 1, seq)
                    seq += 1
        for hdr_kw, sub_kw in SECTIONS:
            hdr = self._find_row(ws, hdr_kw)
            sub = self._find_row(ws, sub_kw)
            if not hdr or not sub:
                continue
            anchor = None
            for r in range(sub - 1, hdr, -1):
                for c in range(2, 6):
                    if '财务未列部分需增列' in self._norm(ws.cell(r, c).value):
                        anchor = r; break
                if anchor: break
            end = (anchor - 1) if anchor else (sub - 1)
            _number_section(hdr, end + 1)
        if r_yanfa and r_nine:
            _number_section(r_yanfa, r_nine)
        self.log.info("   A列序号全节补全完成")

    def _fix_yanfa_detail_fill(self, ws):
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine  = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        if not r_yanfa or not r_nine:
            return
        no_fill = PatternFill(fill_type=None)
        ref_fill_c = ref_fill_d = None
        for r in range(r_yanfa + 1, r_nine):
            d = ws.cell(r, 4).value
            m = ws.cell(r, 13).value
            if d and isinstance(m, (int, float)):
                ref_fill_c = py_copy.copy(ws.cell(r, 3).fill)
                ref_fill_d = py_copy.copy(ws.cell(r, 4).fill)
                break
        fixed = 0
        for r in range(r_yanfa + 1, r_nine):
            d = ws.cell(r, 4).value
            m = ws.cell(r, 13).value
            if d and isinstance(m, (int, float)):
                ws.cell(r, 3).fill = py_copy.copy(ref_fill_c) if ref_fill_c else py_copy.copy(no_fill)
                ws.cell(r, 4).fill = py_copy.copy(ref_fill_d) if ref_fill_d else py_copy.copy(no_fill)
                fixed += 1
        if fixed:
            self.log.info("   研发费用明细 C/D 列填充修正 %d 行", fixed)

    def _clear_yanfa_section_s_fill(self, ws):
        r_yanfa = self._find_row(ws, '八、研发费用')
        r_nine = self._find_row(ws, '九、成本及费用合计') or self._find_row(ws, '八、成本及费用合计')
        if not r_yanfa or not r_nine:
            return
        no_fill = PatternFill(fill_type=None)
        fixed = 0
        for r in range(r_yanfa + 1, r_nine):
            ws.cell(r, 19).fill = py_copy.copy(no_fill)
            fixed += 1
        if fixed:
            self.log.info("   研发费用区间 S 列填充清空 %d 行", fixed)

    def _unify_summary_rows(self, ws, max_col=38):
        r_base = self._find_row(ws, '三、成本合计')
        if not r_base: return
        KEYWORDS = ["四、计提", "五、过程", "六、资金", "七、局投资", "八、研发"]
        for r in range(r_base + 1, ws.max_row + 1):
            val = self._norm(ws.cell(r, 4).value or "")
            if any(k in val for k in KEYWORDS):
                for c in range(1, max_col + 1):
                    src = ws.cell(r_base, c)
                    tgt = ws.cell(r, c)
                    try:
                        if src.font:      tgt.font      = py_copy.copy(src.font)
                        if src.fill:      tgt.fill      = py_copy.copy(src.fill)
                        if src.alignment: tgt.alignment = py_copy.copy(src.alignment)
                    except Exception:
                        pass

    def _apply_borders(self, ws, max_col=38):
        thin = Side(border_style="thin", color="000000")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for r in range(7, ws.max_row + 1):
            if not ws.cell(r, 4).value and not ws.cell(r, 13).value:
                continue
            for c in range(1, max_col + 1):
                ws.cell(r, c).border = border

    def _load_workbook(self):
        wb = openpyxl.load_workbook(self.input_path)
        ws = wb.active
        for sn in wb.sheetnames:
            if any(k in sn for k in ["效益审核", "附表"]):
                ws = wb[sn]
                break
        return wb, ws

    # ====================================================================
    #  执行入口
    # ====================================================================
    def _run_structure_repairs(self, ws):
        self._delete_trailing_empty_rows(ws)
        self._delete_yanfa_ghost_rows(ws)
        self._delete_cost_zero_m_rows(ws)

    def _run_formula_repairs(self, ws):
        self._fix_cost_section_sums(ws)       # 强制重写1-7节 SUM + O/P/Q/R
        self._fix_extended_section_sums(ws)   # 单独修复 S 列及之后的节内 SUM
        self._fix_yanfa_sum(ws)               # 强制重写研发 SUM
        self._fix_anchor_rows(ws)             # 强制重写锚点行 O/P/Q/R
        self._fix_fixed_amount_rows(ws)       # 强制重写固定金额行 O/P/Q/R
        self._fix_cost_total_formula(ws)      # 强制重写三、成本合计
        self._rewrite_detail_row_formulas(ws) # 强制重写所有明细行 O/P/Q/R
        self._normalize_row_arithmetic_formulas(ws) # 清理删行后残留的旧行号
        self._fix_nine_formula(ws)            # 强制重写九合计
        self._fix_profit_formulas(ws)         # 强制重写十二利润/十三利润率
        self._clean_circular_refs(ws)         # 兜底纯自引用
        self._fill_sequence_numbers(ws)       # A列序号

    def _run_style_repairs(self, ws):
        self._fix_yanfa_detail_fill(ws)
        self._clear_yanfa_section_s_fill(ws)
        self._unify_summary_rows(ws)
        self._sync_l_fill_to_j(ws)
        self._apply_borders(ws)

    def _save_workbook(self, wb):
        try:
            wb.save(self.output_path)
            self.log.info("交付版生成完成：%s", self.output_path)
        except PermissionError:
            alt_path = self.output_path.replace(".xlsx", "_修复版.xlsx")
            wb.save(alt_path)
            self.output_path = alt_path
            self.log.warning("原交付版文件被占用，已另存为：%s", alt_path)

    def execute_beautify(self):
        if not os.path.exists(self.input_path):
            self.log.error("找不到输入文件：%s", self.input_path)
            return False
        self.log.info("美化引擎 v3.3 启动（强制重写公式，彻底防循环引用）...")
        try:
            wb, ws = self._load_workbook()
            self._run_structure_repairs(ws)
            self._run_formula_repairs(ws)
            self._run_style_repairs(ws)
            self._save_workbook(wb)
            return True
        except Exception as e:
            self.log.error("美化异常:\n%s", traceback.format_exc())
            return False
