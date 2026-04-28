# -*- coding: utf-8 -*-
import os
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from project_config import ZGD_PREFIX, RESEARCH_OFFSET_MAP

class IntermediateTableBuilder:
    """
    独立的数据清洗与 ETL 模块：
    专职负责将基础大表数据，按照复杂的业务规则，提取并透视为【5_中间汇总表】
    """

    def __init__(self, result_dir, logger):
        self.result_dir = result_dir
        self.log = logger

    def generate(self, df):
        self.log.info("📊 正在提取并生成项目效益审核【中间汇总表】...")
        try:
            def _num(v) -> float:
                n = pd.to_numeric(v, errors='coerce')
                return 0.0 if pd.isna(n) else float(n)

            def _norm_text(v) -> str:
                s = str(v or '').strip()
                return '' if s in ('nan', 'None') else s

            def _path_text(v) -> str:
                return _norm_text(v).replace('/', '\\')

            def _is_blank_text(v) -> bool:
                return _norm_text(v) == ''

            def _has_internal_counterparty(path_text: str) -> bool:
                return ('内部存款' in path_text) or ('内部往来' in path_text)

            def _has_cfk_counterparty(path_text: str) -> bool:
                return ('其他应收款\\内部存款\\可用' in path_text or
                        ('其他应收款' in path_text and '内部存款' in path_text and '可用' in path_text))

            def _is_main_or_aux_material(path_text: str) -> bool:
                return ('原材料\\主材' in path_text or
                        '原材料\\辅材' in path_text or
                        '原材料\\材主' in path_text)

            def _build_negative_material_offset(row_dict: dict, amount: float, subcat: str) -> dict:
                offset_row = dict(row_dict)
                offset_row['成本_财务大类'] = '(三)材料费'
                offset_row['细分科目'] = subcat
                offset_row['客商名称'] = subcat
                offset_row['合同编码'] = ''
                offset_row['最终发生额'] = -abs(amount)
                return offset_row

            df_calc = df.copy()
            df_calc['成本_财务大类'] = None
            df_calc['细分科目'] = None
            df_calc['最终发生额'] = 0.0

            # ========================================================
            # 【前置动作】构建雷达字典
            # ========================================================
            cost_mask = df_calc['总账科目长文本'].astype(str).str.contains(r'合同履约成本\\工程施工成本', na=False)
            cost_df = df_calc[cost_mask].copy()
            cost_df['末级科目'] = cost_df['总账科目长文本'].astype(str).apply(
                lambda x: x.split('\\')[-1] if '\\' in x else x)
            valid_cost_df = cost_df[cost_df['文本'].notna() & (cost_df['文本'].astype(str).str.strip() != '')]
            text_to_cost_subcat = valid_cost_df.drop_duplicates('文本').set_index('文本')['末级科目'].to_dict()

            if '中台单据号' in cost_df.columns:
                valid_doc_df = cost_df[
                    cost_df['中台单据号'].notna() & (cost_df['中台单据号'].astype(str).str.strip() != '')
                ].copy()
                valid_doc_df['中台单据号'] = valid_doc_df['中台单据号'].astype(str).str.strip().str.upper()
                doc_to_cost_subcat = valid_doc_df.drop_duplicates('中台单据号').set_index('中台单据号')['末级科目'].to_dict()
            else:
                doc_to_cost_subcat = {}

            # ========================================================
            # 计算材料库存：原材料借方+贷方汇总，取负数
            # ========================================================
            mat_mask = df_calc['总账科目长文本'].astype(str).str.contains('原材料', na=False)
            mat_debit_sum = pd.to_numeric(df_calc.loc[mat_mask, '借方本位币金额'], errors='coerce').fillna(0).sum()
            mat_credit_sum = pd.to_numeric(df_calc.loc[mat_mask, '贷方本位币金额'], errors='coerce').fillna(0).sum()
            inventory_amt = - (mat_debit_sum + mat_credit_sum)

            # 收集研发费用冲销记录
            offset_records = []

            # --- 核心逐行映射规则引擎 ---
            for idx, row in zip(df_calc.index, df_calc.to_dict('records')):
                gl_text = _norm_text(row.get('总账科目长文本', ''))
                gl_text_path = _path_text(gl_text)
                gl_code = str(row.get('科目号', row.get('总账科目', '')))
                opp_gl_text = _norm_text(row.get('对方科目描述', ''))
                opp_gl_text_path = _path_text(opp_gl_text)
                text_desc = _norm_text(row.get('文本', ''))

                doc_no = _norm_text(row.get('中台单据号', '')).upper()
                contract = _norm_text(row.get('合同编码', row.get('合同', '')))

                debit = _num(row.get('借方本位币金额', 0))
                credit = _num(row.get('贷方本位币金额', 0))
                balance = _num(row.get('余额（本币）', 0))

                cat, sub_cat, amt = None, None, 0.0

                # 1. 收入与税金
                if '主营业务收入' in gl_text_path:
                    cat, amt = '财务累计入账收入(不含增值税)', credit
                elif ('销项税额' in gl_text_path) and ('增值税' in gl_text_path) and ('应交税费' in gl_text_path):
                    cat, amt = '已价税分离的增值税金额(财务账面数据)', credit

                # 2. 成本映射
                elif '直接人工费' in gl_text_path:
                    cat, amt = '(一)人工费', debit
                elif '分包工程支出' in gl_text_path:
                    cat, amt = '(二)分包工程', debit

                # ====================================================
                # 材料费规则重写（覆盖所有情况）
                # ====================================================
                elif '原材料' in gl_text_path:
                    # 1. 普通材料费：原材料/xx，合同不为空且单据为 JSD 起始
                    if contract != '' and doc_no.startswith('JSD'):
                        cat, amt = '(三)材料费', debit
                    # 2. 材料调入：原材料/xx，对方科目为内部存款/内部往来，单据为 DBD
                    elif doc_no.startswith('DBD') and _has_internal_counterparty(opp_gl_text_path):
                        cat, sub_cat, amt = '(三)材料费', '材料调入', debit
                    # 8/9. 原材料主材/辅材/材主，按单据区分 XSD、SQD
                    elif _is_main_or_aux_material(gl_text_path) and doc_no.startswith('XSD'):
                        cat, sub_cat, amt = '(三)材料费', '材料调出+XSD', -credit  # 贷方金额冲减，取负数
                    elif _is_main_or_aux_material(gl_text_path) and doc_no.startswith('SQD'):
                        cat, sub_cat, amt = '(三)材料费', '材料调出+SQD', -credit  # 贷方金额冲减，取负数

                elif '合同履约成本\\工程施工成本\\直接材料费' in gl_text_path:
                    # 5/8. 材料调入+CFK：对方科目为 其他应收款\内部存款\可用，且单据为 CFK
                    if doc_no.startswith('CFK') and _has_cfk_counterparty(opp_gl_text_path):
                        cat, sub_cat, amt = '(三)材料费', '材料调入+CFK', debit
                    # 3. 材料调出：直接材料费，对方科目为内部存款/内部往来，单据为 DBD
                    elif doc_no.startswith('DBD') and _has_internal_counterparty(opp_gl_text_path):
                        cat, sub_cat, amt = '(三)材料费', '材料调出', debit
                    # 6. 废旧物资消耗：对方科目为空，根据 CZD 或 SQD 区分
                    elif doc_no.startswith('CZD') and _is_blank_text(opp_gl_text):
                        cat, sub_cat, amt = '(三)材料费', '废旧物资消耗', debit
                    elif doc_no.startswith('SQD') and _is_blank_text(opp_gl_text):
                        cat, sub_cat, amt = '(三)材料费', '废旧物资消耗（SQD）', debit
                    # 7. 材料费-其他：直接材料费，对方科目为内部存款/内部往来，且单据不为 DBD/CFK
                    elif (not doc_no.startswith('DBD')) and (not doc_no.startswith('CFK')) and _has_internal_counterparty(opp_gl_text_path):
                        cat, sub_cat, amt = '(三)材料费', '其他', debit
                    # 兜底兼容（JSD可能落在履约成本的情形）
                    elif contract != '' and doc_no.startswith('JSD'):
                        cat, amt = '(三)材料费', debit

                # 机械费
                elif '机械使用费' in gl_text_path:
                    cat, amt = '(四)机械租赁费', debit
                # 其他直接费
                elif '其他直接费用' in gl_text_path:
                    cat = '(五)其他直接费'
                    sub_cat = gl_text_path.split('\\')[-1] if '\\' in gl_text_path else gl_text_path
                    amt = debit
                # 间接费
                elif '间接费用' in gl_text_path:
                    cat = '(六)间接费'
                    sub_cat = gl_text_path.split('\\')[-1] if '\\' in gl_text_path else gl_text_path
                    amt = debit

                # 3. 投资收益
                elif '投资收益' in gl_text_path or '以摊余成本计量的金融资产终止确认收益' in gl_text_path:
                    cat, amt = '七、局投资收益（局投资项目选填）', credit

                # 4. 安全生产费
                elif ('专项储备\\安全生产费\\发生数' in gl_text_path) or (
                        '专项储备' in gl_text_path and '安全生产费' in gl_text_path):
                    cat = '(七)安全费'

                    def _safety_bucket(leaf: str) -> str:
                        s = str(leaf or '').replace(' ', '')
                        if '直接人工' in s: return 'labor'
                        if '分包' in s: return 'subcon'
                        if '机械' in s: return 'machine'
                        if '材料' in s: return 'material'
                        return ''

                    SAFETY_BUCKET_NAME_MAP = {
                        'labor': '安全生产-人工费',
                        'subcon': '安全生产-分包工程',
                        'machine': '安全生产-机械使用',
                        'material': '安全生产-材料费',
                    }

                    if '原材料\\' in opp_gl_text_path:
                        sub_cat = '安全生产-材料费'
                    elif '内部存款' in opp_gl_text_path or '可用存款' in opp_gl_text_path:
                        sub_cat = '安全生产费-报销'
                    elif opp_gl_text.strip() in ['', 'nan', 'None']:
                        matched_subcat = doc_to_cost_subcat.get(doc_no) or text_to_cost_subcat.get(text_desc)
                        bucket = _safety_bucket(matched_subcat)
                        if bucket and bucket in SAFETY_BUCKET_NAME_MAP:
                            sub_cat = SAFETY_BUCKET_NAME_MAP[bucket]
                        elif matched_subcat:
                            sub_cat = f'安全生产-{matched_subcat}'
                        else:
                            sub_cat = '劳保用品费' if '劳保' in text_desc else '安全生产费-其他'
                    else:
                        if '合同履约成本' in opp_gl_text_path and '工程施工成本' in opp_gl_text_path:
                            opp_leaf = opp_gl_text_path.split('\\')[-1] if '\\' in opp_gl_text_path else opp_gl_text_path.split(
                                '/')[-1]
                            bucket = _safety_bucket(opp_leaf)
                            if bucket and bucket in SAFETY_BUCKET_NAME_MAP:
                                sub_cat = SAFETY_BUCKET_NAME_MAP[bucket]
                            else:
                                sub_cat = f'安全生产-{opp_leaf}'
                        else:
                            sub_cat = '安全生产费-其他'
                    amt = debit

                    if sub_cat == '安全生产-材料费' and amt != 0:
                        offset_records.append(_build_negative_material_offset(row, amt, sub_cat))

                # 5. 研发费用
                elif ('研发支出' in gl_text_path) or ('研发费用' in gl_text_path):
                    if '研发' in opp_gl_text_path:
                        continue
                    cat = '研发支出'
                    amt = debit
                    offset_cat, offset_subcat = None, None

                    if '材料费' in gl_text_path:
                        sub_cat = '研发费用-材料费'
                        offset_cat, offset_subcat = RESEARCH_OFFSET_MAP.get('材料费', (None, None))
                    elif '人工成本' in gl_text_path:
                        sub_cat = '研发费用-人工成本'
                    elif '研发费用\\租赁及运行费\\设备租赁及运行维护' in gl_text_path:
                        sub_cat = '研发费用-设备租赁及运行维护'
                    elif '租赁及运行维护费' in gl_text_path:
                        sub_cat = '研发费用-机械租赁'
                    else:
                        if '折旧' in gl_text_path:
                            sub_cat = '研发费用-折旧'
                        elif '内部存款' in opp_gl_text_path or '内部往来' in opp_gl_text_path:
                            sub_cat = '研发费用-内部结转'
                        else:
                            sub_cat = '研发费用-其他'

                    if offset_cat and offset_subcat and amt != 0:
                        offset_row = dict(row)
                        offset_row['成本_财务大类'] = offset_cat
                        offset_row['细分科目'] = offset_subcat
                        offset_row['客商名称'] = offset_subcat
                        offset_row['合同编码'] = ''
                        offset_row['最终发生额'] = -amt
                        offset_records.append(offset_row)

                # 6. 资金占用与税金
                elif gl_code.startswith('6603'):
                    cat, amt = '六、资金占用费用', balance
                elif gl_code.startswith('6403'):
                    cat, amt = '十、税金及附加（按财务数据）', debit

                # 7. 待确认进项
                elif '其他应收款\\待确认进项税额' in gl_text_path:
                    cat, amt = '其他应收-待确认进项税额', balance

                # --- ZGD 单据统一追加前缀 ---
                if doc_no.upper().startswith(ZGD_PREFIX):
                    if cat is not None:   # 仅在已分类时处理
                        original_subcat = sub_cat if sub_cat else '未分类'
                        sub_cat = f"{ZGD_PREFIX}-{original_subcat}"
                        df_calc.at[idx, '客商名称'] = 'ZGD'
                        df_calc.at[idx, '合同编码'] = ''
                        amt = debit   # 统一使用借方

                # 回填计算结果
                df_calc.at[idx, '成本_财务大类'] = cat
                df_calc.at[idx, '细分科目'] = sub_cat
                df_calc.at[idx, '最终发生额'] = amt

            # ---------------- 过滤无效与合并多维数据 ----------------
            df_valid = df_calc[df_calc['成本_财务大类'].notna() & (df_calc['最终发生额'] != 0)].copy()
            if offset_records:
                df_valid = pd.concat([df_valid, pd.DataFrame(offset_records)], ignore_index=True)

            if df_valid.empty:
                self.log.warning("⚠️ 按照既定规则，未提取到任何满足条件的效益审核数据。")
                return False

            # 材料库存追加
            if inventory_amt != 0:
                inv_row = {
                    '利润中心': df_valid['利润中心'].iloc[0] if not df_valid.empty else '',
                    '利润中心文本描述': df_valid['利润中心文本描述'].iloc[0] if '利润中心文本描述' in df_valid.columns else '',
                    'WBS元素': df_valid['WBS元素'].iloc[0] if 'WBS元素' in df_valid.columns else '',
                    '客商名称': '无客商-材料库存',
                    '合同编码': '',
                    '中台单据号': '',
                    '总账科目长文本': '原材料',
                    '成本_财务大类': '(三)材料费',
                    '细分科目': '材料库存',
                    '最终发生额': inventory_amt
                }
                df_valid = pd.concat([df_valid, pd.DataFrame([inv_row])], ignore_index=True)

            # 基础字段整理
            def _clean_name_series(series):
                if series is None:
                    return pd.Series(pd.NA, index=df_valid.index, dtype='object')
                if not isinstance(series, pd.Series):
                    series = pd.Series(series, index=df_valid.index)
                cleaned = series.copy()
                cleaned = cleaned.replace(r'^\s*$', pd.NA, regex=True)
                cleaned = cleaned.replace(['nan', 'None'], pd.NA)
                return cleaned

            existing_name = (_clean_name_series(df_valid['客商名称'])
                             if '客商名称' in df_valid.columns
                             else pd.Series(pd.NA, index=df_valid.index, dtype='object'))
            vendor = (_clean_name_series(df_valid['供应商名称'])
                      if '供应商名称' in df_valid.columns
                      else _clean_name_series(df_valid['供应商'])
                      if '供应商' in df_valid.columns
                      else pd.Series(pd.NA, index=df_valid.index, dtype='object'))
            customer = (_clean_name_series(df_valid['客户描述'])
                        if '客户描述' in df_valid.columns
                        else _clean_name_series(df_valid['客户'])
                        if '客户' in df_valid.columns
                        else pd.Series(pd.NA, index=df_valid.index, dtype='object'))

            df_valid['客商名称'] = existing_name.combine_first(vendor).combine_first(customer)

            doc_series = (
                df_valid['中台单据号']
                if '中台单据号' in df_valid.columns
                else pd.Series('', index=df_valid.index, dtype='object')
            )
            doc_prefix = doc_series.astype(str).str.strip().replace(['nan', 'None'], '').str[:3].str.upper()
            df_valid['客商名称'] = df_valid['客商名称'].fillna('').astype(str).str.strip().replace(['nan', 'None'], '')
            df_valid['客商名称'] = df_valid['客商名称'].where(df_valid['客商名称'] != '', doc_prefix)
            df_valid['工程名称'] = df_valid.get('利润中心文本描述', '').astype(str).replace('nan', '')
            df_valid['项目编码'] = df_valid.get('WBS元素', '').astype(str).replace('nan', '')
            df_valid['合同编码'] = df_valid.get('合同', df_valid.get('合同编码', '')).astype(str).replace('nan', '')
            df_valid['利润中心'] = df_valid.get('利润中心', '').astype(str).replace('nan', '')
            df_valid['中台单据号'] = df_valid.get('中台单据号', '').astype(str).replace('nan', '')
            df_valid['总账科目长文本'] = df_valid.get('总账科目长文本', '').astype(str).replace('nan', '')

            # 间接费/其他直接费合并
            merge_mask = df_valid['成本_财务大类'].isin(['(五)其他直接费', '(六)间接费']) & \
                         (df_valid['合同编码'] == '') & \
                         (~df_valid['中台单据号'].str.upper().str.startswith('JSD', na=False))
            df_valid.loc[merge_mask, '中台单据号'] = ''
            df_valid.loc[merge_mask, '客商名称'] = '无客商-零星汇总'
            df_valid.loc[merge_mask, '总账科目长文本'] = df_valid.loc[merge_mask, '细分科目']

            group_cols = [
                '利润中心', '工程名称', '项目编码', '客商名称', '合同编码',
                '中台单据号', '总账科目长文本', '成本_财务大类', '细分科目'
            ]

            agg_df = df_valid.groupby(group_cols, dropna=False)['最终发生额'].sum().reset_index()
            agg_df = agg_df[agg_df['最终发生额'] != 0].round(2)

            sort_order = [
                '(一)人工费', '(二)分包工程', '(三)材料费', '(四)机械租赁费',
                '(五)其他直接费', '(六)间接费', '(七)安全费', '研发支出',
                '六、资金占用费用', '七、局投资收益（局投资项目选填）',
                '十、税金及附加（按财务数据）', '已价税分离的增值税金额(财务账面数据)',
                '财务累计入账收入(不含增值税)'
            ]
            agg_df['成本_财务大类'] = pd.Categorical(agg_df['成本_财务大类'], categories=sort_order, ordered=True)
            agg_df = agg_df.sort_values(by=['成本_财务大类', '细分科目', '客商名称', '合同编码']).reset_index(drop=True)

            out_path = os.path.join(self.result_dir, "5_中间汇总表_效益审核数据源.xlsx")
            agg_df.to_excel(out_path, index=False)
            self._add_subtotal(out_path, agg_df)

            self.log.info(f"🎉 成功提取底层逻辑！已按指定规则排序并生成: {out_path}")
            return True

        except Exception as e:
            self.log.error(f"❌ 生成中间表时失败: {e}")
            import traceback
            self.log.error(traceback.format_exc())
            return False

    def _add_subtotal(self, path, df):
        try:
            wb = openpyxl.load_workbook(path)
            ws = wb.active
            ws.insert_rows(2)
            for i, col in enumerate(df.columns, 1):
                is_amount_col = pd.api.types.is_numeric_dtype(df[col]) or any(
                    k in str(col) for k in ['金额', '余额', '发生额', '差额'])
                if is_amount_col:
                    col_let = get_column_letter(i)
                    ws.cell(row=2, column=i, value=f"=SUBTOTAL(9,{col_let}3:{col_let}{ws.max_row})")
                elif i == 1:
                    ws.cell(row=2, column=i, value="筛选合计：")
            ws.freeze_panes = "A3"
            wb.save(path)
            self.log.info(f"🧮 成功为【5_中间汇总表】添加 SUBTOTAL 筛选动态合计行！")
        except Exception as e:
            self.log.error(f"中间表合计行添加失败: {e}")
